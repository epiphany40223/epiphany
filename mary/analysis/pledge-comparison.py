#!/usr/bin/env python3

# Compare ECC Python pledge numbers to the "Pledge Drive Status Report.xlsx" report from PDS.

import os
import sys
import csv
import copy
import datetime
import openpyxl

# Load the ECC python modules.  There will be a sym link off this directory.
moddir = os.path.join(os.path.dirname(sys.argv[0]), 'ecc-python-modules')
if not os.path.exists(moddir):
    print("ERROR: Could not find the ecc-python-modules directory.")
    print("ERROR: Please make a ecc-python-modules sym link and run again.")
    exit(1)

sys.path.insert(0, moddir)

import ECC
import PDSChurch

pledge_minimum = 0

##############################################################################

def find_family_funding(year, family):
    data = {
        'year'    : year,
        'fid'     : family['FamRecNum'],
        'found'   : False,
        'pledged' : 0,
        'q1'      : 0,
        'q2'      : 0,
        'q3'      : 0,
        'q4'      : 0,
        'total'   : 0,
    }

    key      = 'funds'
    pds_year = f"{year-2000:02d}"
    if key not in family or pds_year not in family[key]:
        return data

    funds = family[key][pds_year]
    for fund in funds.keys():
        # Only look at fund 1, which is general stewardship
        if fund != "1":
            continue

        fund_rate = funds[fund]['fund_rate']
        if fund_rate and fund_rate['FDTotal']:
            data['pledged'] += int(fund_rate['FDTotal'])
            data['found'] = True

        for item in funds[fund]['history']:
            amount = item['item']['FEAmt']
            if amount is None:
                # Yes, this happens.  Sigh.
                continue

            quarter = int((item['item']['FEDate'].month-1) / 3) + 1
            data[f'q{quarter}'] += amount
            data['total'] += amount
            data['found'] = True

    return data

##############################################################################

# Hack, stolen from https://stackoverflow.com/questions/32464280/converting-currency-with-to-numbers-in-python-pandas
import locale
locale.setlocale(locale.LC_ALL,'')

def load_xlsx(filename):
    wb = openpyxl.load_workbook(filename)
    # Assume there's 1 sheet
    name = wb.sheetnames[0]
    ws = wb[name]

    # First data is on row 6
    # Last data is on row 1405
    # Data is in pairs of rows
    results = dict()
    for row in range(6, 1406, 2):
        env = ws.cell(row, 1).value
        family_name = ws.cell(row, 2).value
        pledge = ws.cell(row+1, 6).value

        # Hack, stolen from https://stackoverflow.com/questions/32464280/converting-currency-with-to-numbers-in-python-pandas
        pledge = int(locale.atof(pledge.strip('$')))

        print(f"{env} : {family_name} : {pledge}")
        results[env] = {
            'env'    : env,
            'name'   : family_name,
            'pledge' : pledge,
        }

    return results

##############################################################################

def compare(xlsx, families):
    families_by_env = dict()
    for fam in families.values():
        families_by_env[fam['ParKey'].strip()] = fam
        fam['compare'] = find_family_funding(2021, fam)

    for env in xlsx:
        xlsx_pledge = xlsx[env]['pledge']

        family = families_by_env[env]
        pds_pledge = int(family['compare']['pledged'])

        # Special case
        if pds_pledge == 0 and xlsx_pledge == 1:
            # Skip this
            continue

        if xlsx_pledge != pds_pledge:
            print(f"{env} (fid: {family['FamRecNum']}): ERROR: {xlsx[env]['name']} XLSX pledge: {xlsx_pledge}, PDS pledge {pds_pledge}")

##############################################################################

def main():
    log = ECC.setup_logging(debug=False)

    (pds, families,
     members) = PDSChurch.load_families_and_members(filename='pdschurch.sqlite3',
                                                    active_only=False,
                                                    parishioners_only=False,
                                                    log=log)

    print(f"Loaded {len(members)} total Members")
    print(f"Loaded {len(families)} total Families")

    squyres = 119353
    hall = 549102
    for year in range(2015, 2020+1):
        data = find_family_funding(year, families[hall])
        print(f"Hall {year}: {data}")

    xlsx = load_xlsx("Pledge Drive Status Report.xlsx")
    compare(xlsx, families)

main()
