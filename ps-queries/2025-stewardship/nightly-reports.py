#!/usr/bin/env python3

# Make sure to pip install everything in requirements.txt.

import sys

import collections
import traceback
import datetime
import argparse
import csv
import os
import re

from functools import cmp_to_key
from collections import defaultdict

from pprint import pformat

# We assume that there is a "ecc-python-modules" sym link in this
# directory that points to the directory with ECC.py and friends.
moddir = os.path.join(os.getcwd(), 'ecc-python-modules')
if not os.path.exists(moddir):
    print("ERROR: Could not find the ecc-python-modules directory.")
    print("ERROR: Please make a ecc-python-modules sym link and run again.")
    exit(1)
# On MS Windows, git checks out sym links as a file with a single-line
# string containing the name of the file that the sym link points to.
if os.path.isfile(moddir):
    with open(moddir) as fp:
        dir = fp.readlines()
    moddir = os.path.join(os.getcwd(), dir[0])

sys.path.insert(0, moddir)

import ECC
import ECCEmailer
import Google
import ParishSoftv2 as ParishSoft
import GoogleAuth

import helpers

from datetime import date
from datetime import datetime
from datetime import timedelta

from oauth2client import tools
from googleapiclient.http import MediaFileUpload
from email.message import EmailMessage

import matplotlib
import matplotlib.dates as mdates
import matplotlib.pyplot as plt

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment

#------------------------------------------------------------------------------

from constants import stewardship_fam_prev_year_wg

from constants import stewardship_begin_date
from constants import stewardship_end_date

from constants import gapp_id
from constants import guser_cred_file

from constants import jotform_gsheet_gfile_id
from constants import jotform_gsheet_columns
from constants import upload_team_drive_folder_id
from constants import gsheet_editors

from constants import stewardship_year
from constants import title

from constants import smtp_server
from constants import smtp_from

from constants import jotform

from constants import MAX_PS_FAMILY_MEMBER_NUM

from constants import jotform_gsheet_prev_year_gfile_id
from constants import jotform_gsheet_prev_year_columns

##############################################################################

ecc = '@epiphanycatholicchurch.org'

# Comments and pledge analysis report email
reports_email_to = f'angie{ecc},mary{ecc},jeff@squyres.com'
reports_email_subject = 'Comments and pledge analysis reports'

# Statistics report email
statistics_email_to = f'sdreiss71@gmail.com,angie{ecc},mary{ecc},jeff@squyres.com'
statistics_email_subject = 'Statistics report'

family_participation_email_to = f'mary{ecc},jeff@squyres.com'

pledge_email_to = family_participation_email_to
pledge_email_subject = 'Pledge PS CSV import file'

# Use this field name in the PS member dictionary to store some
# jotform results
member_extras_key = 'jotform extras'

# JMS for debugging/testing
statistics_email_to = 'jeff@squyres.com'
reports_email_to = statistics_email_to
family_participation_email_to = statistics_email_to
pledge_email_to = statistics_email_to

##############################################################################

def _upload_to_gsheet(google, google_folder_id, google_filename, mime_type, local_filename, remove_local, log):
    try:
        log.info(f'Uploading file to google "{local_filename}" -> "{google_filename}"')
        metadata = {
            'name'     : google_filename,
            'mimeType' : Google.mime_types['sheet'],
            'parents'  : [ google_folder_id ],
            'supportsTeamDrives' : True,
            }
        media = MediaFileUpload(local_filename,
                                mimetype=Google.mime_types[mime_type],
                                resumable=True)
        file = google.files().create(body=metadata,
                                     media_body=media,
                                     supportsTeamDrives=True,
                                     fields='id').execute()
        log.debug(f'Successfully uploaded file: "{google_filename}" (ID: {file["id"]})')

    except:
        log.error('Google upload failed for some reason:')
        log.error(traceback.format_exc())
        exit(1)

    # Set permissions on the GSheet to allow the
    # workers group to edit the file (if you are view-only, you
    # can't even adjust the column widths, which will be
    # problematic for the comments report!).
    # JMS Fix me
    if False:
        try:
            perm = {
                'type': 'group',
                'role': 'writer',
                'emailAddress': gsheet_editors,
            }
            out = google.permissions().create(fileId=file['id'],
                                              supportsTeamDrives=True,
                                              sendNotificationEmail=False,
                                              body=perm,
                                              fields='id').execute()
            log.debug(f"Set Google permission for file: {id}")
        except:
            log.error('Google set permission failed for some reason:')
            log.error(traceback.format_exc())
            exit(1)

    # Remove the temp file when we're done
    if remove_local:
        try:
            os.remove(local_filename)
        except:
            pass

    return file['id']

#-----------------------------------------------------------------------------

def _make_filenames(filename, extension):
    extension = f".{extension}"
    if filename.endswith(extension):
        local_filename  = filename
        google_filename = filename[:-len(extension)]
    else:
        local_filename  = f"{filename}{extension}"
        google_fielname = filename

    return local_filename, google_filename

#-----------------------------------------------------------------------------

def upload_csv_to_gsheet(google, google_folder_id, filename, fieldnames, csv_rows, remove_local, log):
    if csv_rows is None or len(csv_rows) == 0:
        return None, None

    # First, write out a CSV file
    csv_filename, google_filename = _make_filenames(filename, 'csv')
    try:
        os.remove(csv_filename)
    except:
        pass

    csvfile = open(csv_filename, 'w')
    writer  = csv.DictWriter(csvfile, fieldnames=fieldnames)
    writer.writeheader()
    for row in csv_rows:
        writer.writerow(row)
    csvfile.close()

    # Now upload that file to Google Drive
    id = _upload_to_gsheet(google,
                        google_folder_id=google_folder_id,
                        google_filename=google_filename,
                        mime_type='csv',
                        local_filename=csv_filename,
                        remove_local=remove_local,
                        log=log)

    return id, None if remove_local else csv_filename

#-----------------------------------------------------------------------------

def upload_xlsx_to_gsheet(google, google_folder_id, filename, workbook, remove_local, log):
    # First, write out the XLSX file
    xlsx_filename, google_filename = _make_filenames(filename, 'xlsx')
    try:
        os.remove(xlsx_filename)
    except:
        pass
    workbook.save(xlsx_filename)

    # Now upload that file to Google Drive
    id = _upload_to_gsheet(google,
                        google_folder_id=google_folder_id,
                        # For some reason, when we upload an XLSX and want
                        # it to convert to a Google Sheet, the google filename
                        # must end in .xlsx.  Sigh.
                        google_filename=xlsx_filename,
                        mime_type='sheet',
                        local_filename=xlsx_filename,
                        remove_local=remove_local,
                        log=log)

    return id, None if remove_local else xlsx_filename

##############################################################################

def _change(label, old_value, new_value, message):
    return {
        'label'     : label,
        'old_value' : old_value,
        'new_value' : new_value,
        'message'   : message,
    }

def _compare(changes, label, jot_value, ps_value):
    if jot_value is None:
        jot_value = ''
    if ps_value is None:
        ps_value = ''

    if jot_value.strip() == ps_value.strip():
        return

    message = ('{label}: {new_value}'
               .format(label=label, new_value=jot_value))
    changes.append(_change(label=label,
                           old_value=ps_value,
                           new_value=jot_value,
                           message=message))

##############################################################################

def comments_to_xlsx(google, jotform_data, id_field, emails_field, name_field,
                     workbook, log):
    sheet = workbook.active

    comments_label    = "Comments"
    pledge_last_label = f'CY{stewardship_year-1} pledge'
    pledge_cur_label  = f'CY{stewardship_year} whole year pledge'

    # Setup the title row
    # Title rows + set column widths
    title_font   = Font(color='FFFF00')
    title_fill   = PatternFill(fgColor='0000FF', fill_type='solid')
    title_align  = Alignment(horizontal='center', wrap_text=True)

    wrap_align   = Alignment(horizontal='general', wrap_text=True)
    right_align  = Alignment(horizontal='right')

    money_format = "$###,###,###"

    xlsx_cols = dict();
    def _add_col(name, width=10):
        col             = len(xlsx_cols) + 1
        xlsx_cols[name] = {'name' : name, 'column' : col, 'width' : width }
    _add_col('Date', width=20)
    _add_col('Family DUID')
    _add_col('Family names', width=30)
    _add_col('Emails', width=50)
    _add_col(pledge_last_label)
    _add_col(pledge_cur_label)
    _add_col('Comments', width=100)

    for data in xlsx_cols.values():
        col            = data['column']
        cell           = sheet.cell(row=1, column=col, value=data['name'])
        cell.fill      = title_fill
        cell.font      = title_font
        cell.alignment = title_align

        col_char = chr(ord('A') - 1 + col)
        sheet.column_dimensions[col_char].width = data['width']

    sheet.freeze_panes = sheet['A2']

    #-------------------------------------------------------------------

    def _extract_money_string(val):
        if val.startswith('$'):
            val = val[1:]
        val = int(float(val.replace(',', '').strip()))

        if val != 0:
            return int(val), money_format
        else:
            return 0, None

    #-------------------------------------------------------------------

    # Now fill in all the data rows
    xlsx_row = 2
    log.info(f"Checking {len(jotform_data)} rows for comments")
    for row in jotform_data:
        # Skip if the comments are empty
        if comments_label not in row:
            continue
        if row[comments_label] == '':
            continue

        if row[pledge_last_label]:
            pledge_last, pledge_last_format = _extract_money_string(row[pledge_last_label])
        else:
            pledge_last = 0
            pledge_last_format = None

        if row[pledge_cur_label]:
            pledge_cur = helpers.jotform_text_to_int(row[pledge_cur_label])
            pledge_cur_format = money_format
        else:
            pledge_cur = 0
            pledge_cur_format = None

        def _fill(col_name, value, align=None, format=None):
            col_data = xlsx_cols[col_name]
            cell     = sheet.cell(row=xlsx_row, column=col_data['column'], value=value)
            if align:
                cell.alignment = align
            if format:
                cell.number_format = format

        _fill('Date', row['SubmitDate'])
        _fill('Family DUID', int(row[id_field]))
        _fill('Family names', row[name_field])
        _fill('Emails', row[emails_field])
        _fill(pledge_last_label, pledge_last, format=pledge_last_format)
        _fill(pledge_cur_label, pledge_cur, format=pledge_cur_format)
        _fill('Comments', row[comments_label], align=wrap_align)

        xlsx_row += 1

    # Return the number of comments we found
    return xlsx_row - 2

def reorder_rows_by_date(jotform_data):
    data = dict()
    for row in jotform_data:
        data[row['SubmitDate']] = row

    out = list()
    for row in sorted(data):
        out.append(data[row])

    return out

def comments_report(args, google, start, end, time_period, jotform_data, log):
    log.info("Composing comments report...")

    # jotform_data is a list of rows, in order by Family DUID.  Re-order them to be
    # ordered by submission date.
    ordered_data = reorder_rows_by_date(jotform_data)

    # Examine the jotform data and see if there are any comments that
    # need to be reported
    workbook = Workbook()
    num_comments = comments_to_xlsx(google, jotform_data=ordered_data,
                     id_field='fduid', name_field='Family names',
                     emails_field='Emails to reply to',
                     workbook=workbook, log=log)

    # If we have any comments, upload them to a Gsheet
    gsheet_id = None
    sheet = workbook.active
    if num_comments == 0:
        log.info("No comments to report")
        return None

    filename   = f'Comments {time_period}.xlsx'
    gsheet_id, _ = upload_xlsx_to_gsheet(google,
                                         google_folder_id=upload_team_drive_folder_id,
                                         filename=filename,
                                         workbook=workbook,
                                         remove_local=True,
                                         log=log)

    return gsheet_id

##############################################################################

def statistics_compute(ps_families, member_workgroups,
                       unique_fduid_jotform, log):
    ret = dict()

    # Total number of active Families in the parish
    ret['num_active'] = len(ps_families)

    #---------------------------------------------------------------

    # Number of active Families who are eligible for electronic stewardship
    # (i.e., we have an email address for the spouse and/or HoH)

    eligible = dict()
    for fduid, family in ps_families.items():
        emails = \
            ParishSoft.family_business_logistics_emails(family,
                                                        member_workgroups,
                                                        log)
        if len(emails) > 0:
            eligible[fduid] = True

    ret['num_eligible'] = len(eligible)

    #-----------------------------------------------------------

    # The unique_fduid_jotform dictionary we have will have, at most, 1 entry per
    # FDUID.  So we can just take the length of it to know how many (unique)
    # families have submitted electronically.
    ret['num_electronic'] = len(unique_fduid_jotform)

    return ret

#------------------------------------------------------------------------

def statistics_graph(ps_members, ps_families, jotform, log):
    def _find_range(data, earliest, latest):
        for row in data:
            submitted = row['SubmitDate']
            submitted_dt = helpers.jotform_date_to_datetime(submitted)
            if submitted_dt < earliest:
                earliest = submitted_dt
            if submitted_dt > latest:
                latest = submitted_dt

        return earliest, latest

    #------------------------------------------------------------------------

    def _compute(start, end, ps_members, ps_families, jotform, log):
        # Compute these values just in the date range:
        # - How many unique family submissions total?

        family_submitted = dict()

        # Check electronic submissions
        for row in jotform:
            if row['fduid'] == 'Family DUID':
                continue # Skip title row

            # Is this row in our date range?
            dt = helpers.jotform_date_to_datetime(row['SubmitDate'])
            if dt < start or dt > end:
                continue

            fduid = int(row['fduid'])
            log.debug(f"Found submission in our date window: {fduid} on {dt}")

            # Make sure the family hasn't been deleted
            if fduid not in ps_families:
                continue

            family_submitted[fduid] = True

        return len(family_submitted)

    #------------------------------------------------------------------------

    one_day  = timedelta(days=1)
    earliest = datetime(year=9999, month=12, day=31)
    latest   = datetime(year=1971, month=1,  day=1)
    earliest, latest = _find_range(jotform, earliest, latest)

    earliest = datetime(year=earliest.year, month=earliest.month, day=earliest.day)
    latest   = datetime(year=latest.year, month=latest.month, day=latest.day)
    latest   += one_day - timedelta(seconds=1)

    log.info(f"Earliest: {earliest}")
    log.info(f"Latest:   {latest}")

    day             = earliest
    dates           = list()
    data_per_day    = list()
    data_cumulative = list()

    # Make lists that we can give to matplotlib for plotting
    while day <= latest:
        log.debug(f"Get per-day stats for {day.date()}")
        per_day    = _compute(day, day + one_day,
                              ps_members, ps_families,
                              jotform, log)
        log.debug(f"Get cumulative stats for {earliest} - {day + one_day}")
        cumulative = _compute(earliest, day + one_day,
                              ps_members, ps_families,
                              jotform, log)

        log.debug(f"Date: {day}: per day {per_day}, cumulative {cumulative}")

        dates.append(day.date())
        data_per_day.append(per_day)
        data_cumulative.append(cumulative)

        day += one_day

    # Make the plot
    fig, ax = plt.subplots()

    # The X label of "days" is obvious, there's no good Y label since we have
    # multiple lines
    n = datetime.now()
    hour = n.hour if n.hour > 0 else 12
    ampm = 'am' if n.hour < 12 else 'pm'
    plt.title("As of {month} {day}, {year} at {hour}:{min:02}{ampm}"
            .format(month=n.strftime("%B"), day=n.day, year=n.year,
                    hour=hour, min=n.minute, ampm=ampm))
    plt.suptitle(title + " electronic submissions statistics")

    ax.plot(dates, data_per_day, label='Unique family electronic submissions per day')
    ax.plot(dates, data_cumulative, label='Cumulative unique family electronic submissions')

    ax.get_xaxis().set_major_locator(mdates.DayLocator(interval=1))
    ax.get_xaxis().set_major_formatter(mdates.DateFormatter("%a %b %d"))
    plt.setp(ax.get_xticklabels(), rotation=20, ha="right",
             rotation_mode="anchor")
    ax.grid()

    plt.legend(loc="upper left")

    # Make a filename based on the time that this was run
    filename = ('statistics-as-of-{year:04}-{month:02}-{day:02}-{hour:02}{min:02}{sec:02}.pdf'
                .format(year=n.year, month=n.month, day=n.day,
                        hour=n.hour, min=n.minute, sec=n.second))
    fig.savefig(filename)

    plt.close(fig)

    return filename

#-----------------------------------------------------------------------------

def statistics_report(args, time_period, ps_members, ps_families,
                      member_workgroups,
                      jotform, log):
    log.info("Composing statistics report...")

    #---------------------------------------------------------------

    graph_filename = statistics_graph(ps_members, ps_families, jotform, log)
    data = statistics_compute(ps_families, member_workgroups, jotform, log)

    #---------------------------------------------------------------------

    electronic_percentage = data['num_electronic'] / data['num_eligible'] * 100

    # Send the statistics report email
    body = list()
    body.append(f"""<html>
<body>
<h2>{title} statistics update</h2>

<h3>Time period: {time_period}</h3>
<ul>
<li> Total number of active PS Families in the parish: {data['num_active']:,d}</li>
<br>
<li> Number of active PS Families eligible for electronic stewardship: {data['num_eligible']:,d}
    <ul>
    <li>This means that we have an email address in PS for the Head, Husband, and/or Wife of a given Family</li>
    </ul></li>
<br>
<li> Number of electronic-stewardship-eligible Families who have electronically submitted: {data['num_electronic']:,d} (out of {data['num_eligible']:,d}, or {electronic_percentage:.1f}%)
    <ul>
    <li>This is the number of families who submitted electronically.</li>
    </ul></li>
</ul>
</body>
</html>""")

    to = statistics_email_to
    subject = f'{statistics_email_subject} ({time_period})'
    attachments = {
        '1' : {
            'filename' : graph_filename,
            'type' : 'pdf',
        },
    }
    try:
        log.info("Sending statistics email...")
        ECCEmailer.send_email('\n'.join(body), 'html', attachments,
                              args.smtp_auth_file,
                              to, subject, smtp_from, log)
        log.info("Sent statistics email...")
    except Exception as e:
        log.error(f"==== Error sending email to {to}")
        log.error(f"Exception: {e}")
        exit(1)

##############################################################################

def unsubmitted_report(args, ps_families, member_workgroups,
                       jotform_data, log):
    log.info("Composing unsubmitted Families report...")

    ekey = 'emailAddress'
    wgkey = 'py workgroups'

    submitted_fduids = {}
    for fduid, data in jotform_data.items():
        fduid = int(fduid)
        submitted_fduids[fduid] = True

    # Produce a list of unsubmitted Families sorted by last name
    unsubmitted_fduids = []
    for fduid, family in ps_families.items():
        if fduid not in submitted_fduids:
            unsubmitted_fduids.append(fduid)

    unsubmitted_fduids = sorted(unsubmitted_fduids,
                                key=cmp_to_key(lambda i1, i2: \
                                               f'{ps_families[i1]["lastName"]} {ps_families[i1]["firstName"]} {ps_families[i1]["familyDUID"]}' < \
                                               f'{ps_families[i2]["lastName"]} {ps_families[i2]["firstName"]} {ps_families[i2]["familyDUID"]}'))

    last_year = stewardship_year - 1
    last_year_wg = f'Active: Stewardship {last_year}'

    filename = 'unsubmitted-families.csv'
    submitted_last_year = f'Submitted in {last_year}'
    with open(filename, 'w') as fp:
        fields = ['Last Name', 'Family Names', 'Family DUID',
                  'eStewardship Eligible',
                  submitted_last_year,
                  'Emails', 'Phones']
        writer = csv.DictWriter(fp, fieldnames=fields)
        writer.writeheader()

        for fduid in unsubmitted_fduids:
            family = ps_families[fduid]

            # If we get here, this family has not submitted
            members = \
                ParishSoft.family_business_logistics_emails_members(family,
                                                                    member_workgroups,
                                                                    log)

            emails = {}
            phones = {}
            for member in members:
                if ekey in member:
                    emails[member[ekey]] = True
                member_phones = ParishSoft.get_member_public_phones(member)
                for phone in member_phones:
                    phones[phone['number'].strip()] = True

            emails_str = '\n'.join(emails.keys())
            phones = '\n'.join(phones.keys())

            sly = 'False'
            if wgkey in family:
                for wg_name in family[wgkey]:
                    if wg_name == last_year_wg:
                        sly = 'True'
                        break

            item = {
                'Last Name' : family['lastName'],
                'Family Names' : f'{family["firstName"]} {family["lastName"]}',
                'Family DUID' : fduid,
                'eStewardship Eligible' : "True" if len(emails) > 0 else "False",
                submitted_last_year : sly,
                'Emails' : emails_str,
                'Phones' : phones,
            }
            writer.writerow(item)

    log.info(f"Wrote {filename}")

##############################################################################

def pledge_comparison_report(google, jotform_this_year, jotform_last_year, log):
    # Extract just 2 fields from the jotform data and return it in a
    # dict indexed by FDUID.
    def _simplify_jotform(jotform_data, participation_fieldname,
                          this_year_pledge_fieldname,
                          last_year_pledge_fieldname, log):
        out = dict()
        for fduid, data in jotform_data.items():
            participate = True
            if participation_fieldname and data[participation_fieldname].startswith("Because"):
                participate = False

            current_pledge = 0
            if participate:
                current_pledge = helpers.jotform_text_to_int(data[this_year_pledge_fieldname])
            previous_pledge = helpers.jotform_text_to_int(data[last_year_pledge_fieldname])
            out[fduid] = {
                'name'           : data['Family names'],
                'participate'    : participate,
                'current pledge' : current_pledge,
                'previous pledge': previous_pledge,
            }

        return out

    # ------------------------------------------------------------------------

    # Compares the dictionaries of pledges from this year to that of last year,
    # and outputs a CSV showing a few statistics relating to which category the
    # pledges falls into (Can't, Reduced, No Change, New, Increased, No Pledge
    # Both Years) and some relevant info / analysis on totals and percentages.
    def _compare(this_year_data, log):
        out = dict()
        for fduid in this_year_data:
            current_pledge = this_year_data[fduid]["current pledge"]
            previous_pledge = this_year_data[fduid]["previous pledge"]
            if this_year_data[fduid]["participate"] == False:
                category = "Cannot pledge"
            # If the family is recorded as having a pledge of 0 OR 1 last year,
            # then they did not have a pledge last year.
            elif previous_pledge == 0 or previous_pledge == 1:
                # Set their previous pledge to zero to accurately reflect how
                # much they *actually* pledged last year.
                previous_pledge = 0
                if current_pledge == 0:
                    category = "No pledge both years"
                # If the family didn't pledge last year, but pledged this year,
                # it's a new pledge.
                elif current_pledge > 0:
                    category = "NEW pledge"
            elif current_pledge == previous_pledge:
                category = "No change"
            elif current_pledge > previous_pledge:
                category = "Increased pledge"
            elif current_pledge < previous_pledge:
                category = "Reduced pledge"

            dollar_impact = current_pledge - previous_pledge

            if category not in out:
                out[category] = {
                    'households'       : 0,
                    'dollar impact'    : 0,
                    'total of pledges' : 0,
                }

            out[category]["households"] += 1
            out[category]["dollar impact"] += dollar_impact
            out[category]["total of pledges"] += current_pledge

        return out

    # ------------------------------------------------------------------------

    def _make_xlsx(comparison, log):
        workbook = Workbook()
        sheet = workbook.active

        comments_label    = "Comments"
        pledge_last_label = f'CY{stewardship_year-1} pledge'
        pledge_cur_label  = f'CY{stewardship_year} pledge'

        # Setup the title rows
        # Title rows + set column widths
        title_font   = Font(color='FFFF00')
        title_fill   = PatternFill(fgColor='0000FF', fill_type='solid')
        title_align  = Alignment(horizontal='center', wrap_text=True)

        wrap_align   = Alignment(horizontal='general', wrap_text=True)
        right_align  = Alignment(horizontal='right')

        money_format = "$##,###,###,###"
        percentage_format = "##.#"

        xlsx_cols = dict();
        def _add_col(name, width=15, format=None):
            col             = len(xlsx_cols) + 1
            xlsx_cols[name] = {'name' : name, 'format' : format,
                               'column' : col, 'width' : width }

        _add_col('Category', width=20)
        _add_col('Number of Households')
        _add_col('%-age of Total Submitted Households')
        _add_col('Dollar Impact')
        _add_col('Total of Pledges Submitted')
        _add_col('%-age of Total Pledges Submitted')

        # Make 2 rows of merged cells for wide titles
        def _make_merged_title_row(row, value):
            cell           = sheet.cell(row=row, column=1, value=value)
            cell.fill      = title_fill
            cell.font      = title_font
            cell.alignment = title_align
            end_col_char   = chr(ord('A') - 1 + len(xlsx_cols))
            sheet.merge_cells(f'A{row}:{end_col_char}{row}')

        _make_merged_title_row(row=1, value='eStewardship Pledge Analysis')
        _make_merged_title_row(row=2, value='')

        # Now add all the column titles
        for data in xlsx_cols.values():
            col            = data['column']
            cell           = sheet.cell(row=3, column=col, value=data['name'])
            cell.fill      = title_fill
            cell.font      = title_font
            cell.alignment = title_align

            col_char = chr(ord('A') - 1 + col)
            sheet.column_dimensions[col_char].width = data['width']

        # Finally, fill in all the data rows.
        # First, compute totals so that we can compute percentages.
        total_households = 0
        total_pledges    = 0
        total_impact     = 0
        for data in comparison.values():
            total_households += data['households']
            total_pledges    += data['total of pledges']
            total_impact     += data['dollar impact']

        def _fill(column, value, align=None, format=None):
            cell = sheet.cell(row=xlsx_row, column=column, value=value)

            want_format = True
            if (type(value) is int or type(value) is float) and value == 0:
                    want_format = False

            if want_format:
                if align:
                    cell.alignment = align
                if format:
                    cell.number_format = format

        xlsx_row = 4
        for category, data in comparison.items():
            _fill(1, category)
            _fill(2, data['households'])
            _fill(3, data['households'] / total_households * 100.0,
                format=percentage_format)
            _fill(4, data['dollar impact'],
                format=money_format)
            _fill(5, data['total of pledges'],
                format=money_format)
            _fill(6, data['total of pledges'] / total_pledges * 100.0,
                format=percentage_format)

            xlsx_row += 1

        _fill(1, 'Totals', align=right_align)
        _fill(2, total_households)
        _fill(4, total_impact, format=money_format)
        _fill(5, total_pledges, format=money_format)

        return workbook

    # ------------------------------------------------------------------------

    # Make simplified data structures from the full jotform data.  These will be
    # easier to compare. We use this year's jotform for last year's data because
    # it (this year's jotform) has pre-filled info on how much the family gave
    # last year.
    this_year_data = _simplify_jotform(jotform_this_year,
                                       f'CY{stewardship_year} participation',
                                       f'CY{stewardship_year} whole year pledge',
                                       f'CY{stewardship_year-1} pledge', log)



    # JMS
    filename = 'pledge-data-for-angie.csv'
    with open(filename, 'w') as fp:
        fields = ['Name', 'Family DUID', 'Participate', 'CY25 pledge', 'CY24 pledge']
        writer = csv.DictWriter(fp, fieldnames=fields)
        writer.writeheader()

        for duid, data in this_year_data.items():
            item = {
                'Name' : data['name'],
                'Family DUID' : duid,
                'Participate' : data['participate'],
                'CY25 pledge' : data['current pledge'],
                'CY24 pledge' : data['previous pledge'],
            }
            writer.writerow(item)

    log.info(f"JMS Wrote {filename}")

    filename = 'pledged-last-year-but-not-this-year.csv'
    with open(filename, 'w') as fp:
        fields = ['Name', 'Family DUID', 'Participate', 'CY25 pledge', 'CY24 pledge']
        writer = csv.DictWriter(fp, fieldnames=fields)
        writer.writeheader()

        for duid, data in this_year_data.items():
            item = {
                'Name' : data['name'],
                'Family DUID' : duid,
                'Participate' : data['participate'],
                'CY25 pledge' : data['current pledge'],
                'CY24 pledge' : data['previous pledge'],
            }
            writer.writerow(item)

    log.info(f"JMS Wrote {filename}")

    exit(1)


    # Do the comparison
    comparison = _compare(this_year_data, log)

    # Make an XLSX report
    workbook = _make_xlsx(comparison, log)

    # Upload the XLS to google
    now      = datetime.now()
    filename = f'{now.year:04}-{now.month:02}-{now.day:02} Pledge analysis.xlsx'
    gsheet_id, _ = upload_xlsx_to_gsheet(google,
                    google_folder_id=upload_team_drive_folder_id,
                    filename=filename,
                    workbook=workbook,
                    remove_local=True,
                    log=log)

    return gsheet_id

##############################################################################

def send_reports_email(time_period, comments_gfile, pledge_analysis_gfile, args, log):
    # Send the comments report email
    body = list()
    body.append(f"""<html>
<body>
<h2>{title} comments and pledge analysis reports</h2>

<h3>Time period: {time_period}</h3>""")

    if comments_gfile is None:
        body.append("<p>No comments submitted during this time period.</p>")
    else:
        url = 'https://docs.google.com/spreadsheets/d/{id}'.format(id=comments_gfile)
        body.append(f'<p><a href="{url}">Link to Google sheet containing comments for this timeframe</a>.</p>')

    url = 'https://docs.google.com/spreadsheets/d/{id}'.format(id=pledge_analysis_gfile)
    body.append(f'''<p><a href="{url}">Link to Google sheet containing pledge analysis.</a>.</p>
</body>
</html>''')

    to = reports_email_to
    subject = f'{reports_email_subject} ({time_period})'
    try:
        log.info(f'Sending "{subject}" email to {to}')
        ECCEmailer.send_email('\n'.join(body), 'html', None,
                              args.smtp_auth_file,
                              to, subject, smtp_from, log)
    except:
        log.error(f"==== Error sending email to {to}")
        log.error(f"Exception: {e}")
        exit(1)

##############################################################################

def _family_comparison_reports(ps_families, family_workgroups, jotform, log):
    stewardship_fduids = list()
    census_fduids = list()

    # Find everyone who did census last year
    census_wg_name = f'Active: Census {stewardship_year - 1}'
    for family_wg in family_workgroups.values():
        if family_wg['name'] == census_wg_name:
            for entry in family_wg['membership']:
                fduid = entry['familyId']
                census_fduids.append(fduid)

    # Index everyone who did estewardship this year
    # Jotform
    for row in jotform:
        stewardship_fduids.append(row['fduid'])

    # Compare
    census_not_stewardship_fduids = list()
    stewardship_not_census_fduids = list()

    for fduid in census_fduids:
        if fduid not in stewardship_fduids:
            census_not_stewardship_fduids.append(fduid)

    for fduid in stewardship_fduids:
        if fduid not in census_fduids:
            # First, ensure that the FDUID is still in PS
            if fduid not in ps_families:
                continue

            stewardship_not_census_fduids.append(fduid)

    # Make CSV data
    def _doit(rows):
        out = list()
        for fduid in rows:
            family = ps_families[fduid]
            item = {
                'Family DUID' : fduid,
                'Family Name' : f"{family['firstName']} {family['lastName']}",
                'Last Name'   : family['lastName'],
            }
            out.append(item)

        return out

    census_not_stewardship = _doit(census_not_stewardship_fduids)
    stewardship_not_census = _doit(stewardship_not_census_fduids)

    log.info(f"Found {len(census_not_stewardship)} families who submitted census but not stewardship")
    log.info(f"Found {len(stewardship_not_census)} families who submitted stewardship but not census")

    return census_not_stewardship, stewardship_not_census

#-----------------------------------------------------------------------------

# Find who did census earlier this year, but did not submit
# stewardship in this campaign.
#
# Also find who did stewardship in this compaign but did not submit
# census earlier this year.
def family_comparison_reports(args, google, ps_families,
                              family_workgroups, jotform, log):
    def _doit(filename, rows):
        if len(rows) == 0:
            return

        gsheet_id = None
        gsheet_id, csv_filename = upload_csv_to_gsheet(google,
                                        google_folder_id=upload_team_drive_folder_id,
                                        filename=filename,
                                        fieldnames=rows[0].keys(),
                                        csv_rows=rows,
                                        remove_local=False,
                                        log=log)

        url = 'https://docs.google.com/spreadsheets/d/{id}'.format(id=gsheet_id)
        log.info(f"Uploaded to {url}")

    #-------------------------------------------------------------------------

    census_not_stewardship, stewardship_not_census = \
        _family_comparison_reports(ps_families, family_workgroups, jotform, log)

    # If we have pledges, upload them to a Google sheet
    _doit('Families who submitted census but not stewardship.csv',
          census_not_stewardship)
    _doit('Families who submitted stewardship but not census.csv',
          stewardship_not_census)

##############################################################################

# Columns that I need
# - fduid
# - Recurring Charge Name:
# - Terms / Frequency: Weekly, Biweekly, Monthly, Bimonthly, Semimonthly, Quarterly, Semiannually, Annually
# - Begin Date: 01/01/2020
# - End Date: 12/31/2020
# - Rate
# - Total pledge
def convert_pledges_to_ps_import(ps_families, jotform, log):
    def _map_to_freqrate(pledge):
        freq = pledge[f'CY{stewardship_year} how fullfill']
        amount = helpers.jotform_text_to_int(pledge[f'CY{stewardship_year} whole year pledge'])

        if freq == 'Weekly donations':
            return 'Weekly', amount / 52
        elif freq == 'Monthly donations':
            return 'Monthly', amount / 12
        elif freq == 'Quarterly donations':
            return 'Quarterly', amount / 4
        elif freq == 'One annual donation':
            return 'Annually', amount
        else:
            return None, None

    #-------------------------------------------------------------------------

    out = list()

    for pledge in jotform:
        # Skip the title row
        fduid = pledge['fduid']
        if 'Family DUID' in fduid:
            continue

        # Here's something that can happen: A family may be deleted from PS
        # even though they submitted.  In this case, skip them.
        fduid = int(fduid)
        if fduid not in ps_families:
            log.warning(f"WARNING: Family FDUID {fduid} / {pledge['Family names']} submitted a pledge, but is no longer in PS")
            continue

        # If there is a $0 pledge, Per Lynne's comment, we'll
        # transform this into a $1 annual pledge -- just so that this
        # person is on the books, so to speak.
        pledge_field  = f'CY{stewardship_year} whole year pledge'
        pledge_amount = helpers.jotform_text_to_int(pledge[pledge_field])
        if not pledge_amount or pledge_amount == '' or pledge_amount == 0:
            pledge_amount = 1
            pledge[pledge_field] = pledge_amount
            pledge[f'CY{stewardship_year} how fullfill'] = 'One annual donation'

        frequency, rate = _map_to_freqrate(pledge)

        # Round pledge value and rate to 2 decimal points, max
        if rate == None:
            rate = 0
        rate  = float(int(rate * 100)) / 100.0
        total = float(int(pledge_amount) * 100) / 100.0

        # Use an OrderedDict to keep the fields in order
        row = collections.OrderedDict()
        row['Family DUID']   = fduid
        row['RecChargeName'] = 'Due/Contributions'
        row['Frequency']     = frequency
        row['BeginDate']     = stewardship_begin_date.strftime('%m/%d/%Y')
        row['EndDate']       = stewardship_end_date.strftime('%m/%d/%Y')
        row['PledgeRate']    = rate
        row['TotalPledge']   = total
        row['SubmitDate']    = pledge['SubmitDate']
        row['Names']         = pledge['Family names']
        # JMS This whole section may be scrapped anyway...?
        #row['Envelope ID']   = helpers.pkey_url(pledge['EnvId'])[2:]
        #row['Stupid Envelope ID']   = helpers.pkey_url(pledge['EnvId'])[4:]

        # Calculate family pledge values for last CY
        family = ps_families[fduid]
        helpers.calculate_family_values(family, stewardship_year - 2, log)

        row[f'CY{stewardship_year - 2} YTD gifts'] = family['calculated']['gifts']

        # Calculate family pledge values for this CY
        helpers.calculate_family_values(family, stewardship_year - 1, log)

        row[f'CY{stewardship_year - 1} YTD gifts'] = family['calculated']['gifts']
        row[f'CY{stewardship_year - 1} pledge']    = family['calculated']['pledged']

        # Add column for how they want to fullfill their pledge
        row[f'CY{stewardship_year} how fullfill'] = pledge[f'CY{stewardship_year} how fullfill']
        row[f'CY{stewardship_year} mechanism'] = pledge[f'CY{stewardship_year} how']

        # Add a column for this Family's "Envelope user" value
        row['PS Envelope User'] = family['EnvelopeUser']

        # Add a column for whether the Family selected the "offeratory
        # envelopes" option on the Jotform.
        val = False
        if 'Offertory' in pledge[f'CY{stewardship_year} how']:
            val = True
        row['Jotform asked for Envelopes'] = val

        out.append(row)

    return out

#-----------------------------------------------------------------------------

def family_pledge_csv_report(args, google, ps_families, jotform, log):
    pledges = convert_pledges_to_ps_import(ps_families, jotform, log)

    # If we have pledges, upload them to a Google sheet
    gsheet_id = None
    if len(pledges) > 0:
        filename = 'Family Pledge PS import.csv'
        gsheet_id, csv_filename = upload_csv_to_gsheet(google,
                                        google_folder_id=upload_team_drive_folder_id,
                                        filename=filename,
                                        fieldnames=pledges[0].keys(),
                                        csv_rows=pledges,
                                        remove_local=False,
                                        log=log)

    # Send the statistics report email
    body = list()
    body.append(f"""<html>
<body>
<h2>{title} pledge update</h2>

""")

    if gsheet_id:
        url = 'https://docs.google.com/spreadsheets/d/{id}'.format(id=gsheet_id)
        body.append(f'<p><a href="{url}">Link to Google sheet containing pledge updates for this timeframe</a>.</p>')
        body.append("<p>See the attachment for a CSV to import directly into PS.</p>")
    else:
        body.append("<p>There were no pledge submissions during this timeframe.<p>")

    body.append("""</body>
</html>""")

    try:







        # JMS Convert to ECCEmailer







        to = pledge_email_to
        log.info(f'Sending "{pledge_email_subject}" email to {to}')
        with smtplib.SMTP_SSL(host=smtp_server,
                              local_hostname='epiphanycatholicchurch.org') as smtp:

            # This assumes that the file has a single line in the format of username:password.
            with open(args.smtp_auth_file) as f:
                line = f.read()
                smtp_username, smtp_password = line.split(':')

            # Login; we can't rely on being IP whitelisted.
            try:
                smtp.login(smtp_username, smtp_password)
            except Exception as e:
                log.error(f'Error: failed to SMTP login: {e}')
                exit(1)

            msg = EmailMessage()
            msg['Subject'] = pledge_email_subject
            msg['From'] = smtp_from
            msg['To'] = to
            msg.set_content('\n'.join(body))
            msg.replace_header('Content-Type', 'text/html')

            # If there were results, attach the CSV
            if gsheet_id:
                with open(filename, "rb") as f:
                    csv_data = f.read()
                msg.add_attachment(csv_data, maintype='text', subtype='csv', filename=filename)

            smtp.send_message(msg)

            if gsheet_id:
                os.unlink(filename)
    except:
        print("==== Error with {email}".format(email=to))
        print(traceback.format_exc())

##############################################################################

MINISTRY_PARTICIPATE = 1
MINISTRY_INTERESTED  = 2
MINISTRY_INACTIVE    = 3

def _convert_jotform_ministry_status(jotform_status):
    if jotform_status is None:
        return MINISTRY_INACTIVE

    if 'PARTICIPATE' in jotform_status:
        return MINISTRY_PARTICIPATE
    elif 'INTERESTED' in jotform_status:
        return MINISTRY_INTERESTED
    elif 'NO LONGER' in jotform_status:
        return MINISTRY_INACTIVE
    else:
        # Catch-all
        return MINISTRY_INACTIVE

def _status_to_str(status):
    if status == MINISTRY_PARTICIPATE:
        return 'I already participate'
    elif status == MINISTRY_INTERESTED:
        return 'I am interested'
    else:
        return 'I do not participate'

##############################################################################

# Generic helper function that traverses each Family in the jotform,
# and each Member's submission in that Family.  Callback hook
# functions to process each.
def traverse_jotform_members(ps_families, ps_members,
                             jotform_list,
                             per_family_func, per_family_state,
                             per_member_func, per_member_state,
                             log):
    # Each row is a family
    for jrow in jotform_list:
        fduid = int(jrow['fduid'])

        # Make sure the family is still active
        if fduid not in ps_families:
            log.warn(f"WARNING: Family {fduid} submitted, but cannot be found -- skipping")
            continue

        family = ps_families[fduid]
        log.info(f"Processing Jotform Family submission: {family['firstName']} {family['lastName']} (FDUID {fduid})")

        if per_family_func:
            per_family_func(jrow, family, per_family_state, log)

        # Check the members in this row
        for member_num in range(MAX_PS_FAMILY_MEMBER_NUM):
            column_names = jotform_gsheet_columns['members'][member_num]
            # The 0th entry in each Member is the MDUID
            mduid = jrow[column_names[0]]

            # If there's no MDUID for this Member, then there's no Member.
            # We're done with the loop.
            if mduid == '':
                break

            # Here's something that can happen: a MDUID was submitted, but is no
            # longer an active member in PS.  In that case, ignore the submission.
            mduid = int(mduid)
            if mduid not in ps_members:
                log.warn(f"WARNING: Member {mduid} submitted, but cannot be found -- skipping")
                continue

            member = ps_members[mduid]
            log.info(f"  Processing Jotform member {member['emailAddress']} (MDUID {mduid})")

            if per_member_func:
                per_member_func(jrow, member_num, member,
                                per_member_state, log)

#-----------------------------------------------------------------------------

# Fields I need:
# - MDUID
# - Ministry
# - Status
# - Start date (from constants.py)
# - End date (from constants.py)
# Other fields requested by the staff:
# - Member nickname + last name
# - Member email addresses
# - Member phones
# - Member status on ministry
#
# Will generate several outputs:
# - Likely can be directly imported
#   - NOT PARTICIPATE -> INTERESTED
# - Likely cannot be directly imported (still need to test)
#   - PARTICPATE -> NOT PARTICIPATE
# - Needs to be examined by a human
#   - NOT PARTICIPATE -> ALREADY PARTICIPATE
#   - PARTICIPATE -> INTERESTED
#
# Output will be a dictionary:
# - PS ministry name (in the case of multi-names, like lectors, use the jotform name): dictionary of
#   - interested: list of members
#   - no_longer_interested: list of members
#   - needs_human: list of members
def analyze_member_ministry_submissions(ps_members, ps_families, ps_ministries,
                                        jotform_csv, log):
    def _get_ps_member_ministry_status(member, ministry_names):
        ministry = None
        for m_duid, m_data in ps_ministries.items():
            if m_data['name'] in ministry_names:
                ministry = m_data
                break

        if ministry is None:
            log.critical(f"Cannot find PS ministry with any of these names: {ministry_names}")
            exit(1)

        duid = member['memberDUID']
        for ministry_member in ministry['membership']:
            if ministry_member['memberId'] == duid:
                return MINISTRY_PARTICIPATE, 'Active'

        return MINISTRY_INACTIVE, 'Inactive'

    #-------------------------------------------------------------------------

    def _status_to_ps(status):
        if status == MINISTRY_PARTICIPATE:
            return 'Actively Involved'
        elif status == MINISTRY_INACTIVE:
            return 'No Longer Involved'
        elif status == MINISTRY_INTERESTED:
            return 'Interested'

    #-------------------------------------------------------------------------

    # Analyze one Member's submission in a jotform row.  This is a
    # callback/hook function for traverse_jotform_members().
    def _analyze_member(jrow, member_num, member, state, log):
        member_interested = False
        member_no_longer_interested = False
        member_needs_human = False

        output = state['output']

        # Go through the list of ministry grids from the jotform
        for grid in jotform.ministry_grids:
            # Go through the rows in the jotform ministry grid
            for mrow in grid.rows:
                # Each row has its PS ministry name
                ministry_entry = mrow['ps_ministry']

                # Some ministry rows are lists because we want to treat them equivalently
                if type(ministry_entry) is list:
                    output_key = mrow['row_heading']
                    ministries = ministry_entry
                else:
                    output_key = ministry_entry
                    ministries = [ ministry_entry ]

                # Get the Member's status in this Ministry from the jotform
                jotform_column_name = mrow['jotform_columns'][member_num]
                jotform_status_str  = jrow[jotform_column_name]
                jotform_status      = _convert_jotform_ministry_status(jotform_status_str)

                # Get their status from PS
                ps_status, ps_status_string = _get_ps_member_ministry_status(member, ministries)
                key = 'jotform'
                if key not in member:
                    member[key] = dict()
                member[key][output_key] = ps_status_string

                # If they're the same, nothing to do
                if jotform_status == ps_status:
                    continue

                if output_key not in output:
                    output[output_key] = dict()

                # If PS INACTIVE -> INTERESTED
                if (ps_status == MINISTRY_INACTIVE and
                    jotform_status == MINISTRY_INTERESTED):
                    key = 'Interested'
                    if key not in output[output_key]:
                        output[output_key][key] = list()
                    output[output_key][key].append(member)
                    member_interested = True

                elif (ps_status == MINISTRY_PARTICIPATE and
                      jotform_status == MINISTRY_INACTIVE):
                    key = 'No longer interested'
                    if key not in output[output_key]:
                        output[output_key][key] = list()
                    output[output_key][key].append(member)
                    member_no_longer_interested = True

                elif (ps_status == MINISTRY_INACTIVE and
                      jotform_status == MINISTRY_PARTICIPATE):
                    key = 'Needs human: PS=inactive, but Jotform=active'
                    if key not in output[output_key]:
                        output[output_key][key] = list()
                    output[output_key][key].append(member)
                    member_needs_human = True

                elif (ps_status == MINISTRY_PARTICIPATE and
                      jotform_status == MINISTRY_INTERESTED):
                    key = 'Needs human: PS=active, but Jotform=interested'
                    if key not in output[output_key]:
                        output[output_key][key] = list()
                    output[output_key][key].append(member)
                    member_needs_human = True

        if member_interested:
            state['num_members_interested'] += 1
        if member_no_longer_interested:
            state['num_members_no_longer_interested'] += 1
        if member_needs_human:
            state['num_members_needs_human'] += 1

    #----------------------------------------------

    state = {
        'output' : dict(),
        'num_members_interested' : 0,
        'num_members_no_longer_interested' : 0,
        'num_members_needs_human' : 0,
    }

    traverse_jotform_members(ps_families, ps_members,
                             jotform_csv,
                             None, None,
                             _analyze_member, state,
                             log)

    log.info(f"Total number of Members who were interested in a new ministry: {state['num_members_interested']}")
    log.info(f"Total number of Members who were no longer interested in an existing ministry: {state['num_members_no_longer_interested']}")
    log.info(f"Total number of Members who submitted an ambiguous response: {state['num_members_needs_human']}")

    return state['output']

#-----------------------------------------------------------------------------

def member_ministry_csv_report(args, google, start, end, time_period,
                               ps_members, ps_families, ps_ministries,
                               jotform_csv, log):
    def _find_all_phones(member):
        found = list()

        for tuple in [('mobilePhone', 'cell'),
                      ('homePhone', 'home'),
                      ('workPhone', 'work')]:
            key = tuple[0]
            name = tuple[1]
            if key in member and member[key]:
                found.append(f'{member[key]} ({name})')

        # When used with XLSX word wrapping alignment, this will
        # across put each phone number on a separate line, but all
        # within a single cell.
        return '\r\n'.join(found)

    #--------------------------------------------------------------------

    def _find_family_home_phone(member):
        family = member['py family']
        key = 'family_HomePhone'
        if key in family and family[key]:
            return family[key]

        key = 'homePhone'
        if key in member and member[key]:
            return member[key]

        return ""

    #--------------------------------------------------------------------

    today  = date.today()
    output = analyze_member_ministry_submissions(ps_members, ps_families,
                                                 ps_ministries,
                                                 jotform_csv, log)

    #--------------------------------------------------------------------

    title_font  = Font(color='FFFF00')
    title_fill  = PatternFill(fgColor='0000FF', fill_type='solid')
    title_align = Alignment(horizontal='center', wrap_text=True)

    wrap_align  = Alignment(horizontal='general', wrap_text=True)
    right_align = Alignment(horizontal='right')

    #--------------------------------------------------------------------

    def _doit(output, filename_suffix, category_match, log):

        def _setup_new_workbook(xlsx_cols):
            workbook    = Workbook()
            sheet       = workbook.active

            for data in xlsx_cols.values():
                col            = data['column']
                cell           = sheet.cell(row=1, column=col, value=data['name'])
                cell.fill      = title_fill
                cell.font      = title_font
                cell.alignment = title_align

                col_char = chr(ord('A') - 1 + col)
                sheet.column_dimensions[col_char].width = data['width']

            sheet.freeze_panes = sheet['A2']

            return workbook

        #--------------------------------------------------------------------

        def _fill(col_name, value, align=None, format=None):
            col_data = xlsx_cols[col_name]
            cell     = sheet.cell(row=xlsx_row, column=col_data['column'], value=value)
            if align:
                cell.alignment = align
            if format:
                cell.number_format = format

        #--------------------------------------------------------------------

        def _add_col(name, width=10):
            col             = len(xlsx_cols) + 1
            xlsx_cols[name] = {'name' : name, 'column' : col, 'width' : width }

        #--------------------------------------------------------------------

        xlsx_cols = dict();
        _add_col('Full Name', width=20)
        _add_col('First')
        _add_col('Last')
        _add_col('Age')
        _add_col('Mem DUID')
        _add_col('Email', width=30)
        _add_col('Member phones', width=20)
        _add_col('Family home phone', width=20)
        _add_col('Category', width=25)
        _add_col('Current ministry status', width=20)
        _add_col('PS ministry name', width=50)

        for ministry_name in sorted(output.keys()):
            workbook = _setup_new_workbook(xlsx_cols)
            sheet    = workbook.active

            data = output[ministry_name]
            xlsx_row = 2
            for category in sorted(data.keys()):
                match = re.search(category, category_match)
                if match is None:
                    continue

                for member in data[category]:
                    family = member['py family']

                    _fill('Full Name', member['py friendly name FL'])
                    _fill('First', member['firstName'])
                    _fill('Last', member['lastName'])
                    if member['birthdate']:
                        age = today - member['birthdate']
                        _fill('Age', int(age.days / 365))
                    _fill('Mem DUID', member['memberDUID'])
                    _fill('Email', member['emailAddress'])
                    _fill('Member phones', _find_all_phones(member), align=wrap_align)
                    _fill('Family home phone', _find_family_home_phone(member))
                    _fill('Category', category.capitalize(), align=wrap_align)
                    _fill('Current ministry status', member['jotform'][ministry_name],
                          align=wrap_align)
                    _fill('PS ministry name', ministry_name)

                    xlsx_row += 1

            # If there was no data here, don't bother writing out the XLSX
            if xlsx_row == 2:
                continue

            # Write out the XLSX with the results
            filename = f'{ministry_name} {filename_suffix}.xlsx'.replace('/', '-')
            if os.path.exists(filename):
                os.unlink(filename)
            workbook.save(filename)
            log.info(f"Wrote to filename: {filename}")

    _doit(output, "jotform results", '.', log)
    _doit(output, "no longer interested", 'No longer interested', log)

#-----------------------------------------------------------------------------

# Columns that I need
# - fduid
# - per member:
#   - talent
#   - prayer group
#   - small prayer group (yes/no)
#   - protect (dishwasher, recycle helper)
def extract_member_extras(ps_families, ps_members, jotform, log):
    def _extract(jotform_row, member_jotform_num, member, log):
        item = dict()
        for name in jotform_gsheet_columns['per-member epilog']:
            jcolumn_name = f'{name} {member_jotform_num+1}'
            value = jotform_row[jcolumn_name]
            if value and len(value) > 0:
                item[name] = value.split('\n')

        member[member_extras_key] = item

    traverse_jotform_members(ps_families, ps_members, jotform_list,
                             None, None,
                             _extract, None,
                             log)

#-----------------------------------------------------------------------------

def _member_extras_csv(ps_members, google, key, filename, log):
    # Find all the possible values of this field
    values = dict()
    for mduid, member in ps_members.items():
        if member_extras_key not in member:
            continue
        if key not in member[member_extras_key]:
            continue
        for value in member[member_extras_key][key]:
            values[value] = True

    # Write these all out into a file
    with open(filename, 'w') as fp:
        fields = ['Member', 'MDUID' ]
        fields.extend(sorted(values.keys()))

        writer = csv.DictWriter(fp, fieldnames=fields)
        writer.writeheader()

        for mduid, member in ps_members.items():
            if member_extras_key not in member:
                continue
            if key not in member[member_extras_key]:
                continue

            item = {
                'Member' : member['py friendly name FL'],
                'MDUID' : mduid,
            }
            for value in member[member_extras_key][key]:
                item[value] = 'Yes'
            writer.writerow(item)

    log.info(f"Wrote {filename}")

def member_talent_csv_report(ps_members, google, log):
    # This key field is defined in constants.py
    key = 'member special talent'
    filename = 'member-talents.csv'
    _member_extras_csv(ps_members, google, key, filename, log)

def member_group_prayer_events_csv_report(ps_members, google, log):
    # This key field is defined in constants.py
    key = 'member group prayer events'
    filename = 'member-group-prayer-events.csv'
    _member_extras_csv(ps_members, google, key, filename, log)

def member_small_group_prayer_csv_report(ps_members, google, log):
    # This key field is defined in constants.py
    key = 'member small prayer group'
    filename = 'member-small-prayer-group.csv'
    _member_extras_csv(ps_members, google, key, filename, log)

def member_protect_csv_report(ps_members, google, log):
    # This key field is defined in constants.py
    key = 'member protect'
    filename = 'member-protect.csv'
    _member_extras_csv(ps_members, google, key, filename, log)

##############################################################################

def count_ministry_participation(members, families, log):
    min_age = 14
    buckets = defaultdict(int)
    count_of_members_lt_min_age = 0
    count_of_members_ge_min_age = 0
    count_of_members_no_age = 0
    families_with_active_member = dict()

    for mduid, member in members.items():
        key = 'age'
        if key not in member:
            count_of_members_no_age += 1
            continue
        if not member[key]:
            count_of_members_no_age += 1
            continue
        if member[key] < min_age:
            count_of_members_lt_min_age += 1
            continue

        count_of_members_ge_min_age += 1

        key = 'py ministries'
        if key not in member:
            continue

        count = len(member[key])
        if count > 0:
            buckets[count] += 1
            fduid = member['familyDUID']
            families_with_active_member[fduid] = True

    log.info(f"Number of Members who have no age: {count_of_members_no_age}")
    log.info(f"Number of Members who are <{min_age}: {count_of_members_lt_min_age}")
    log.info(f"Number of Members who are >={min_age}: {count_of_members_ge_min_age}")

    log.info(f"Buckets of Members who are >={min_age} years old who are active in at least one ministry:")
    total = 0
    for num_ministries in sorted(buckets):
        count = buckets[num_ministries]
        total += count
        log.info(f"Active in {num_ministries} ministries: {count}")
    log.info(f"Total active in at least 1 ministry: {total}")

    log.info(f"Number of Families with >=1 Member who is >={min_age} years old who are active in at least one ministry: {len(families_with_active_member)}")

##############################################################################

def _export_gsheet_to_csv(service, start, end, google_sheet_id, fieldnames, log):
    response = service.files().export(fileId=google_sheet_id,
                                      mimeType=Google.mime_types['csv']).execute()

    # Write the result into a temp file so that csv can re-read the file and
    # parse newlines in fields correctly.
    filename = 'tempfile-google-data.csv'
    if os.path.exists(filename):
        os.unlink(filename)
    with open(filename, 'w') as fp:
        raw_data = response.decode('utf-8')
        fp.write(raw_data)

    with open(filename, newline='') as fp:
        csvreader = csv.DictReader(fp,
                                   fieldnames=fieldnames)

        rows = list()
        for row in csvreader:
            # Skip title row
            if 'Submission' in row['SubmitDate']:
                continue
            if row['fduid'] == '':
                continue

            # As of Sep 2021, Google Sheets CSV export sucks. :-(
            # The value of the "Edit Submission" field from Jotform is something
            # like:
            #
            # =HYPERLINK("https://www.jotform.com/edit/50719736733810","Edit Submission")
            #
            # Google Sheet CSV export splits this into 2 fields.  The first one
            # has a column heading of "Edit Submission" (which is what the
            # Jotform-created sheet column heading it) and contains the long number
            # in the URL.  The 2nd one has no column heading, and is just the words
            # "Edit Submission".  :-(  CSV.DictReader therefore puts a value of
            # "Edit Submission" in a dict entry of "None" (because it has no column
            # heading).
            #
            # For our purposes here, just delete the "None" entry from the
            # DictReader.
            if None in row and row[None] == ['Edit Submission']:
                del row[None]

            # Is this submission between start and end?
            if start is not None and end is not None:
                submit_date = helpers.jotform_date_to_datetime(row['SubmitDate'])
                if submit_date < start or submit_date > end:
                    continue

            rows.append(row)

    # Remove the temporary file used to store the google results
    os.unlink(filename)

    return rows

#-----------------------------------------------------------------------------

def read_jotform_gsheet(google, start, end, fieldnames, gfile_id, log):
    log.info(f"Downloading Jotform raw data ({gfile_id})...")

    # Some of the field names will be lists.  In those cases, use the first
    # field name in the list.
    final_fieldnames = list()
    final_fieldnames.extend(fieldnames['prelude'])
    for member in fieldnames['members']:
        final_fieldnames.extend(member)
    final_fieldnames.extend(fieldnames['family'])
    final_fieldnames.extend(fieldnames['epilog'])

    csv_data = _export_gsheet_to_csv(google, start, end, gfile_id,
                                      final_fieldnames, log)

    # Deduplicate: save the last row number for any given FDUID
    # (we only really care about the *last* entry that someone makes)
    out_dict = dict()
    for row in csv_data:
        fduid = row['fduid']

        # Skip the title row
        if fduid == 'Family DUID':
            continue

        out_dict[fduid] = row

    # Turn this dictionary into a list of rows
    out_list = [ out_dict[fduid] for fduid in sorted(out_dict) ]

    return out_list, out_dict

##############################################################################

def setup_args():
    tools.argparser.add_argument('--gdrive-folder-id',
                                 help='If specified, upload a Google Sheet containing the results to this Team Drive folder')

    tools.argparser.add_argument('--all',
                                 action='store_const',
                                 const=True,
                                 help='If specified, run the comparison for all time (vs. running for the previous time period')

    tools.argparser.add_argument('--smtp-auth-file',
                                 required=True,
                                 help='File containing SMTP AUTH username:password')

    global gapp_id
    tools.argparser.add_argument('--app-id',
                                 default=gapp_id,
                                 help='Filename containing Google application credentials')
    global guser_cred_file
    tools.argparser.add_argument('--user-credentials',
                                 default=guser_cred_file,
                                 help='Filename containing Google user credentials')

    # ParishSoft args
    tools.argparser.add_argument('--ps-api-keyfile',
                                 default='parishsoft-api-key.txt',
                                 help='File containing the ParishSoft API key')
    tools.argparser.add_argument('--ps-cache-dir',
                                 default='ps-data',
                                 help='Directory to cache the ParishSoft data')

    tools.argparser.add_argument('--debug', action='store_true',
                                 help='Enable additional debug logging')

    args = tools.argparser.parse_args()

    # Read the PS API key
    if not os.path.exists(args.ps_api_keyfile):
        print(f"ERROR: ParishSoft API keyfile does not exist: {args.ps_api_keyfile}")
        exit(1)
    with open(args.ps_api_keyfile) as fp:
        args.api_key = fp.read().strip()

    return args

##############################################################################

def main():
    global families, members

    args = setup_args()
    log = ECC.setup_logging(debug=args.debug)

    #---------------------------------------------------------------

    # Calculate the start and end of when we are analyzing in the
    # source data
    start = None
    end = None
    if args.all:
        time_period = 'all results to date'

    else:
        # If not supplied on the command line:
        # Sun: skip
        # Mon: everything from last 3 days (Sat+Sun)
        # Tue-Fri: everything from yesterday
        # Sat: skip
        today = end.strftime('%a')
        if today == 'Sat' or today == 'Sun':
            print("It's the weekend.  Nothing to do!")
            exit(0)
        elif today == 'Mon':
            start = end - timedelta(days=3)
        else:
            start = end - timedelta(days=1)

        # No one wants to see the microseconds
        start = start - timedelta(microseconds=start.microsecond)
        end   = end   - timedelta(microseconds=end.microsecond)

        time_period = '{start} - {end}'.format(start=start, end=end)

    log.info("Time period: {tp}".format(tp=time_period))

    #---------------------------------------------------------------

    # Read in all the ParishSoft data
    log.info("Loading ParishSoft data...")
    families, members, family_workgroups, member_workgroups, ministries = \
        ParishSoft.load_families_and_members(api_key=args.api_key,
                                             cache_dir=args.ps_cache_dir,
                                             active_only=True,
                                             parishioners_only=True,
                                             log=log)
    log.info(f"Loaded {len(families)} ParishSoft Families")
    log.info(f"Loaded {len(members)} ParishSoft Members")

    #---------------------------------------------------------------

    apis = {
        'drive' : { 'scope'       : Google.scopes['drive'],
                    'api_name'    : 'drive',
                    'api_version' : 'v3', },
    }
    services = GoogleAuth.service_oauth_login(apis,
                                              app_json=args.app_id,
                                              user_json=args.user_credentials,
                                              log=log)
    google = services['drive']

    #---------------------------------------------------------------

    # Load all the results
    log.info(f"Reading Jotform {stewardship_year} data...")
    jotform_all_list, jotform_all_dict = read_jotform_gsheet(google,
                                    start=None, end=None,
                                    fieldnames=jotform_gsheet_columns,
                                    gfile_id=jotform_gsheet_gfile_id,
                                    log=log)

    # Load a range of results
    if start is None:
        jotform_range_list = jotform_all_list.copy()
    else:
        jotform_range_list, jotform_range_dict = read_jotform_gsheet(google,
                                    start=start, end=end,
                                    fieldnames=jotform_gsheet_columns,
                                    gfile_id=jotform_gsheet_gfile_id,
                                    log=log)

    #---------------------------------------------------------------

    # These are periodic reports that are run during the campaign
    # JMS Ran the final one of these for Angie/Don/Mary
    if False:
        # Stats of how many families have submitted, etc.
        statistics_report(args, time_period, members, families,
                          member_workgroups,
                          jotform_all_list, log)

    # Get a list of who has not submitted yet
    if False:
        unsubmitted_report(args, families, member_workgroups,
                           jotform_all_dict, log)

    # JMS Ran this for Angie
    comments_gfile = None
    if False:
        # A collection of all the random text comments that people
        # submitted (so that staff members can act on them).
        comments_gfile = comments_report(args, google, start, end, time_period,
                                         jotform_range_list, log)

    # JMS: Ran this for Mary and Angie
    if False:
        # A comparison of this year's pledges vs. last year's pledges.
        pledge_gfile = None
        # Looks like we don't really need to load the Jotform from the
        # prior year -- we have the 1 field we need in this year's
        # jotform.
        jotform_last_year = None

        pledge_gfile = pledge_comparison_report(google, jotform_all_dict,
                                                jotform_last_year, log)

        #send_reports_email(time_period, comments_gfile, pledge_gfile, args, log)

    if False:
        # Compare who submitted census vs. stewardship
        family_comparison_reports(args, google, families,
                                  family_workgroups,
                                  jotform_all_list, log)

    ##########################################################

    # These reports are generally run after the campaign
    # JMS: Ran this for Mary
    if False:
        # Raw list of pledges (I think this is importable to PS...?)
        family_pledge_csv_report(args, google, families, jotform_all_list, log)

    # JMS Sent these to Angie
    if False:
        # Per-ministry CSVs showing member status changes (given to
        # staff members to review, and ultimately to make phone calls
        # to followup).
        member_ministry_csv_report(args, google, start, end, time_period,
                                   members, families, ministries,
                                   jotform_range_list, log)

    # JMS Sent these to Angie
    if False:
        # Per-ministry CSVs showing member status changes (given to
        # staff members to review, and ultimately to make phone calls
        # to followup).
        extract_member_extras(families, members, jotform_all_list, log)

        member_talent_csv_report(members, google, log)
        member_group_prayer_events_csv_report(members, google, log)
        member_small_group_prayer_csv_report(members, google, log)
        member_protect_csv_report(members, google, log)

    # JMS Sent these to Angie
    if True:
        count_ministry_participation(members, families, log)

main()
