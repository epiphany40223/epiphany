#!/usr/bin/env python3

import os
import re
import sys
import csv
import uuid
import time
import smtplib
import sqlite3
import datetime
import calendar
import argparse
import openpyxl

# We assume that there is a "ecc-python-modules" sym link in this
# directory that points to the directory with ECC.py and friends.
moddir = os.path.join(os.getcwd(), 'ecc-python-modules')
if not os.path.exists(moddir):
    print("ERROR: Could not find the ecc-python-modules directory.")
    print("ERROR: Please make a ecc-python-modules sym link and run again.")
    exit(1)
# On MS Windows, git checks out sym links as a file with a single-line
# string containing the name of the file that the sym link points to.
if os.path.isfile(moddir):
    with open(moddir) as fp:
        dir = fp.readlines()
    moddir = os.path.join(os.getcwd(), dir[0])

sys.path.insert(0, moddir)

import ECC
import Google
import GoogleAuth
import ParishSoftv2 as ParishSoft

from oauth2client import tools

from pprint import pprint
from pprint import pformat

#--------------------------------------------------------------------------

import helpers

from constants import stewardship_fam_cur_year_wg
from constants import stewardship_fam_prev_year_wg

from constants import gapp_id
from constants import guser_cred_file
from constants import jotform_gsheet_gfile_id
from constants import jotform_gsheet_columns

from constants import stewardship_year
from constants import title

from constants import smtp_server
from constants import smtp_from

smtp_subject = 'Epiphany ' + title

from constants import email_image_url
from constants import api_base_url

from constants import jotform

from constants import COL_AM_INVOLVED
from constants import COL_NOT_INVOLVED

from constants import MAX_PS_FAMILY_MEMBER_NUM

#--------------------------------------------------------------------------

# JMS Kinda yukky that "args" is global
args = None

###########################################################################

exp_regex = re.compile(r"^\d+E\d+$")

# Returns the redirect URL
def insert_url_cookie(fduid, url, cookies, log):

    # See if there's a cookie for this Family DUID already
    def _get_cookie():
        # Get just the latest entry
        query = 'SELECT cookie FROM cookies WHERE fduid=:fduid ORDER BY creation_timestamp DESC LIMIT 1'
        values = {
            'fduid' : fduid,
        }
        results = cookies.execute(query, values)
        cookie = None
        row = results.fetchone()
        if row:
            cookie = row[0]

        return cookie

    # Make a cookie that is unique
    def _generate_cookie():
        while True:
            raw_uuid = uuid.uuid4()
            str_uuid = str(raw_uuid)
            cookie = str_uuid[0:6].upper()

            # Make sure that this is an acceptable cookie.
            # These cookies are written to a CSV/Spreadsheet, and we don't want
            # them interpreted as numbers.  So make sure we don't have any of them.
            # Reject cookies that start with 0
            if cookie[0] == '0':
                continue

            # Reject cookies that are all digits
            if cookie.isdigit():
                continue

            # Also reject cookies of the form \d+E\d+, because Excel/Google
            # Sheets will interpret that as a number, too.
            match = exp_regex.match(cookie)
            if match:
                continue

            # Finally, reject cookies that are already in the database.
            # Each cookie must be unique.
            query = 'SELECT cookie FROM cookies WHERE cookie=:cookie ORDER BY creation_timestamp DESC LIMIT 1'
            values = {
                'cookie' : cookie,
            }
            results = cookies.execute(query, values)
            row = results.fetchone()
            if not row:
                # Return when we found a cookie that is not yet in the
                # database
                return cookie

    #------------------------------------------------------------

    # We're basically basing the lookup on the Family DUID.  But the
    # Family DUIDs used by PS have a fairly dense distribution --
    # given one Family DUID, it's pretty easy to find other valid
    # Family DUIDs.  So instead, we generate a random 6-digit hex
    # number to use instead of the Family DUID.  Using only 6 digits
    # means that this code can be conveyed verbally, on a printed
    # sheet (e.g., a snail mail), etc.
    cookie = _get_cookie()
    if cookie is not None:
        log.debug(f"Using existing cookie: {cookie}")
    else:
        cookie = _generate_cookie()
        log.debug(f"Using new cookie: {cookie}")

    # Insert the URL in the cookies database
    url_escaped = url.replace("'", "''")
    query       = ("INSERT INTO cookies "
                   "(cookie,fduid,url,creation_timestamp) "
                   "VALUES (:cookie, :fduid, :url, :ts);")
    values      = {
        "cookie" : cookie,
        "fduid"  : fduid,
        "url"    : url_escaped,
        "ts"     : int(calendar.timegm(time.gmtime())),
    }
    cookies.execute(query, values)

    # Write the cookie to the DB immediately (in case someone is
    # reading/copying the sqlite3 database behind the scenes).
    cookies.connection.commit()

    if log:
        log.debug(f"SQL query: {query} / {values}")

    return f'{api_base_url}{cookie}', cookie

##############################################################################

# The ministry URL is comprised of three sections:
#
# 1. The base jotform URL
# 2. Some Family-specific field data
# 3. N sets of Member-specific field data
#
# This routine constructs #1 and #2.
def make_jotform_base_url(family, log):
    char = '?'
    url  = jotform.url
    for entry in jotform.pre_fill_data['family']:
        # Recall: the "fields" entry will be a single field for global data
        field  = entry['fields']
        value  = entry['value_func'](family)
        next   = f'{char}{field}={value}'
        url   += next
        char   = '&'

        if log:
            log.debug(f"Ministry base URL added: {next}")

    return url

# This routine constructs #3 (from above)
def make_ministries_url_portion(member, member_number, log):
    def _check(ministry, member):
        # member['py ministries'] is a dict() with the ministry name
        # as the key
        for member_ministry in member['py ministries']:
            if ministry == member_ministry:
                return True
        return False

    #-----------------------------------------------------------------------

    # First, add some per-Member non-minstry data
    url = ''
    for entry in jotform.pre_fill_data['per_member']:
        # The "fields" member will contain an array of fields; we
        # want the (member_number)th one.
        field = entry['fields'][member_number]
        value = entry['value_func'](member)
        next  = f'&{field}={value}'
        url  += next

        if log:
            log.debug(f"Ministry member URL portion: {next}")

    #-----------------------------------------------------------------------

    # Now add in all the per-Member ministry data
    for grid in jotform.ministry_grids:
        field = grid.member_fields[member_number]

        for row_num, row in enumerate(grid.rows):
            column      = COL_NOT_INVOLVED
            ps_ministry = row['ps_ministry']

            match = False
            if type(ps_ministry) is list:
                for m in ps_ministry:
                    match = _check(m, member)
                    if match:
                        break
            else:
                match = _check(ps_ministry, member)

            if match:
                column = COL_AM_INVOLVED

            # If relevant, pre-populate the correct column
            if column != None:
                url += f"&{field}[{row_num}][{column}]=true"
                # Example: parishLeadership[1][2]=true

    return url

#--------------------------------------------------------------------------

def send_family_email(message_body, family, submissions, cookies, smtp, log):
    # We won't get here unless there's at least one email address to
    # which to send.  But do a sanity check anyway.
    data = family['stewardship']
    if len(data['to_addresses']) == 0:
        return 0

    bounce_url    = data['bounce_url']
    text          = f'{stewardship_year} Epiphany Stewardship Renewal'
    ministry_link = f'<a href="{bounce_url}">{text}</a>'

    #---------------------------------------------------------------------

    smtp_to = ",".join(data['to_addresses'])

    # JMS DEBUG
    was = smtp_to
    smtp_to = "Jeff Squyres <jeff@squyres.com>"
    log.info(f"Sending to (OVERRIDE): {smtp_to} (was {was})")
    #---------------------------------------------------------------------

    names = ' / '.join(data['to_names'])
    log.info(f"    Sending to Family {names} at {smtp_to}")

    # Note: we can't use the normal string.format() to substitute in
    # values because the HTML/CSS includes a bunch of instances of {}.
    #message_body = message_initial()
    message_body = message_body.replace("{img}", email_image_url)
    message_body = message_body.replace("{family_names}",
                             f'{family["firstName"]} {family["lastName"]}')
    message_body = message_body.replace("{bounce_url}", ministry_link)
    message_body = message_body.replace("{family_code}", family['stewardship']['code'])

    # JMS kinda yukky that "args" is global
    global args
    if args.do_not_send:
        log.info("NOT SENDING EMAIL (--do-not-send)")
    else:
        ECC.send_email_existing_smtp(message_body, 'text/html',
                                     smtp_to, smtp_subject, smtp_from,
                                     smtp, log)

    return len(data['to_addresses'])

###########################################################################

def _send_family_emails(message_body, families, member_workgroups, submissions, cookies, log):
    # Send one email to the recipient in each family in the
    # dictionary.  Send them in Family FUID order so that if we get
    # interrupted and have to start again, we can do so
    # deterministically.

    email_sent = list()
    email_not_sent = list()

    # Open just one connection to the SMTP server
    # JMS Kinda gross that we're taking values from ECC globals
    with ECC.open_smtp_connection() as smtp:
        # Iterate through all the family emails that we need to send
        sorted_fduids = sorted(families)
        for i, fduid in enumerate(sorted_fduids):
            family = families[fduid]
            log.info(f"=== Family: {family['firstName']} {family['lastName']} ({i} of {len(sorted_fduids)})")

            family['stewardship'] = {
                'sent_email' : True,
                'reason not sent' : '',
                'to_addresses' : '',
                'to_names' : '',
                'code' : '',
                'bounce_url' : '',
            }

            # As of September 2024, Jotform cannot handle more than 7 Members'
            # worth of data in a single form (except with Chrome on a
            # laptop/desktop -- all other cases fail on the final submit).  So
            # if this Family has more than 7 Members, do not send to them.
            #
            # And since this Family will not be processed by Jotform,
            # we don't even need to generate a Family Jotform code for
            # them.  So just skip to the next Family.
            if len(family['py members']) > MAX_PS_FAMILY_MEMBER_NUM:
                family['stewardship']['sent_email'] = False
                family['stewardship']['reason not sent'] = 'Too many Members in Family'
                if log:
                    log.info(f"    *** Too many Members in Family ({len(family['py members'])} > {MAX_PS_FAMILY_MEMBER_NUM}) -- will not send")

                # This family will not be processed by Jotform.  So we skip
                # this family.
                continue

            #----------------------------------------------------------------

            members_by_mduid = { member['memberDUID'] : member for member in family['py members'] }
            to_names = { member['lastName'] : True for member in family['py members'] }

            # Scan through the Members and generate a list of names
            # and email addresses that we need.  Find the email
            # addresses of the Members of this Family who should
            # receive the Business Logistics Emails.
            to_addresses = ParishSoft.family_business_logistics_emails(family,
                                                                       member_workgroups,
                                                                       log)

            # Save the things we have computed to far on the family
            # (we use this data in making the ministry jotform base URL, below)
            family['stewardship']['to_addresses'] = to_addresses
            family['stewardship']['to_names'] = to_names

            #----------------------------------------------------------------

            # If we have no email addresses for the Family, skip emailing them.
            if len(to_addresses) == 0:
                family['stewardship']['sent_email'] = False
                family['stewardship']['reason not sent'] = 'Could not find relevant emails for Family'
                if log:
                    log.info(f"    *** Have no HoH/Spouse emails for Family {family['firstName']} {family['lastName']}")

                # Note that we fall through and still process this family, even
                # though we won't send them an email (because they might get
                # their Family Code some other way -- e.g., by calling the
                # office -- and do Jotform eStewardship that way).

            #----------------------------------------------------------------

            # Construct the base jotform URL with the Family-global data
            jotform_url = make_jotform_base_url(family, log)

            # Now add to the ministry jotform URL all the Member data
            # Since we might truncate the number of family members, do them in
            # a deterministic order.
            for member_number, mduid in enumerate(sorted(members_by_mduid)):
                member = members_by_mduid[mduid]

                # Add to the overall ministry URL with data for this Member
                jotform_url += make_ministries_url_portion(member,
                                                            member_number, log)

            # Now that we have the entire ministry jotform URL,
            # make a bounce URL for it
            bounce_url, cookie  = insert_url_cookie(fduid, jotform_url, cookies, log=log)
            family['stewardship']['bounce_url'] = bounce_url
            family['stewardship']['code'] = cookie

            #----------------------------------------------------------------

            # Check if we're going to send
            if not family['stewardship']['sent_email']:
                email_not_sent.append(family)
                continue

            #----------------------------------------------------------------

            send_count = send_family_email(message_body, family,
                                           submissions,
                                           cookies, smtp, log=log)
            email_sent.append(family)

    return email_sent, email_not_sent

###########################################################################

def send_all_family_emails(args, families, member_workgroups, family_workgroups,
                           submissions, cookies, log=None):
    return _send_family_emails(args.email_content, families, member_workgroups,
                               submissions, cookies, log)

#--------------------------------------------------------------------------

# JMS THIS NEEDS TO BE UPDATED FOR MDUIDS
def send_file_family_emails(args, families, member_workgroups, family_workgroups,
                            submissions, cookies, log=None):



    log.error("JMS This function needs to be updated for MDUIDs")
    log.error("JMS Exiting early")
    exit(1)



    some_families = dict()

    log.info("Reading Envelope ID file...")
    env_ids = list()
    with open(args.env_id_file, "r") as ef:
        lines = ef.readlines()
        for line in lines:
            env_ids.append(line.strip())
    log.info(env_ids)

    log.debug("Looking for Envelope IDs...")
    for fduid, f in families.items():
        fname = f'{f["firstName"]} {f["lastName"]}'
        env = f['ParKey'].strip()
        if env in env_ids:
            log.info(f"Found Envelope ID {env} in list (Family: {fname})")
            some_families[fduid] = f

    return _send_family_emails(args.email_content, some_families,
                               member_workgropus, submissions,
                               cookies, log)

#--------------------------------------------------------------------------

def _find_fam_wg(family_workgroups, name):
    for family_wg in family_workgroups.values():
        if family_wg['name'] == name:
            return family_wg
    return None

def send_submitted_family_emails(args, families, member_workgroups, family_workgroups,
                                 submissions, cookies, log=None):
    # Find all families that have a valid Jotform submission or are in
    # the Family Workgroup for this year's Stewardship.
    log.info(f"Looking for Families with a Jotform submission or are in the PS Workgroup '{stewardship_fam_cur_year_wg}'")

    family_wg = _find_fam_wg(family_workgroups,
                             stewardship_fam_cur_year_wg)

    some_families = dict()
    for fduid, f in families.items():
        want = False
        fname = f'{f["firstName"]} {f["lastName"]}'

        # Does the Family have a Jotform submission?
        if fduid in submissions:
            log.info(f"Family {fname} (Family FUID {fduid}) has an electronic submission")
            want = True

        # Is the Family in the designated WG?
        for entry in family_wg['membership']:
            if entry['familyId'] == fduid:
                want = True

        if want:
            log.info(f"Submitted family (Family FUID {fduid}): {fname}")
            some_families[fduid] = f
        else:
            log.info(f"Unsubmitted family (Family DUID {fduid}): {fname} -- skipping")

    return _send_family_emails(args.email_content, some_families,
                               member_workgroups, submissions,
                               cookies, log)

#--------------------------------------------------------------------------

def send_unsubmitted_family_emails(args, families, member_workgroups, family_workgroups,
                                   submissions, cookies, log=None):
    # Find all families that do not have a valid Jotform submission
    # and are not in the Family Workgroup for this year's Stewardship.
    log.info(f"Looking for Families without a Jotform submission and are not in the PS Workgroup '{stewardship_fam_cur_year_wg}'")

    family_wg = _find_fam_wg(family_workgroups,
                             stewardship_fam_cur_year_wg)

    some_families = dict()
    for fduid, f in families.items():
        fname = f'{f["firstName"]} {f["lastName"]}'
        want = False

        # If the Family has no Jotform submission, we want them!
        if fduid not in submissions:
            log.info(f"Family {fname} (Family DUID {fduid}) no electronic submission")
            want = True

        # If the Family is in the designated Family WG, then we don't
        # want them
        for entry in family_wg['membership']:
            if entry['familyId'] == fduid:
                want = False

        if want:
            log.info(f"Unsubmitted family (Family DUID {fduid}): {fname}")
            some_families[fduid] = f
        else:
            log.info(f"Completed family (Family DUID {fduid}): {fname} -- skipping")

    return _send_family_emails(args.email_content, some_families,
                               member_workgroups, submissions,
                               cookies, log)

#--------------------------------------------------------------------------

def send_some_family_emails(args, families, member_workgroups, family_workgroups,
                            submissions, cookies, log=None):
    target = args.email
    some_families = dict()

    log.info(f"Looking for email addresses: {target}")

    for fduid, f in families.items():
        found = False

        for m in f['py members']:
            emails = m['py emailAddresses']
            if emails is None:
                continue
            if any(email in target for email in emails):
                log.info(f"Found Member in Family with target email address: {emails}")
                found = True
                break

        if found:
            some_families[fduid] = f

    log.info(f"Found {len(some_families)} families with target email addresses")

    return _send_family_emails(args.email_content, some_families,
                               member_workgroups, submissions,
                               cookies, log)

###########################################################################

def cookiedb_create(filename, log=None):
    cur = cookiedb_open(filename)
    query = ('CREATE TABLE cookies ('
             'cookie text not null,'
             'fduid not null,'
             'url text not null,'
             'creation_timestamp integer not null'
             ')')

    cur.execute(query)

    if log:
        log.debug(f"Initialized cookie db: {filename}")

    return cur

def cookiedb_open(filename, log=None):
    conn = sqlite3.connect(filename)
    cur = conn.cursor()

    if log:
        log.debug(f"Opened cookie db: {filename}")

    return cur

###########################################################################

def write_email_csv(family_list, filename, extra, log):
    csv_family_fields = {
        "fduid"              : 'Family DUID',
        "household"          : 'Household names',
        'email'              : 'Email addresses',
    }

    csv_extra_family_fields = {
        'Code'                  : lambda fam: fam['stewardship']['code'],
        'Salulation'            : lambda fam: f"{fam['firstName']} {fam['lastName']}",
        'Street Address 1'      : lambda fam: fam['primaryAddress1'],
        'Street Address 2'      : lambda fam: fam['primaryAddress2'],
        'City/State'            : lambda fam: f"{fam['primaryCity']}, {fam['primaryState']}",
        'Zip Code'              : lambda fam: fam['primaryPostalCode'],
        'Send no mail'          : lambda fam: fam['sendNoMail'],
        'Num Family Members'    : lambda fam: len(fam['py members']),
        'Reason email not sent' : lambda fam: fam['stewardship']['reason not sent'],
    }

    csv_member_fields = {
        "mduid"              : 'Member DUID',
        "name"               : 'Name',
    }

    fieldnames = list()

    # Add the fields from both forms
    fieldnames.extend(csv_family_fields.values())
    if extra:
        fieldnames.extend(csv_extra_family_fields.keys())
    fieldnames.extend(csv_member_fields.values())

    csv_fieldnames = fieldnames.copy()

    with open(filename, 'w', newline='') as csvfile:
        writer = csv.DictWriter(csvfile, fieldnames=csv_fieldnames)
        writer.writeheader()

        for family in family_list:
            row = dict()

            for ff, column in csv_family_fields.items():
                log.debug(f"Looking for: {ff}")
                for item in jotform.pre_fill_data['family']:
                    if item['fields'] == ff:
                        func = item['value_func']
                        value = func(family)

                        # The pledge values will have HTML-ized "%24" instead of "$".
                        if isinstance(value, str):
                            value = value.replace("%24", "$")

                        row[column] = value

            if extra:
                for column, func in csv_extra_family_fields.items():
                    row[column] = func(family)

            for member in family['py members']:
                for mf, column in csv_member_fields.items():
                    log.debug(f"Looking for member item: {mf}")
                    for item in jotform.pre_fill_data['per_member']:
                        for field in item['fields']:
                            if field.startswith(mf):
                                func = item['value_func']
                                row[column] = func(member)

                log.debug(f"Writing row: {row}")
                writer.writerow(row)

    log.info(f"Wrote {filename}")

###########################################################################

def write_code_csv(families, filename, log):
    code_field = f'eStewardship {stewardship_year} code'
    fields = [
        'Family DUID',
        'Family name',
        code_field,
    ]

    with open(filename, 'w', newline='') as csvfile:
        writer = csv.DictWriter(csvfile, fieldnames=fields)
        writer.writeheader()

        for family in families.values():
            if 'stewardship' not in family:
                continue

            code = family['stewardship']['code']
            if len(code.strip()) == 0:
                # If this family will not be processed by Jotform for some
                # reason (e.g., too many members), there will be no code.  Skip
                # the family.
                continue

            row = {
                'Family DUID' : family['familyDUID'],
                'Family name' : f"{family['firstName']} {family['lastName']}",
                code_field : family['stewardship']['code'],
            }

            writer.writerow(row)

    log.info(f"Wrote {filename}")

###########################################################################

def setup_google(args, log):
    apis = {
        'drive' : { 'scope'       : Google.scopes['drive'],
                    'api_name'    : 'drive',
                    'api_version' : 'v3', },
    }
    services = GoogleAuth.service_oauth_login(apis,
                                              app_json=args.app_id,
                                              user_json=args.user_credentials,
                                              log=log)
    google = services['drive']

    #---------------------------------------------------------------------

    def _read_jotform_gsheet(google, gfile_id):
        response = google.files().export(fileId=gfile_id,
                                          mimeType=Google.mime_types['csv']).execute()

        # The ordering of these fields is critical, although the names are not
        fields = jotform_gsheet_columns['prelude']

        csvreader = csv.DictReader(response.decode('utf-8').splitlines(),
                                   fieldnames=fields)
        return csvreader

    #---------------------------------------------------------------------
    # All we need are quick-lookup dictionaries to know if a given
    # Family DUID has submitted or not.
    #
    # So convert the CSV data structures to a quick-lookup dictionary.
    def _convert(csv_data, log):
        output_data = dict()

        first = True
        for row in csv_data:
            # Skip the first / title row
            if first:
                first = False
                continue

            fduid = row['fduid']
            if fduid:
                fduid = int(fduid)
            else:
                fduid = 0
            output_data[fduid] = True

        return output_data

    log.info("Loading Jotform submissions Google sheet")
    jotform_csv = _read_jotform_gsheet(google, jotform_gsheet_gfile_id)
    submissions = _convert(jotform_csv, log=log)

    return submissions

###########################################################################

def process_pledge_data(filename, families, log):
    def _safe_float(s):
        if s is None:
            return 0
        s = str(s).strip()
        if len(s) == 0:
            return 0
        # Remove any currency symbols or commas
        s = s.replace('$', '').replace(',', '')
        try:
            return float(s)
        except ValueError:
            return 0

    # Build a mapping from Envelope IDs to Family DUIDs, since
    # the PS pledge report is indexed by Envelope ID.
    envelope_to_fduid = {}
    for fduid, family in families.items():
        envelope_id = family.get('envelopeNumber')
        if envelope_id is not None:
            envelope_to_fduid[envelope_id] = fduid

    key = 'calculated'
    log.info(f"Reading pledge data from Excel file {filename}")

    # Load the workbook
    workbook = openpyxl.load_workbook(filename, read_only=True)

    for sheet_name in workbook.sheetnames:
        log.debug(f"Processing worksheet: {sheet_name}")
        worksheet = workbook[sheet_name]

        envelope_number = None
        pledged_amount = None
        posted_amount = None

        # Search for envelope number, pledged amount, and posted amount
        for row in worksheet.iter_rows(values_only=True):
            for col_idx, cell_value in enumerate(row):
                if cell_value is None:
                    continue

                cell_str = str(cell_value).strip()

                # Look for "Envelope Number: xxxx" in a single cell
                envelope_match = re.search(r'Envelope Number:\s*(\d+)', cell_str, re.IGNORECASE)
                if envelope_match:
                    envelope_number = int(envelope_match.group(1))
                    log.debug(f"Found envelope number: {envelope_number}")
                    continue

                # Look for "Pledged" with amount 2 cells to the right
                if cell_str.lower() == 'pledged':
                    val = row[col_idx + 2]
                    if val and type(val) is str and val.strip().lower() == 'adjustments':
                        # This isn't the right "pledged" cell -- skip it
                        continue
                    pledged_amount = _safe_float(val)
                    log.debug(f"Found pledged amount: {pledged_amount}")
                    continue

                # Look for "Posted" with amount 2 cells to the right
                if cell_str.lower() == 'posted' and col_idx + 2 < len(row):
                    posted_amount = _safe_float(row[col_idx + 2])
                    log.debug(f"Found posted amount: {posted_amount}")
                    continue

        # If we found all required data, store it
        if envelope_number is not None and pledged_amount is not None and posted_amount is not None:
            if envelope_number in envelope_to_fduid:
                fduid = envelope_to_fduid[envelope_number]
                families[fduid][key] = {
                    'pledged': pledged_amount,
                    'gifts': posted_amount,
                }
                log.debug(f"Stored data for family {fduid}: pledged={pledged_amount}, gifts={posted_amount}")
            else:
                log.warning(f"Envelope number {envelope_number} not found in families data")
        else:
            log.warning(f"Incomplete data found in worksheet {sheet_name}: "
                       f"envelope={envelope_number}, pledged={pledged_amount}, posted={posted_amount}")

    workbook.close()
    log.info(f"Finished processing Excel file {filename}")

###########################################################################

def setup_args():
    # These options control which emails are sent.
    # You can only use one of these options at a time.
    group = tools.argparser.add_mutually_exclusive_group()
    group.add_argument('--all', action='store_true',
                        help='Send all emails')
    group.add_argument('--email',
                        action='append',
                        help='Send only to this email address')
    # JMS THIS NEEDS TO BE UPDATED FOR MDUIDS / FDUIDS
    group.add_argument('--env-id-file',
                        help='Send only to families with envelope IDs in the file')
    group.add_argument('--unsubmitted',
                        action='store_true',
                        help='Send only to families who have not yet submitted all their data (i.e., all Members have submitted ministry, and a Family pledge has been recorded)')

    group.add_argument('--submitted',
                        action='store_true',
                        help='Send only to families who have already submitted all their data')

    # JMS: Get an XLS report from ParishSoft ledgers: From the Home
    # page select Offering--Reports (click on Pledge)--Select
    # Stewardship contributions then pledge details--Input dates to
    # track-- export and select Excel instead of PDF.
    #
    # Then load that XLS file in Excel and save it as XLSX.  The
    # resulting XLSX file can be read here.
    #
    # ALTERNATIVELY: I think all this pledge / contribution data might
    # be available via the ParishSoft API (see
    # https://ps-fs-external-api-prod.azurewebsites.net/index.html),
    # but we're not [yet] downloading it as part of ParishSoftv2.py.
    # It could be extended to obtain that data (maybe only load it via
    # an optional parameter to load_families_and_members()?) since we
    # don't need it often, and it'll likely be a bunch of additional
    # API calls).
    tools.argparser.add_argument('--pledge-data',
                                 required=True,
                                 help='XLSX file from ParishSoft containing previous year pledge info')

    tools.argparser.add_argument('--do-not-send',
                                action='store_true',
                                help='If specified, do not actually send any emails, but do everything else')
    tools.argparser.add_argument('--email-content',
                                required=True,
                                help='File containing the templated content of the email to be sent')

    tools.argparser.add_argument('--smtp-auth-file',
                                 required=True,
                                 help='File containing SMTP AUTH username:password')

    # These options are for Google Authentication
    global gapp_id
    tools.argparser.add_argument('--app-id',
                                 default=gapp_id,
                                 help='Filename containing Google application credentials.  Only necessary if sending an email that contains a {*_reminder} tag.')
    global guser_cred_file
    tools.argparser.add_argument('--user-credentials',
                                 default=guser_cred_file,
                                 help='Filename containing Google user credentials.  Only necessary if sending an email that contains a {*_reminder} tag.')

    # ParishSoft args
    tools.argparser.add_argument('--ps-api-keyfile',
                                 default='parishsoft-api-key.txt',
                                 help='File containing the ParishSoft API key')
    tools.argparser.add_argument('--ps-cache-dir',
                                 default='ps-data',
                                 help='Directory to cache the ParishSoft data')

    tools.argparser.add_argument('--debug', action='store_true',
                                 help='Enable additional debug logging')

    # These options control the generated cookie database
    tools.argparser.add_argument('--cookie-db', required=True,
                        help='Name of the SQLite3 database to output to')
    tools.argparser.add_argument('--append', action='store_true',
                        help='If specified, append to the existing SQLite3 database')

    args = tools.argparser.parse_args()

    # If a database already exists and --append was not specified, barf
    if os.path.exists(args.cookie_db) and not args.append:
        print(f"Error: database {args.cookie_db} already exists and --append was not specified")
        print("Cowardly refusing to do anything")
        exit(1)

    # Need either --all or --email or --env-id-file or --submitted or
    # --ubsibmitted
    if (not args.all and not args.email and
        not args.env_id_file and not args.submitted
        and not args.unsubmitted):
        print("Error: must specify either --all, --email, --env-id-file, --submitted, or --unsubmitted")
        print("Cowardly refusing to do anything")
        exit(1)

    # Helper to check if a path exists
    def _check_path(filename):
        if not os.path.exists(filename):
            print(f"Error: file '{filename}' is not readable")
            exit(1)

    # Check to make sure the env ID file is valid
    if args.env_id_file:
        _check_path(args.env_id_file)

    # Check to make sure pledge data file is valid
    _check_path(args.pledge_data)

    # Check to make sure the email content file is valid
    _check_path(args.email_content)

    # Read the email content
    with open(args.email_content, 'r') as f:
        args.email_content = f.read()

    # If the email content contains any of the "_reminder" tags, then check the
    # Google arguments, too.
    need_google = False
    if (args.submitted or args.unsubmitted or
        re.search("_reminder[s]*}", args.email_content)):
        need_google = True
        if (args.app_id is None or
            args.user_credentials is None):
            print("Error: email content file contains a *_reminder template, but not all of --app-id, --user-credentials were specified")
            exit(1)

        _check_path(args.app_id)
        _check_path(args.user_credentials)

    # Read the PS API key
    if not os.path.exists(args.ps_api_keyfile):
        print(f"ERROR: ParishSoft API keyfile does not exist: {args.ps_api_keyfile}")
        exit(1)
    with open(args.ps_api_keyfile) as fp:
        args.api_key = fp.read().strip()

    return args, need_google

###########################################################################

def main():
    global families, members

    # JMS Kinda yukky that "args" is global
    global args
    args, need_google = setup_args()

    logfilename = 'log.txt'
    try:
        os.unlink(logfilename)
    except:
        pass
    log = ECC.setup_logging(debug=args.debug, logfile=logfilename)

    # We need Google if the email content contains a *_reminder template
    # (i.e., we need to login to Google and download some data)
    if need_google:
        # Submissions is a dictionary indexed by Family DUID
        submissions = setup_google(args, log=log)
    else:
        submissions = dict()

    # Read in all the ParishSoft data
    log.info("Loading ParishSoft data...")
    families, members, family_workgroups, member_workgroups, ministries = \
        ParishSoft.load_families_and_members(api_key=args.api_key,
                                             cache_dir=args.ps_cache_dir,
                                             active_only=True,
                                             parishioners_only=True,
                                             log=log)
    log.info(f"Loaded {len(families)} ParishSoft Families")
    log.info(f"Loaded {len(members)} ParishSoft Members")

    # Open the cookies DB
    if not args.append or not os.path.exists(args.cookie_db):
        func = cookiedb_create
    else:
        func = cookiedb_open
    cookies = func(args.cookie_db)

    # Load pledge data
    process_pledge_data(args.pledge_data, families, log)

    # Setup the SMTP connection parameters
    ECC.setup_email(args.smtp_auth_file, log=log)

    # Send the desired emails
    if args.all:
        func = send_all_family_emails
    elif args.env_id_file:
        func = send_file_family_emails
    elif args.submitted:
        func = send_submitted_family_emails
    elif args.unsubmitted:
        func = send_unsubmitted_family_emails
    else:
        func = send_some_family_emails

    sent, not_sent = func(args, families, member_workgroups, family_workgroups,
                          submissions, cookies, log)

    # Record who/what we sent
    ts = datetime.datetime.now().strftime('%Y-%m-%d-%H%M%S')
    write_email_csv(sent,     f'emails-sent-{ts}.csv',     extra=True, log=log)
    write_email_csv(not_sent, f'emails-not-sent-{ts}.csv', extra=True, log=log)
    write_code_csv(families,  f'family-codes-{ts}.csv',    log)

    # Close the databases
    cookies.connection.close()

# Need to make these global for the lambda functions
members = 1
families = 1

main()
