#!/usr/bin/env python3

# Make sure to pip install everything in requirements.txt.

import os
import sys

# We assume that there is a "ecc-python-modules" sym link in this
# directory that points to the directory with ECC.py and friends.
moddir = os.path.join(os.getcwd(), 'ecc-python-modules')
if not os.path.exists(moddir):
    print("ERROR: Could not find the ecc-python-modules directory.")
    print("ERROR: Please make a ecc-python-modules sym link and run again.")
    exit(1)
# On MS Windows, git checks out sym links as a file with a single-line
# string containing the name of the file that the sym link points to.
if os.path.isfile(moddir):
    with open(moddir) as fp:
        dir = fp.readlines()
    moddir = os.path.join(os.getcwd(), dir[0])

sys.path.insert(0, moddir)

import csv
import argparse

import ECC
import Google
import GoogleAuth
import ParishSoftv2 as ParishSoft

from google.api_core import retry

import constants

from datetime import date

from collections import defaultdict

from pprint import pprint, pformat

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

start_of_interest = date(month=1, day=1, year=constants.stewardship_year - 1)

##############################################################################

gapp_id         = 'client_id.json'
guser_cred_file = 'user-credentials.json'

##############################################################################

def write_xlsx(ministry_name, interested, log):
    # Make a worksheet
    filename = f"{ministry_name}-interested.xlsx"
    filename = filename.replace('/', '-')
    log.debug(f"Writing {filename}")

    wb = Workbook()
    ws = wb.active

    #---------------------------------------------------------------------

    # Set narrow margins so that we can get as much in as possible if people print it out
    ws.page_margins.left  = 0.4
    ws.page_margins.right = ws.page_margins.left

    ws.page_setup.orientation = ws.ORIENTATION_LANDSCAPE

    #---------------------------------------------------------------------

    title_font    = Font(name='Arial', size=12, bold=True)
    thin_side     = Side(style='thin')
    bottom_border = Border(bottom=thin_side)

    def _make_title(cell, value):
        ws[cell]        = value
        ws[cell].font   = title_font
        ws[cell].border = bottom_border

    _make_title('A1', 'Ministry')
    _make_title('B1', '')
    _make_title('C1', ministry_name)
    _make_title('A2', 'Ministry Chair')
    _make_title('B2', '')
    _make_title('C2', '')

    #---------------------------------------------------------------------

    outcome_font  = Font(name='Arial', size=12, bold=True)
    outcome_align = Alignment(horizontal='center', wrap_text=True)
    box_border    = Border(top=thin_side, bottom=thin_side, left=thin_side, right=thin_side)

    row = 3
    ws.merge_cells(f'G{row}:I{row}')
    cell               = f'G{row}'
    ws[cell]           = 'OUTCOME'
    ws[cell].font      = outcome_font
    ws[cell].alignment = outcome_align
    ws[cell].border    = box_border

    # The "Ouycom" box is merged, so we have to set the top cell border on all the cells
    ws[f'H{row}'].border = Border(top=thin_side, bottom=thin_side)
    ws[f'I{row}'].border = Border(top=thin_side, bottom=thin_side, right=thin_side)

    #---------------------------------------------------------------------

    def _make_heading(cell, value, width, fill=None):
        ws[cell]           = value
        ws[cell].font      = heading_font
        ws[cell].alignment = heading_align
        if fill:
            ws[cell].fill  = fill
        ws[cell].border    = box_border

        ws.column_dimensions[cell[0]].width = width

    heading_font  = Font(name='Arial', size=10)
    heading_align = outcome_align

    join_fill     = PatternFill(fgColor='D9E9D4', fill_type='solid')
    not_int_fill  = PatternFill(fgColor='FFF1CE', fill_type='solid')
    no_resp_fill  = PatternFill(fgColor='FEC7CE', fill_type='solid')

    row += 1
    _make_heading(f'A{row}', 'Name', 20)
    _make_heading(f'B{row}', 'Family DUID', 9)
    _make_heading(f'C{row}', 'Email addresses', 35)
    _make_heading(f'D{row}', 'Date of email', 10)
    _make_heading(f'E{row}', 'Phone numbers', 25)
    _make_heading(f'F{row}', 'Date of phone call', 10)
    join_col = 'G'
    _make_heading(f'{join_col}{row}', 'Join ministry', 10, join_fill)
    not_int_col = 'H'
    _make_heading(f'{not_int_col}{row}', 'No longer interested', 10, not_int_fill)
    no_resp_col = 'I'
    _make_heading(f'{no_resp_col}{row}', 'No response / followup', 10, no_resp_fill)

    #---------------------------------------------------------------------

    ws.freeze_panes = ws[f'A{row+1}']

    #---------------------------------------------------------------------

    def _set(cell, value):
        ws[cell]           = value
        ws[cell].font      = value_font
        ws[cell].alignment = value_align
        ws[cell].border    = box_border

    def _set_color(cell, fill):
        ws[cell].fill      = fill
        ws[cell].alignment = outcome_align

    def _find_phones(entity, phones):
        key = 'phones'
        if key not in entity:
            return

        for phone in entity[key]:
            if 'emergency' not in phone['type'].lower():
                item = {
                    'number' : phone['number'],
                    'string' : phone['number'],
                    'type'   : '',
                }
                if phone['type']:
                    item['string'] += f" ({phone['type']})"
                    item['type']    = phone['type']

                phones[phone['number']] = item

    value_font  = Font(name='Arial', size=10)
    value_align = Alignment(wrap_text=True)

    for member_name in sorted(interested):
        member = interested[member_name]
        log.debug(f"- Member: {member['py friendly name FL']} (MDUID: {member['memberDUID']})")

        row += 1

        _set(f'A{row}', member['py friendly name FL'])
        _set(f'B{row}', member['py family']['familyDUID'])

        value = ''
        num_emails = 0
        key = 'emailAddress'
        if key in member and member[key]:
            value = member[key]
        _set(f'C{row}', value)

        # Take non-Emergency phone numbers
        phones = dict()
        phones = ParishSoft.get_member_public_phones(member)
        #_find_phones(member['py family'], phones)
        #_find_phones(member, phones)
        num_phones = 0
        if len(phones) > 0:
            value = ''
            for phone in phones:
                if value:
                    value += '\r\n'
                value += f'{phone["number"]} {phone["type"]}'
                num_phones += 1
            _set(f'E{row}', value)

        # Add cell borders for cells that have no values
        for column in ['D', 'F', 'G', 'H', 'I']:
            ws[f"{column}{row}"].border = box_border

        # Set colors
        _set_color(f'{join_col}{row}', join_fill)
        _set_color(f'{not_int_col}{row}', not_int_fill)
        _set_color(f'{no_resp_col}{row}', no_resp_fill)

        # Set the row height
        num_lines = num_emails if num_emails > num_phones else num_phones
        num_lines = 1 if num_emails == 0 else num_lines
        actual_height = 15 * num_lines
        if actual_height < 30:
            ws.row_dimensions[row].height = 30

    #---------------------------------------------------------------------

    wb.save(filename)
    log.info(f'Wrote {filename}')

##############################################################################

def write_all_interested(all_interested, log):
    for ministry_name in sorted(all_interested):
        write_xlsx(ministry_name, all_interested[ministry_name], log)

##############################################################################

def read_staff_gsheets(google, folder_id, ps_ministries, log):

    #---------------------------------------------------------------------

    @retry.Retry(predicate=Google.retry_errors)
    def _find_staff_google_sheets(google, folder_id, log):
        log.info(f"Loading Staff Google Sheets from folder {folder_id}")
        mime = Google.mime_types['sheet']
        q = f'"{folder_id}" in parents and mimeType="{mime}" and trashed=false'
        response = google.files().list(corpora='drive',
                                       driveId=constants.google_shared_drive_id,
                                       includeTeamDriveItems=True,
                                       supportsAllDrives=True,
                                       q=q,
                                       fields='files(name,id)').execute()
        sortable_gfiles = {
            gfile['name'] : gfile for gfile in response.get('files', [])
        }

        return sortable_gfiles

    #---------------------------------------------------------------------

    def _find_ministry(name, ps_ministries, log):
        try:
            num = int(name[:3])
        except Exception as e:
            return None, None

        for grid in constants._all_ministry_grids:
            found = False
            for row in grid.rows:
                row_num = int(row['ps_ministry'][:3])
                if row_num != num:
                    continue

                # Get the official PS ministry name.
                ps_ministry_name = row['ps_ministry']

                # Look up that name in the list of PS ministries.
                for mduid, ministry in ps_ministries.items():
                    if ministry['name'] == ps_ministry_name:
                        return num, ministry

        # If we get here, we didn't find it
        return None, None

    #---------------------------------------------------------------------

    def _load_all_gsheets(google, folder_id, gfiles, ps_ministries, log):

        # We make a bunch of Google API calls; sometimes we get a
        # "rate exceeded" type of error.  Wrap up this API call in a
        # retry so that it won't fail automatically.
        @retry.Retry(predicate=Google.retry_errors)
        def _export(google, id, mimetype):
            return google.files().export(fileId=id,
                                         mimeType=mimetype).execute()

        data = defaultdict(list)
        for name in sorted(gfiles):
            gfile = gfiles[name]
            log.debug(f"Loading {name} ({gfile['id']})")

            # Use the filename to look up the corresponding ParishSoft
            # ministry
            ministry_num, ps_ministry = _find_ministry(name,
                                                       ps_ministries,
                                                       log)
            if ps_ministry is None:
                # If we errored, skip this file
                log.error(f"File {name} does not parse into ministry name -- SKIPPING")
                continue

            log.debug("Found corresponding PS ministry")

            # Download the Gsheet as a CSV (so that we can parse it)
            response = _export(google, gfile['id'], Google.mime_types['csv'])

            # Parse the Gsheet into actual data
            # The first row is the field names
            lines = response.decode('utf-8').splitlines()
            fields = lines[0].split(',')
            csvreader = csv.DictReader(lines[1:], fieldnames=fields)
            for row in csvreader:
                mduid = row['Mem DUID']
                if mduid:
                    mduid = int(mduid)
                else:
                    mduid = 0

                # This value is supplied by previous python, but may
                # be altered by Staff members.  So make it all lower
                # case in an attempt to normalize the values.
                category = row['Category'].lower()
                current = row['Current ministry status'].lower()

                # Save this result in our output data
                data[ps_ministry['name']].append({
                    'mduid' : mduid,
                    'ps ministry' : ps_ministry,
                    'category' : category,
                    'current' : current,
                })

        log.debug("Loaded all staff GSheets")
        return data

    #---------------------------------------------------------------------

    gfiles = _find_staff_google_sheets(google, folder_id, log)
    data = _load_all_gsheets(google, folder_id, gfiles, ps_ministries, log)

    return data

##############################################################################

def find_interested(members, google, sheets_folder_id, ps_ministries, log):
    interested = defaultdict(dict)
    discrepancies = defaultdict(dict)

    # Returns a dictionary:
    # ministry_name : { mduid, ps_ministry, category }
    gsheets = read_staff_gsheets(google, sheets_folder_id, ps_ministries, log)

    key = 'py friendly name FL'
    sorted_keys = sorted(gsheets.keys())
    for ministry_name in sorted_keys:
        log.info(f"Examining results from ministry: {ministry_name}")

        # Make: out[ministry_name][member name] = member
        for record in gsheets[ministry_name]:
            mduid = record['mduid']
            if mduid not in members:
                log.error(f"Could not find MDUID {mduid} in current list of members -- skipping")
                continue

            member = members[mduid]
            name = member[key]

            # Record interested members

            if record['category'] == 'interested':
                log.info(f"  Interested: {name}")
                interested[ministry_name][name] = member

            # Look for weird discrepancies
            if record['category'] != record['current']:
                discrepancies[ministry_name][name] = member

    return interested, discrepancies

##############################################################################

def setup_google(args, log):
    apis = {
        'drive' : { 'scope'       : Google.scopes['drive'],
                    'api_name'    : 'drive',
                    'api_version' : 'v3', },
    }
    services = GoogleAuth.service_oauth_login(apis,
                                              app_json=args.app_id,
                                              user_json=args.user_credentials,
                                              log=log)
    google = services['drive']
    return google

##############################################################################

def setup_args():
    parser = argparse.ArgumentParser()
    parser.add_argument('--debug',
                                action='store_true')

    # ParishSoft args
    parser.add_argument('--ps-api-keyfile',
                        default='parishsoft-api-key.txt',
                        help='File containing the ParishSoft API key')
    parser.add_argument('--ps-cache-dir',
                        default='ps-data',
                        help='Directory to cache the ParishSoft data')

    # Google args
    global gapp_id
    parser.add_argument('--app-id',
                        default=gapp_id,
                        help='Filename containing Google application credentials')
    global guser_cred_file
    parser.add_argument('--user-credentials',
                        default=guser_cred_file,
                        help='Filename containing Google user credentials')

    parser.add_argument('--sheets-folder',
                        required=True,
                        help='Google folder ID containing Sheets to download + analyze')

    args = parser.parse_args()

    # Read the PS API key
    if not os.path.exists(args.ps_api_keyfile):
        print(f"ERROR: ParishSoft API keyfile does not exist: {args.ps_api_keyfile}")
        exit(1)
    with open(args.ps_api_keyfile) as fp:
        args.api_key = fp.read().strip()

    return args

##############################################################################

def main():
    args = setup_args()
    log = ECC.setup_logging(debug=args.debug)

    #---------------------------------------------------------------

    log.info("Loading ParishSoft data...")
    families, members, family_workgroups, member_workgroups, ministries = \
        ParishSoft.load_families_and_members(api_key=args.api_key,
                                             cache_dir=args.ps_cache_dir,
                                             active_only=True,
                                             parishioners_only=True,
                                             log=log)
    log.info(f"Loaded {len(families)} ParishSoft Families")
    log.info(f"Loaded {len(members)} ParishSoft Members")

    #---------------------------------------------------------------

    google = setup_google(args, log)

    #---------------------------------------------------------------

    interested, discrepancies = find_interested(members, google, args.sheets_folder,
                                 ministries, log)
    write_all_interested(interested, log)




    # JMS Need to write out discrepancies





main()
