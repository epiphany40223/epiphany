#!/usr/bin/env python3

# Make sure to pip install everything in requirements.txt.

import sys
import argparse
import os
import copy

# We assume that there is a "ecc-python-modules" sym link in this
# directory that points to the directory with ECC.py and friends.
moddir = os.path.join(os.getcwd(), 'ecc-python-modules')
if not os.path.exists(moddir):
    print("ERROR: Could not find the ecc-python-modules directory.")
    print("ERROR: Please make a ecc-python-modules sym link and run again.")
    exit(1)

sys.path.insert(0, moddir)

import ECC
import PDSChurch

import helpers

from datetime import date
from datetime import datetime
from datetime import timedelta

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment

#------------------------------------------------------------------------------

def compare_pds_pledges(second_year, pds_families, log):
    def _compare():
        log.info(f"Comparing CY{second_year-1} to CY{second_year}")

        y1 = f'calculated {second_year - 1}'
        y2 = f'calculated {second_year}'

        out = dict()
        for family in pds_families.values():
            previous_pledge = 0
            if y1 in family:
                previous_pledge = family[y1]['pledged']

            current_pledge = 0
            if y2 in family:
                current_pledge = family[y2]['pledged']

            if previous_pledge == 0:
                if current_pledge == 0:
                    category = "No pledge both years"
                else:
                    category = "NEW pledge"

            else:
                if current_pledge < previous_pledge:
                    category = "Reduced pledge"
                elif current_pledge == previous_pledge:
                    category = "No change"
                elif current_pledge > previous_pledge:
                    category = "Increased pledge"

            dollar_impact = current_pledge - previous_pledge

            if category not in out:
                out[category] = {
                    'households'       : 0,
                    'dollar impact'    : 0,
                    'total of pledges' : 0,
                }

            out[category]["households"]       += 1
            out[category]["dollar impact"]    += dollar_impact
            out[category]["total of pledges"] += current_pledge

        return out

    # ------------------------------------------------------------------------

    def _make_xlsx(comparison):
        workbook = Workbook()
        sheet = workbook.active

        comments_label    = "Comments"
        pledge_last_label = f'CY{second_year-1} pledge'
        pledge_cur_label  = f'CY{second_year} pledge'
        amount_label      = f'CY{second_year-1} gifts'

        # Setup the title rows
        # Title rows + set column widths
        title_font   = Font(color='FFFF00')
        title_fill   = PatternFill(fgColor='0000FF', fill_type='solid')
        title_align  = Alignment(horizontal='center', wrap_text=True)

        wrap_align   = Alignment(horizontal='general', wrap_text=True)
        right_align  = Alignment(horizontal='right')

        money_format = "$##,###,###,###"
        percentage_format = "##.#"

        xlsx_cols = dict();
        def _add_col(name, width=15, format=None):
            col             = len(xlsx_cols) + 1
            xlsx_cols[name] = {'name' : name, 'format' : format,
                               'column' : col, 'width' : width }

        _add_col('Category', width=20)
        _add_col('Number of Households')
        _add_col('%-age of Total Submitted Households')
        _add_col('Dollar Impact')
        _add_col('Total of Pledges Submitted')
        _add_col('%-age of Total Pledges Submitted')

        # Make 2 rows of merged cells for wide titles
        def _make_merged_title_row(row, value):
            cell           = sheet.cell(row=row, column=1, value=value)
            cell.fill      = title_fill
            cell.font      = title_font
            cell.alignment = title_align
            end_col_char   = chr(ord('A') - 1 + len(xlsx_cols))
            sheet.merge_cells(f'A{row}:{end_col_char}{row}')

        _make_merged_title_row(row=1, value='eStewardship Pledge Analysis')
        _make_merged_title_row(row=2, value='')

        # Now add all the column titles
        for data in xlsx_cols.values():
            col            = data['column']
            cell           = sheet.cell(row=3, column=col, value=data['name'])
            cell.fill      = title_fill
            cell.font      = title_font
            cell.alignment = title_align

            col_char = chr(ord('A') - 1 + col)
            sheet.column_dimensions[col_char].width = data['width']

        # Finally, fill in all the data rows.
        # First, compute totals so that we can compute percentages.
        total_households = 0
        total_pledges    = 0
        total_impact     = 0
        for data in comparison.values():
            total_households += data['households']
            total_pledges    += data['total of pledges']
            total_impact     += data['dollar impact']

        def _fill(column, value, align=None, format=None):
            cell = sheet.cell(row=xlsx_row, column=column, value=value)

            want_format = True
            if (type(value) is int or type(value) is float) and value == 0:
                    want_format = False

            if want_format:
                if align:
                    cell.alignment = align
                if format:
                    cell.number_format = format

        xlsx_row = 4
        for category in sorted(comparison.keys()):
            data = comparison[category]
            _fill(1, category)
            _fill(2, data['households'])
            _fill(3, data['households'] / total_households * 100.0,
                format=percentage_format)
            _fill(4, data['dollar impact'],
                format=money_format)
            _fill(5, data['total of pledges'],
                format=money_format)

            value = 0
            if total_pledges > 0:
                value = data['total of pledges'] / total_pledges * 100.0
            _fill(6, value, format=percentage_format)

            xlsx_row += 1

        _fill(1, 'Totals', align=right_align)
        _fill(2, total_households)
        _fill(4, total_impact, format=money_format)
        _fill(5, total_pledges, format=money_format)

        return workbook

    # ------------------------------------------------------------------------

    # Do the comparison
    comparison = _compare()

    # Make an XLSX report
    workbook   = _make_xlsx(comparison)
    filename   = f'Pledge comparison CY{second_year-1}-CY{second_year}.xlsx'
    workbook.save(filename)
    log.info(f"Saved: {filename}")

###############################################################################

def main():
    global families, members

    log = ECC.setup_logging(debug=False)










    # JMS Update for PS
    # JMS Compare to XLSX that Doug provided?

    log.info("Reading PDS data...")
    (pds, pds_families,
     pds_members) = PDSChurch.load_families_and_members(filename='pdschurch.sqlite3',
                                                        parishioners_only=True,
                                                        log=log)

    first_year = 2020
    last_year  = 2022
    for year in range(first_year - 1, last_year + 1):
        log.info(f"Calculating family data for year {year}")
        for family in pds_families.values():
            helpers.calculate_family_values(family, year, log)
            family[f'calculated {year}'] = family['calculated']
            del family['calculated']

    for year in range(first_year, last_year + 1):
        compare_pds_pledges(year, pds_families, log)

main()
