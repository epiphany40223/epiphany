#!/usr/bin/env python3

import os
import re
import sys
import csv
import copy
import datetime
import argparse
import openpyxl

# Load the ECC python modules.  There will be a sym link off this directory.
moddir = os.path.join(os.path.dirname(sys.argv[0]), 'ecc-python-modules')
if not os.path.exists(moddir):
    print("ERROR: Could not find the ecc-python-modules directory.")
    print("ERROR: Please make a ecc-python-modules sym link and run again.")
    exit(1)

sys.path.insert(0, moddir)

import ECC
import PDSChurch
import Google
import GoogleAuth

##############################################################################

import helpers
import constants

from oauth2client import tools

email_keys = [PDSChurch.pkey, PDSChurch.npkey]

###########################################################################

def env_id(family):
    return "'" + family['ParKey'].strip()

#--------------------------------------------------------------------------

# Strip all whitespace, (, ), -, etc. from a phone number so that we're
# left with just a series of digits (i.e,. easier to compare).
def reduce_phone_to_digits(phone):
    phone = re.sub(r'[\s\(\)\-]+', '', phone)
    return phone

#--------------------------------------------------------------------------

def _jotform_boolean(value):
    if value is None:
        return False
    return True if value.lower() == 'yes' else False

###########################################################################

# This routine gets the "complete" submissions, meaning that the
# dictionary is indexed by the Jotform fieldnames (not our "simple" fieldnames).
def comments(submissions, families, log):
    num_comments = 0
    filename = 'census2021-comments.csv'
    with open(filename, 'w') as fp:
        fields = ['FID', 'Env ID', 'Family name', 'Comment']
        writer = csv.DictWriter(fp, fieldnames=fields)
        writer.writeheader()

        for fid in sorted(submissions):
            row = submissions[fid]
            comment = row['Do you have any questions, comments, or other information for us?']
            if comment == '':
                continue

            if fid in families:
                family = families[fid]
                env  = env_id(family)
                name = family['Name']
            else:
                env  = '???'
                name = f"Not a parishioner: {row['Household name']}"

            item = {
                'FID'         : fid,
                'Env ID'      : env,
                'Family name' : name,
                'Comment'     : comment,
            }

            writer.writerow(item)
            num_comments += 1

    log.info(f"Found {num_comments} comments")
    log.info(f"Wrote {filename}")

###########################################################################

def member_attendance(member_submissions, members, log):
    official_answers = {
        'In-person (100%)' : False,
        'Online / Virtual (100%)' : False,
        'Hybrid (combination of in-person and online)' : False,
        'None of these - I am not interested in attending': False,
    }
    all_answers = dict()

    responses = dict()
    for mid, submission in member_submissions.items():
        key = 'attendance'
        if key not in submission:
            continue
        if submission[key] is None:
            continue

        # This is a multi-value answer (like the survey results)
        responses[mid] = dict()
        answers = submission[key].splitlines()
        for answer in answers:
            responses[mid][answer] = True
            all_answers[answer] = True

    num_results = 0
    filename = f'census2021-member-attendance.csv'
    with open(filename, 'w') as fp:
        fields = ['MID', 'Name']
        # Put the official answers first (before write-in/"other" answers)
        for answer in official_answers:
            fields.append(answer)
        for answer in all_answers:
            if answer not in official_answers:
                fields.append(answer)

        writer = csv.DictWriter(fp, fieldnames=fields)
        writer.writeheader()

        for mid, results in responses.items():
            member = members[mid]

            item = {
                'MID'  : mid,
                'Name' : member['Name'],
            }
            for answer in results:
                item[answer] = True
            writer.writerow(item)
            num_results += 1

    log.info(f"Found {num_results} responses to the per-member attendance question")
    log.info(f"Wrote {filename}")

    num_results = 0
    filename = f'census2021-member-attendance-other-values.txt'
    with open(filename, 'w') as fp:
        for answer in all_answers:
            if answer not in official_answers:
                fp.write(f"{answer}\n")
                num_results += 1

    log.info(f"Found {num_results} 'other' responses to the per-member attendance question")
    log.info(f"Wrote {filename}")

###########################################################################

def _simple_pds_jotform_member_compare(member_submissions, members, jotform_key, pds_key, log):
    results = dict()

    for mid, submission in member_submissions.items():
        member = members[mid]

        jotform_value = submission[jotform_key]
        if jotform_value is None:
            jotform_value = ''
        jotform_value = str(jotform_value)

        pds_value = ''
        if pds_key in member:
            pds_value = member[pds_key]

        if pds_value != jotform_value:
            results[mid] = jotform_value

    return results

###########################################################################

def _simple_csv_member_write(filename, results, members, fields, description, user_function, log):
    with open(filename, 'w') as fp:
        writer = csv.DictWriter(fp, fieldnames=fields)
        writer.writeheader()

        num_updates = 0
        for mid, value in results.items():
            member = members[mid]
            item = {
                'MID' : mid,
                'Name' : member['Name']
            }
            user_function(mid, member, value, item)
            writer.writerow(item)
            num_updates += 1

        log.info(f"Write {num_updates} {description}")
        log.info(f"Wrote {filename}")

###########################################################################

def member_sex_updates(member_submissions, members, log):
    results = _simple_pds_jotform_member_compare(member_submissions, members,
        'sex', 'Gender', log)

    def _make_item(mid, member, value, item):
        item['Sex'] = value

    filename = 'census2021-member-sex-updates.csv'
    fields = ['MID', 'Name', 'Sex']
    _simple_csv_member_write(filename, results, members,
        fields, 'Member sex updates', _make_item, log)

###########################################################################

def member_marital_status_updates(member_submissions, members, log):
    results = _simple_pds_jotform_member_compare(member_submissions, members,
        'marital', 'marital_status', log)

    def _make_item(mid, member, value, item):
        item['Marital status'] = value

    filename = 'census2021-member-marital-status-updates.csv'
    fields = ['MID', 'Name', 'Marital status']
    _simple_csv_member_write(filename, results, members,
        fields, 'Member marital status updates', _make_item, log)

###########################################################################

def member_marriage_date_updates(member_submissions, members, log):
    results = dict()

    jotform_key = 'wedding'
    pds_key = 'marriage_date'
    for mid, submission in member_submissions.items():
        member = members[mid]

        jotform_value = submission[jotform_key]
        if jotform_value == '':
            jotform_value = None
        if jotform_value is not None:
            # It's (usually) a datetime -- convert it to a date, because that's what PDS's data is
            if type(jotform_value) is datetime.datetime:
                jotform_value = jotform_value.date()
            else:
                jotform_value = None

        pds_value = None
        if pds_key in member:
            pds_value = member[pds_key]
        if pds_value == PDSChurch.date_never:
            pds_value = None

        if pds_value != jotform_value:
            results[mid] = jotform_value

    def _make_item(mid, member, value, item):
        item['Marriage date'] = value

    filename = 'census2021-member-marriage-date-updates.csv'
    fields = ['MID', 'Name', 'Marriage date']
    _simple_csv_member_write(filename, results, members,
        fields, 'Member marriage date updates', _make_item, log)

###########################################################################

def member_student_updates(member_submissions, members, log):
    results = dict()

    for mid, submission in member_submissions.items():
        member = members[mid]

        # Only process those who said "yes" to student
        jotform_student = _jotform_boolean(submission['student'])
        if not jotform_student:
            continue

        jotform_school = submission['school']

        pds_occupation = None
        pds_key = 'occupation'
        if pds_key in member:
            pds_occupation = member[pds_key]

        pds_location = None
        pds_key = 'Location'
        if pds_key in member:
            pds_location = member[pds_key]

        if pds_occupation != 'Student' or pds_location != jotform_school:
            results[mid] = jotform_school

    def _make_item(mid, member, value, item):
        item['Occupation'] = 'Student'
        item['Location'] = value

    filename = 'census2021-member-student-updates.csv'
    fields = ['MID', 'Name', 'Occupation', 'Location']
    _simple_csv_member_write(filename, results, members,
        fields, 'Member school updates', _make_item, log)

    return results

###########################################################################

def member_employment_updates(member_submissions, members, log):
    results = dict()

    for mid, submission in member_submissions.items():
        member = members[mid]

        # Only process those who said "no" to student
        jotform_student = _jotform_boolean(submission['student'])
        if jotform_student:
            continue

        jotform_employed = _jotform_boolean(submission['employed'])
        if not jotform_employed:
            continue

        jotform_occupation = submission['occupation']
        jotform_employer = submission['employer']

        pds_occupation = None
        pds_key = 'occupation'
        if pds_key in member:
            pds_occupation = member[pds_key]
        if pds_occupation == '':
            pds_occupation = None

        pds_location = None
        pds_key = 'Location'
        if pds_key in member:
            pds_location = member[pds_key]
        if pds_location == '':
            pds_location = None

        if pds_occupation != jotform_occupation or pds_location != jotform_employer:
            results[mid] = {
                'occupation' : jotform_occupation,
                'employer' : jotform_employer,
            }

    def _make_item(mid, member, value, item):
        item['Occupation'] = value['occupation']
        item['Location'] = value['employer']

    filename = 'census2021-member-employer-updates.csv'
    fields = ['MID', 'Name', 'Occupation', 'Location']
    _simple_csv_member_write(filename, results, members,
        fields, 'Member employer updates', _make_item, log)

###########################################################################

# PDS stores languages in a single field.
# ECC has the following conventions:
#
# - languages are separated by /
# - first one in the list is the primary language
#
# We ask on Jotform "primary" and "additional" languages.
# We collapse these down to a single string to compare to the PDS value.
def member_language_updates(member_submissions, members, log):
    results = dict()

    def _make_str(parts):
        return '/'.join(sorted(parts))

    for mid, submission in member_submissions.items():
        member = members[mid]

        jotform_language = submission['language_primary']
        if jotform_language is None:
            jotform_language = ''
        jotform_add = submission['language_additional']
        if jotform_add is not None:
            parts = jotform_add.split('\n')
            # Some people put the same language in primary and additional :-(
            parts = [x for x in parts if x != jotform_language]
            if len(parts) > 0:
                jotform_language += '/' + _make_str(parts)

        pds_key = 'language'
        pds_language = member[pds_key] if pds_key in member else ''
        if pds_language is None:
            pds_language = ''

        if jotform_language != pds_language:
            results[mid] = jotform_language

    def _make_item(mid, member, value, item):
        item['Language'] = value

    filename = 'census2021-member-language-updates.csv'
    fields = ['MID', 'Name', 'Language']
    _simple_csv_member_write(filename, results, members,
        fields, 'Member language updates', _make_item, log)

###########################################################################

def member_find_yes_email_but_no_address(member_submissions, members, log):
    nonsense = dict()
    for mid, submission in member_submissions.items():

        member = members[mid]
        emails = PDSChurch.find_any_email(member)

        want_email = True if submission['email_eh'] or submission['email_obit'] or submission['email_taxes'] or submission['email_business'] or submission['email_weekday'] else False

        if want_email and len(emails) == 0:
            nonsense[mid] = True

    def _make_item(mid, member, value, item):
        pass

    filename = f'census2021-member-who-wants-emails-but-has-no-email-address.csv'
    fields = ['MID', 'Name']
    _simple_csv_member_write(filename, nonsense, members,
        fields, 'Member marital status updates', _make_item, log)

###########################################################################

def family_annual_donations_email(member_submissions, members, families, log):
    # Make per-Family lists of desired email addresses
    fid_submissions = dict()
    family_emails = dict()
    for mid, submission in member_submissions.items():
        member = members[mid]
        fid = member['FamRecNum']
        fid_submissions[fid] = True

        jotform_value = False
        if submission['email_taxes'] and 'Yes'in submission['email_taxes']:
            jotform_value = True

        # If this Member doesn't want the annual donations email, skip them
        if not jotform_value:
            continue

        # If we're here, then this Member wants the annual donations email
        emails = PDSChurch.find_any_email(member)
        if len(emails) > 0:
            fid = member['family']['FamRecNum']
            if fid not in family_emails:
                family_emails[fid] = list()

            for email in emails:
                family_emails[fid].append(email)

    # Now find families who submitted census results but did not have
    # any member say "yes, I want the annual tax email"
    families_with_no_one_designated = dict()
    for fid in sorted(fid_submissions.keys()):
        if fid not in family_emails:
            families_with_no_one_designated[fid] = True

    # Now find families who submitted census results and selected one or more Members
    # to receive the annual tax email, but we have no email addresses.
    families_with_no_emails = dict()
    for fid in sorted(fid_submissions.keys()):
        if fid in family_emails and len(family_emails[fid]) == 0:
            families_with_no_emails[fid] = True

    def _write_results(filename, data, email):
        with open(filename, 'w') as fp:
            fields = ['FID', 'Name']
            if email:
                fields.append('Email')
            writer = csv.DictWriter(fp, fieldnames=fields)
            writer.writeheader()

            num_updates = 0
            for fid, value in data.items():
                family = families[fid]

                item = {
                    'FID'   : fid,
                    'Name'  : family['Name'],
                }
                if email:
                    for e in value:
                        item['Email'] = email
                        writer.writerow(item)
                        num_updates += 1
                else:
                    writer.writerow(item)
                    num_updates += 1

            if email:
                log.info(f"Write {num_updates} Member emails as the Family emails for {len(family_emails)} Families")
            else:
                log.info(f"Write {num_updates} Families")
            log.info(f"Wrote {filename}")

    # We're going to run an external C++ pxlib program that will remove all
    # remove all Family email addresses for all FIDs that submitted census.
    # Therefore, make a CSV that can be imported to PDS to set all the email
    # addresses that we want on the Family.
    _write_results('census2021-family-annual-tax-emails.csv',
        family_emails, email=True)

    _write_results('census2021-family-annual-tax-no-one-designated.csv',
        families_with_no_one_designated, email=False)
    _write_results('census2021-family-annual-tax-no-emails.csv',
        families_with_no_emails, email=False)

###########################################################################

def _member_keyword_addremove(member_submissions, members, keyword, extract_jotform_fn, filename, log):
    members_to_add = dict()
    members_to_remove = dict()

    key = 'keywords'
    for mid, submission in member_submissions.items():
        jotform_value = extract_jotform_fn(submission)

        member = members[mid]
        pds_value = False
        if key in member and keyword in member[key]:
            pds_value = True

        log.debug(f"Member keyword: {member['Name']}, jotform {jotform_value}, pds {pds_value}")

        if jotform_value == pds_value:
            # What the Member wants is already in PDS.
            # Nothing to do.
            continue

        if jotform_value:
            members_to_add[mid] = keyword
        else:
            members_to_remove[mid] = keyword

    def _write_results(type, filename, results):
        filename = f'census2021-member-keyword-{type}-{filename}-updates.csv'
        with open(filename, 'w') as fp:
            fields = ['MID', 'Name', f'{type} keyword']
            writer = csv.DictWriter(fp, fieldnames=fields)
            writer.writeheader()

            num_updates = 0
            for mid, value in results.items():
                member = members[mid]

                item = {
                    'MID'             : mid,
                    'Name'            : member['Name'],
                    f'{type} keyword' : value,
                }
                writer.writerow(item)
                num_updates += 1

            log.info(f"Write {num_updates} Member '{keyword}' keyword {type} updates")
            log.info(f"Wrote {filename}")

    _write_results("add", filename, members_to_add)
    _write_results("remove", filename, members_to_remove)

###########################################################################

def member_weekday_mass_email(member_submissions, members, log):
    def _extract_jotform_pwe(submission):
        jotform_value = False
        if submission['email_weekday']:
            jotform_value = True if 'Yes' in submission['email_weekday'] else False

        return jotform_value

    # JMS is this keyword too long?
    _member_keyword_addremove(member_submissions, members, "Weekday Mass Email",
                            _extract_jotform_pwe, 'weekday-mass-email', log)

###########################################################################

def member_business_logistics_email(member_submissions, members, log):
    def _extract_jotform_pwe(submission):
        jotform_value = False
        if submission['email_business']:
            jotform_value = True if 'Yes' in submission['email_business'] else False

        return jotform_value

    # JMS is this keyword too long?
    _member_keyword_addremove(member_submissions, members, "Business Logistics Email",
                            _extract_jotform_pwe, 'business-logistics-email', log)

###########################################################################

def member_obit_email(member_submissions, members, log):
    def _extract_jotform_pwe(submission):
        jotform_value = False
        if submission['email_obit']:
            jotform_value = True if 'Yes' in submission['email_obit'] else False

        return jotform_value

    _member_keyword_addremove(member_submissions, members, "Obituary Email",
                            _extract_jotform_pwe, 'obituary-email', log)

###########################################################################

def member_parish_wide_email(member_submissions, members, log):
    def _extract_jotform_pwe(submission):
        jotform_value = False
        if submission['email_eh']:
            jotform_value = True if 'Yes' in submission['email_eh'] else False

        return jotform_value

    _member_keyword_addremove(member_submissions, members, "Parish-wide Email",
                            _extract_jotform_pwe, 'parish-wide-email', log)

###########################################################################

def member_cell_updates(member_submissions, members, log):
    # Calculate phone numbers to add to a Member and phone numbers to
    # remove from a Member
    additions     = dict()
    removals      = dict()

    for mid, submission in member_submissions.items():
        member = members[mid]
        item   = {
            'MID'  : mid,
            'Name' : member['Name'],
        }

        jotform_phone = submission['cell']
        if jotform_phone is None:
            jotform_phone = ''
        jotform_phone = str(jotform_phone)

        # Iterate through all the Member phones:
        # - For all phones that are not the target phone, remove them.
        # - If we didn't find the target phone, add it.
        found = False
        jotform_digits = reduce_phone_to_digits(jotform_phone)
        jotform_ac     = jotform_digits[0:3]
        jotform_rest   = f'{jotform_digits[3:6]}-{jotform_digits[6:]}'

        key = 'phones'
        if key in member:
            for entry in member[key]:
                digits = reduce_phone_to_digits(entry['number'])
                if digits == jotform_digits:
                    # If this is the phone number they gave us on the jotform,
                    # huzzah!
                    found = True
                else:
                    # If this is not the phone number they gave us on the
                    # jotform, remove it.
                    if mid not in removals:
                        removals[mid] = list()
                    removals[mid].append(entry['number'])

        # If we didn't find the phone number they gave us on the jotform,
        # add it.
        if not found and jotform_phone != '':
            additions[mid] = {
                'ac'        : jotform_ac,
                'phone'     : jotform_rest,
                'type'      : 'Cell',
            }

    # Write the removals
    def _make_removal_item(mid, member, value, item):
        item['Phone to remove'] = value

    filename = 'census2021-member-cell-removals.csv'
    fields = ['MID', 'Name', 'Phone to remove']
    _simple_csv_member_write(filename, removals, members,
        fields, 'Member cell removals', _make_removal_item, log)

    # Write the additions
    def _make_add_item(mid, member, value, item):
        item['Area code'] = value['ac']
        item['Phone'] = value['phone']
        item['Type'] = value['type']

    filename = 'census2021-member-cell-additions.csv'
    fields = ['MID', 'Name', 'Type', 'Area code', 'Phone']
    _simple_csv_member_write(filename, additions, members,
        fields, 'Member cell additions', _make_add_item, log)

###########################################################################

def member_email_updates(member_submissions, members, log):
    emails_to_remove = dict()
    emails_to_add = dict()

    for mid, submission in member_submissions.items():
        member = members[mid]

        jotform_email = submission['email']
        if jotform_email is None:
            jotform_email = ''
        else:
            jotform_email = jotform_email

        # Remove everything except the jotform email
        found = False
        for key in email_keys:
            for entry in member[key]:
                email = entry['EMailAddress'].strip()

                # Keep this one!
                if jotform_email == email:
                    found = True
                    continue

                # Remove all others
                if mid not in emails_to_remove:
                    emails_to_remove[mid] = list()
                emails_to_remove[mid].append(email)

        if not found and jotform_email != "":
            emails_to_add[mid] = [ jotform_email ]

    def _write_results(type, results):
        filename = f'census2021-member-email-{type}-updates.csv'
        with open(filename, 'w') as fp:
            fields = ['MID', 'Name', 'Email']
            writer = csv.DictWriter(fp, fieldnames=fields)
            writer.writeheader()

            num_updates = 0
            for mid, values in results.items():
                member = members[mid]

                item = {
                    'MID'   : mid,
                    'Name'  : member['Name'],
                }
                for email in values:
                    item['Email'] = email
                    writer.writerow(item)
                    num_updates += 1

            log.info(f"Write {num_updates} Member email {type} updates")
            log.info(f"Wrote {filename}")

    _write_results('remove', emails_to_remove)
    _write_results('add', emails_to_add)

###########################################################################

def member_birthday_updates(member_submissions, members, log):
    months = {
        'January': 1,
        'February': 2,
        'March': 3,
        'April': 4,
        'May': 5,
        'June': 6,
        'July': 7,
        'August': 8,
        'September': 9,
        'October': 10,
        'November': 11,
        'December': 12,
    }

    num_updates = 0
    filename = f'census2021-member-birthday-updates.csv'
    with open(filename, 'w') as fp:
        fields = ['MID', 'Name', f'Birthday']
        writer = csv.DictWriter(fp, fieldnames=fields)
        writer.writeheader()

        for mid, submission in member_submissions.items():
            member = members[mid]

            # Load up the values from jotform
            mob = submission['mob']
            yob = submission['yob']

            # If they didn't submit anything, skip this entry
            if mob is None and yob is None:
                continue

            # Load up the old values from PDS
            key = 'DateOfBirth'
            if key not in member or member[key] == '0000-00-00':
                year  = None
                month = None
                date  = None
            else:
                year  = member['YearOfBirth']
                month = member['MonthOfBirth']
                if month == '':
                    month = 1
                else:
                    month = months[month]
                date  = member['DayOfBirth']

            # Was there a change?
            changed = False
            if mob is not None and months[mob] != month:
                changed = True
                month = months[mob]
            if yob is not None and int(yob) != member['YearOfBirth']:
                changed = True
                year = int(yob)

            # If there was no change, skip this entry
            if not changed:
                continue

            # Set semi-arbitrary values for any value we don't have
            if date is None:
                date = 1
            if month is None:
                month = 1
            if year is None:
                year = 1899

            # Make a new birthday.
            birthday = f'{year:04d}-{month:02d}-{date:02d}'
            item = {
                'MID'      : mid,
                'Name'     : member['Name'],
                'Birthday' : birthday,
            }
            writer.writerow(item)
            num_updates += 1

    log.info(f"Write {num_updates} name birthday updates")
    log.info(f"Wrote {filename}")

###########################################################################

def member_name_updates(member_submissions, members, log):
    # To make these easy for human review, we're going to do a different file
    # for each type of name update.
    first_updates    = dict()
    nickname_updates = dict()
    middle_updates   = dict()
    last_updates     = dict()
    suffix_updates   = dict()

    for mid, submission in member_submissions.items():
        member = members[mid]

        def _check(pds_key, submission_key, results):
            if submission[submission_key] == "":
                submission[submission_key] = None

            if member[pds_key] != submission[submission_key]:
                results[mid] = submission[submission_key]

        _check('first', 'first', first_updates)
        _check('nickname', 'nick', nickname_updates)
        _check('middle', 'middle', middle_updates)
        _check('last', 'last', last_updates)
        _check('suffix', 'suffix', suffix_updates)

    def _write_results(type, results):
        filename = f'census2021-member-name-{type}-updates.csv'
        with open(filename, 'w') as fp:
            fields = ['MID', 'Name', f'New {type}']
            writer = csv.DictWriter(fp, fieldnames=fields)
            writer.writeheader()

            num_updates = 0
            for mid, value in results.items():
                member = members[mid]

                item = {
                    'MID'         : mid,
                    'Name'        : member['Name'],
                    f'New {type}' : value,
                }
                writer.writerow(item)
                num_updates += 1

            log.info(f"Write {num_updates} name {type} updates")
            log.info(f"Wrote {filename}")

    _write_results("first", first_updates)
    _write_results("nickname", nickname_updates)
    _write_results("middle", middle_updates)
    _write_results("last", last_updates)
    _write_results("suffix", suffix_updates)

###########################################################################

def member_extract_submissions(ws, members, log):
    log.info("Extracting Member updates...")

    member_jotform_fields = [
        'mid',
        'title',
        'first',
        'nick',
        'middle',
        'last',
        'suffix',
        'mob',
        'yob',
        'cell',
        'email',
        'email_eh',
        'email_obit',
        'email_taxes',
        'email_business',
        'email_weekday',
        'sex',
        'marital',
        'wedding',
        'student',
        'school',
        'employed',
        'occupation',
        'employer',
        'language_primary',
        'language_additional',
    ]

    def _find_mid_fields(ws):
        mid_fields = list()
        base = 1
        for member_num in range(0, constants.MAX_PDS_FAMILY_MEMBER_NUM+1):
            target = f'mid{member_num}'
            for col in range(base, 9999):
                val = ws.cell(row=1, column=col).value
                if val == target:
                    mid_fields.append(col)
                    base = col
                    break

        return mid_fields

    def _extract_member_data(ws, mid_fields):
        mid_submissions = dict()
        for row in range(2, 9999):
            val = ws.cell(row=row, column=1).value
            if val == '':
                break

            for member_num in range(0, constants.MAX_PDS_FAMILY_MEMBER_NUM):
                results = dict()
                for id, field in enumerate(member_jotform_fields):
                    val = ws.cell(row=row, column=mid_fields[member_num]+id).value
                    if type(val) is str:
                        val = val.strip()
                    results[field] = val

                # The attendance question was only asked for the HoH and Spouse.
                # If we have the attendance field for this member, it's one past
                # the last field we just obtained
                if member_num < 2:
                    val = ws.cell(row=row, column=mid_fields[member_num]+id+1).value
                    results['attendance'] = val

                mid_val = results['mid']
                if mid_val is None:
                    break

                # Only save current members
                mid = int(mid_val)
                if mid in members:
                    mid_submissions[int(mid_val)] = results

        return mid_submissions

    mid_fields      = _find_mid_fields(ws)
    mid_submissions = _extract_member_data(ws, mid_fields)

    log.info(f"Extracted {len(mid_submissions)} Member submissions")

    return mid_submissions

###########################################################################

# This routine gets the "complete" submissions, meaning that the
# dictionary is indexed by the Jotform fieldnames (not our "simple" fieldnames).
def survey(wb, ws, families, log):
    currents = dict()
    prefers  = dict()

    additional_currents = dict()
    additional_prefers  = dict()

    # Find the column with questions
    def _find_column(ws, text):
        for i in range(1, 999):
            val = ws.cell(row=1, column=i).value
            if val == text:
                return i
        return None

    fid_col = _find_column(ws, 'fid')
    q1_col = _find_column(ws, 'How does your household currently receive Epiphany announcements?  Select all that apply.')
    q2_col = _find_column(ws, 'How does your household prefer to receive Epiphany announcements or news?  Rank JUST the top 3.')

    official_answers = [
        'Epiphany Happenings weekly email',
        'Bulletin (printed)',
        'Bulletin (online)',
        'Epiphany Facebook',
        'Epiphany Twitter',
        'Epiphany Instagram',
        'MyParish App',
        'Verbal announcements at end of Mass',
        'ReachAlert calls / texts',
    ]

    for row in range(2, 9999):
        val = ws.cell(row=row, column=1).value
        if val == '':
            break

        fid = ws.cell(row=row, column=fid_col).value
        # Only care about survey results from families
        if not fid:
            continue
        fid = int(fid)
        if fid not in families:
            continue

        current = ws.cell(row=row, column=q1_col).value
        prefer  = ws.cell(row=row, column=q2_col).value

        if current:
            responses = current.splitlines()
            currents[fid] = responses
            for response in responses:
                if response not in official_answers:
                    if response not in additional_currents:
                        additional_currents[response] = 0
                    additional_currents[response] += 1
        if prefer:
            responses = prefer.splitlines()
            prefers[fid] = responses
            for response in responses:
                if response not in official_answers:
                    if response not in additional_prefers:
                        additional_prefers[response] = 0
                    additional_prefers[response] += 1

    # Write out the current values
    def _write_results(name, columns, responses, additional):
        num_responses = 0
        filename = f'census2021-survey-{name}.csv'
        with open(filename, 'w') as fp:
            fields = ['FID', 'Env ID', 'Family name']
            for response in sorted(columns):
                fields.append(response)

            writer = csv.DictWriter(fp, fieldnames=fields)
            writer.writeheader()

            for fid in responses:
                family = families[fid]
                item = {
                    'FID'         : fid,
                    'Env ID'      : env_id(family),
                    'Family name' : family['Name'],
                }
                for response in responses[fid]:
                    if response in official_answers:
                        item[response] = 1
                writer.writerow(item)
                num_responses += 1

            # Write a blank line
            item = { 'FID' : '' }
            writer.writerow(item)

            # Write all the additional responses
            for response, count in additional.items():
                item = {
                    'FID' : count,
                    'Env ID' : response,
                }
                writer.writerow(item)

        log.info(f"Found {num_responses} response to the '{name}' survey question")
        log.info(f"Wrote {filename}")

    _write_results('current', official_answers, currents, additional_currents)
    _write_results('prefer', official_answers, prefers, additional_prefers)

###########################################################################

def family_emergency_contacts(active, families, log):
    num_relations = 0
    filename = 'census2021-family-emergency-contacts.csv'
    with open(filename, 'w') as fp:
        fields = ['FID', 'Env ID', 'Family name',
                  'Emergency name', 'Emergency phone', 'Emergency relationship']
        writer = csv.DictWriter(fp, fieldnames=fields)
        writer.writeheader()

        for fid, submission in active.items():
            family = families[fid]
            env    = env_id(family)

            name     = ''
            if submission["emergency first"] or submission["emergency last"]:
                name = f'{submission["emergency first"]} {submission["emergency last"]}'
            phone    = submission["emergency phone"]
            relation = submission["emergency relationship"]

            if name == "" and phone == "" and relation == "":
                continue

            item   = {
                'FID'                    : fid,
                'Env ID'                 : family['ParKey'],
                'Family name'            : family['Name'],
                'Emergency name'         : name,
                'Emergency phone'        : phone,
                'Emergency relationship' : relation,
            }
            writer.writerow(item)
            num_relations += 1

    log.info(f"Added {num_relations} family emergency contacts")
    log.info(f"Wrote {filename}")

###########################################################################

def family_phone_updates(submissions, families, members, log):
    # Convert from the answer we had on Jotform to the appropriate PDS phone type
    type_map = {
        # Have an empty entry for entries with no phone number
        '' : '',
        'Cell / Mobile' : 'Cell',
        'Landline' : 'Home',
    }

    # Calculate phone numbers to add to a Family and phone numbers to
    # remove from a Family
    additions     = dict()
    removals      = dict()

    # We might also end up adding phone numbers to individual members
    mid_additions = dict()

    for fid in sorted(submissions):
        row    = submissions[fid]
        family = families[fid]
        item   = {
            'FID'         : fid,
            'Env ID'      : env_id(family),
            'Family name' : family['Name'],
        }

        # The phone entered on the Jotform will be in one of two fields
        # (we had 2 fields just so that we could ask the question with two
        # different wordings -- only one of the two fields will be filled)

        jotform_phone = row['house phone1']
        if jotform_phone == '':
            jotform_phone = row['house phone2']

        # Iterate through all the Family phones:
        # - For all phones that are not the target phone, remove them.
        # - If we didn't find the target phone, add it.
        found = False
        jotform_digits = reduce_phone_to_digits(jotform_phone)
        jotform_ac     = jotform_digits[0:3]
        jotform_rest   = f'{jotform_digits[3:6]}-{jotform_digits[6:]}'

        key = 'phones'
        if key in family:
            for entry in family[key]:
                digits = reduce_phone_to_digits(entry['number'])
                if digits == jotform_digits:
                    # If this is the phone number they gave us on the jotform,
                    # huzzah!
                    found = True
                else:
                    # If this is not the phone number they gave us on the
                    # jotform, remove it.
                    if fid not in removals:
                        removals[fid] = list()
                    removals[fid].append(entry['number'].strip())

        # If we didn't find the phone number they gave us on the jotform,
        # add it.
        if not found and jotform_phone != '':
            # See if we can find the owner Member of the phone in this Family
            # This is a heuristic, and won't always work.
            owner_mid   = None
            owner_first = row['house phone owner first']
            owner_last  = row['house phone owner last']
            owner       = f"{owner_first} {owner_last}"

            owner_first_lc = owner_first.lower()
            owner_last_lc  = owner_last.lower()
            owner_lc       = owner.lower()
            for member in family['members']:
                if member['email_name'].lower() == owner_lc:
                    owner_mid = member['MemRecNum']
                    break
                elif (owner_first_lc == member['first'].lower() and
                      owner_last_lc  == member['last'].lower()):
                    owner_mid = member['MemRecNum']
                    break
                elif ('nickname' in member and
                      member['nickname'] and
                      owner_first_lc == member['nickname'].lower() and
                      owner_last_lc  == member['last'].lower()):
                    owner_mid = member['MemRecNum']
                    break

            # If we found a MID, does that member have this phone number already?
            if owner_mid:
                member = members[owner_mid]
                found = False
                if key in member:
                    for entry in member[key]:
                        digits = reduce_phone_to_digits(entry['number'])
                        if digits == jotform_digits:
                            # If this is the phone number they gave us on the jotform,
                            # huzzah!
                            found = True
                            break

                # If we didn't find the phone number on that Member, queue it
                # up to be added
                if not found:
                    mid_additions[owner_mid] = {
                        'ac'    : jotform_ac,
                        'phone' : jotform_rest,
                    }

            additions[fid] = {
                'ac'        : jotform_ac,
                'phone'     : jotform_rest,
                'type'      : type_map[row['house phone type']],
                'owner'     : owner,
                'owner_mid' : owner_mid,
            }

    # Write the removals
    num_removals = 0
    filename = 'census2021-family-phone-removals.csv'
    with open(filename, 'w') as fp:
        fields = ['FID', 'Env ID', 'Family name', 'Family phone to remove']
        writer = csv.DictWriter(fp, fieldnames=fields)
        writer.writeheader()

        for fid in sorted(removals):
            family = families[fid]
            env    = env_id(family)

            for phone in removals[fid]:
                item   = {
                    'FID'         : fid,
                    'Env ID'      : env,
                    'Family name' : family['Name'],
                    'Family phone to remove' : phone,
                }
                writer.writerow(item)
                num_removals += 1

    log.info(f"Removed {num_removals} family phones")
    log.info(f"Wrote {filename}")

    # Write the Family additions
    num_additions = 0
    filename = 'census2021-family-phone-additions.csv'
    with open(filename, 'w') as fp:
        fields = ['FID', 'Env ID', 'Family name', 'Area code', 'Phone', 'Type', 'Owner']
        writer = csv.DictWriter(fp, fieldnames=fields)
        writer.writeheader()

        for fid in sorted(additions):
            family = families[fid]
            env    = env_id(family)
            entry  = additions[fid]

            item   = {
                'FID'         : fid,
                'Env ID'      : env,
                'Family name' : family['Name'],
                'Area code'   : entry['ac'],
                'Phone'       : entry['phone'],
                'Type'        : entry['type'],
                'Owner'       : entry['owner'],
            }
            writer.writerow(item)
            num_additions += 1

    log.info(f"Added {num_additions} family phones")
    log.info(f"Wrote {filename}")

    # Write the Family additions
    num_mid_additions = 0
    filename = 'census2021-member-phone-additions.csv'
    with open(filename, 'w') as fp:
        fields = ['MID', 'Member name', 'Area code', 'Phone', 'Type']
        writer = csv.DictWriter(fp, fieldnames=fields)
        writer.writeheader()

        for mid in sorted(mid_additions):
            member = members[mid]
            entry  = mid_additions[mid]

            item   = {
                'MID'         : mid,
                'Member name' : member['Name'],
                'Area code'   : entry['ac'],
                'Phone'       : entry['phone'],
                'Type'        : 'Cell',
            }
            writer.writerow(item)
            num_mid_additions += 1

    log.info(f"Added {num_additions} member phones")
    log.info(f"Wrote {filename}")

###########################################################################

def family_mass_updates(submissions, families, log):
    # PDS can only import additions to keywords.
    # The scheme we use is this:
    #
    # - Run a PDS Family "Quick process":
    #   - Remove keywords
    #   - Use a selection of a comma-delimited list of all Families who submitted census electronically
    #   - Do this to remove ALL mass keywords
    # - Then run a Family import of keywords
    #   - This import will contain ALL the mass keywords for each Family

    # Go through each submission and remove all stale keywords from the Family,
    # and add each new keyword to the Family
    key = 'keywords'
    for fid in sorted(submissions):
        row    = submissions[fid]
        family = families[fid]
        item   = {
            'FID'         : fid,
            'Env ID'      : env_id(family),
            'Family name' : family['Name'],
        }

        weekday   = 'Weekly Mass - '
        spanish   = 'Spanish Mass - '
        mass_530  = '5:30pm Sat. Mass - '
        mass_900  = '9:00am Sun. Mass - '
        mass_1130 = '11:30am Sun. Mass - '

        changed = False
        for tuple in [('weekday frequency', weekday),
                      ('spanish frequency', spanish),
                      ('530 frequency', mass_530),
                      ('900 frequency', mass_900),
                      ('1130 frequency', mass_1130)]:
            row_field = tuple[0]
            prefix    = tuple[1]

            # If the jotform value is empty, assume "None"
            jotform_value = row[row_field]
            if jotform_value == '':
                jotform_value = 'Never'
            jotform_keyword = f'{prefix}{jotform_value}'

            # Does the jotform response already exist as a keyword?
            # If so, we're done with this entry.
            if key in family and jotform_keyword in family[key]:
                continue

            # So the jotform answer does not exist as a keyword.
            # This qualifies as an addition.
            # Now we need to check and see if we need to remove another
            # existing keyword with the same prefix.
            if key in family:
                to_remove = list()
                for i, pds_keyword in enumerate(family[key]):
                    if pds_keyword.startswith(prefix):
                        to_remove.append(i)

                to_remove.reverse()
                for i in to_remove:
                    del family[key][i]

            if key not in family:
                family[key] = list()
            family[key].append(jotform_keyword)

    # Write out all keywords for all Families who submitted (not *all* Families
    # PDS -- just the ones who submitted census)
    fields   = ['FID', 'Env ID', 'Family name', 'Keyword']
    filename = f'census2021-family-mass-keywords.csv'
    with open(filename, 'w') as fp:
        writer = csv.DictWriter(fp, fieldnames=fields)
        writer.writeheader()

        for fid in sorted(submissions):
            family = families[fid]
            item   = {
                'FID'         : fid,
                'Env ID'      : env_id(family),
                'Family name' : family['Name'],
            }

            for keyword in family[key]:
                if (keyword.startswith(weekday) or
                    keyword.startswith(spanish) or
                    keyword.startswith(mass_530) or
                    keyword.startswith(mass_900) or
                    keyword.startswith(mass_1130)):
                    item['Keyword'] = keyword
                    writer.writerow(item)

    log.info(f"Wrote {filename}")

    # Write out the comma-delimited list of envelope IDs for all Families who
    # submitted census
    filename = f'census2021-family-env-ids.txt'
    with open(filename, 'w') as fp:
        env_ids = list()
        for fid in sorted(submissions):
            env_ids.append(families[fid]['ParKey'].strip())
        fp.write(', '.join(env_ids) + '\n')

    log.info(f"Wrote {filename}")

###########################################################################

def family_address_updates(submissions, families, log):
    num_changes = 0
    filename = 'census2021-family-address-updates.csv'
    with open(filename, 'w') as fp:
        fields = ['FID', 'Env ID', 'Family name',
                  'Address 1 (import)', 'Address 2 (import)', 'City+State (import)', 'Zip (import)',
                  'Address 1 (old)', 'Address 2 (old)', 'City+State (old)', 'Zip (old)',
                  'Address 1 (new)', 'Address 2 (new)', 'City+State (new)', 'Zip (new)' ]
        writer = csv.DictWriter(fp, fieldnames=fields)
        writer.writeheader()

        for fid in sorted(submissions):
            row    = submissions[fid]
            family = families[fid]
            item   = {
                'FID'         : fid,
                'Env ID'      : env_id(family),
                'Family name' : family['Name'],
            }

            # Jotform gathered the City and State separately, but they're in a
            # single field in PDS (!).
            row['city_state'] = f"{row['city']} {row['state']}"
            # PDS stores the value as "City State", but we must insert the value
            # as "City, State".  Arrgh!
            city_state_value = f"{row['city']}, {row['state']}"

            changed = False
            for tuple in [('street1', None, 'StreetAddress1', 'Address 1'),
                          ('street2', None, 'StreetAddress2', 'Address 2'),
                          ('city_state', city_state_value, 'city_state', 'City+State'),
                          ('zip', None, 'StreetZip', 'Zip')]:
                row_field  = tuple[0]
                insert_val = tuple[1]
                pds_field  = tuple[2]
                csv_base   = tuple[3]

                csv_import = f'{csv_base} (import)'
                csv_old    = f'{csv_base} (old)'
                csv_new    = f'{csv_base} (new)'

                item[csv_import] = family[pds_field]
                if row[row_field] != family[pds_field].strip():
                    changed = True
                    if insert_val:
                        item[csv_import] = insert_val
                        item[csv_old]    = family[pds_field].strip()
                        item[csv_new]    = insert_val
                    else:
                        item[csv_import] = row[row_field]
                        item[csv_old]    = family[pds_field].strip()
                        item[csv_new]    = row[row_field]

            if changed:
                writer.writerow(item)
                num_changes += 1

    log.info(f"Found {num_changes} address changes")
    log.info(f"Wrote {filename}")

###########################################################################

def family_census2021_keywords(submissions, families, log):
    filename = 'census2021-family-keywords.csv'
    with open(filename, 'w') as fp:
        fields = ['FID', 'Env ID', 'Family name', 'Keyword']
        writer = csv.DictWriter(fp, fieldnames=fields)
        writer.writeheader()

        for fid in sorted(submissions):
            family = families[fid]
            item = {
                'FID'         : fid,
                'Env ID'      : env_id(family),
                'Family name' : family['Name'],
                'Keyword'     : 'Active: Census 2021',
            }
            writer.writerow(item)

    log.info(f"Wrote {filename}")

###########################################################################

def family_no_longer_active(submissions, families, log):
    filename = 'census2021-inactive-families.csv'
    with open(filename, 'w') as fp:
        fields = [ 'Envelope ID', 'Name' ]
        writer = csv.DictWriter(fp, fieldnames=fields)
        writer.writeheader()

        for fid, submission in submissions.items():
            if fid in families:
                family = families[fid]
                env    = env_id(family)
                name   = family['Name']
            else:
                env  = '???'
                name = submission['name']

            log.info(f"Found inactive family: {name}")
            item = {
                'Envelope ID' : env,
                'Name'        : name,
            }
            writer.writerow(item)

    log.info(f"Wrote {filename}")

###########################################################################

# We read the Jotform Google Sheet 3 different ways for 3 different
# purposes:
#
# 1. CSV DictReader with some "simplified" field names.  We do this for several
# chases where using the "simple" field names is good enough to access the
# fields (vs. Jotform's unweildy column/field names).
#
# 2. CSV Dict Reader with Jotform's names.  We do this for accessing some of the
# later fields.  For example: we don't want to have to make up "simple" fields
# for a few hundred fields just to get to the "comments" questions at the very
# end.  So we just let the DictReader use Jotform's column names, even though
# they're unweildy.
#
# 3. XSLX.  We do this because some of the multi-select Jotform fields are
# stored as "A\nB\nC" in the cell values.  But downloading and reading them as
# CSV values loses the \n somewhere along the way.  Reading the data as an XLSX
# preserves the \n so that we can properaly separate out the multi-values by the
# seperator \n's.
#
def read_census_google_spreadsheet(args, families, log):
    apis = {
        'drive' : { 'scope'       : Google.scopes['drive'],
                    'api_name'    : 'drive',
                    'api_version' : 'v3', },
    }
    services = GoogleAuth.service_oauth_login(apis,
                                              app_json=args.app_id,
                                              user_json=args.user_credentials,
                                              log=log)
    google = services['drive']

    #---------------------------------------------------------------------

    # Just return the raw CSV data from the Jotform Google sheet, split into
    # lines.
    def _read_jotform_gsheet(google, gfile_id):
        csv_response = google.files().export(fileId=gfile_id,
                                          mimeType=Google.mime_types['csv']).execute()
        xlsx_response = google.files().export(fileId=gfile_id,
                                          mimeType=Google.mime_types['xlsx']).execute()
        return csv_response.decode('utf-8').splitlines(), xlsx_response

    #---------------------------------------------------------------------

    def _simple_csvreader(lines):
        # The ordering of these fields is critical, although the names are not.
        # We use an abbreviated set of "simple" fieldnames (since the Jotform
        # fieldnames are... unweildy)
        fields = [
            "SubmissionDate",
            "active",
            "emailReply",

            'fid',
            'name',
            'street1',
            'street2',
            'city',
            'state',
            'zip',
            'country',

            'house phone1',
            'house phone2',
            'house phone type',
            'house phone owner first',
            'house phone owner last',

            'emergency first',
            'emergency last',
            'emergency relationship',
            'emergency phone',

            'weekday frequency',
            'spanish frequency',
            '530 frequency',
            '900 frequency',
            '1130 frequency',

            'survey current',
            'suvery prefer',
            # There are other fields in these spreadsheets, but we only care about
            # the ones above (actually, we only care about the fid)
        ]

        csvreader = csv.DictReader(lines, fieldnames=fields)
        return csvreader

    #---------------------------------------------------------------------

    def _complete_csvreader(lines):
        # This will make a dictionary based on the field names in the first row
        # of the CSV data (i.e., Jotform's field names)
        csvreader = csv.DictReader(lines)
        return csvreader

    #---------------------------------------------------------------------

    def _convert(csv_data, log):
        output_data = dict()

        first = True
        for row in csv_data:
            # Skip the first / title row
            if first:
                first = False
                continue

            fid = row['fid']
            if fid:
                fid = int(fid)
            else:
                fid = 0
            output_data[fid] = row

        return output_data

    #---------------------------------------------------------------------

    def _convert_to_wb(raw_xlsx):
        # There almost certainly a more efficient way to do this
        tmpfile = 'results-google.xlsx'
        with open(tmpfile, "wb") as fp:
            fp.write(raw_xlsx)

        wb = openpyxl.load_workbook(tmpfile)
        # Assume there's 1 sheet
        name = wb.sheetnames[0]
        ws = wb[name]

        os.unlink(tmpfile)

        return wb, ws

    log.info("Loading Jotform submissions Google sheet")
    raw_csv_lines, raw_xlsx = _read_jotform_gsheet(google, constants.jotform_gsheet_gfile_id)
    simple_csvreader        = _simple_csvreader(raw_csv_lines)
    simple_submissions      = _convert(simple_csvreader, log=log)
    complete_csvreader      = _complete_csvreader(raw_csv_lines)
    complete_submissions    = _convert(complete_csvreader, log=log)
    wb, ws                  = _convert_to_wb(raw_xlsx)
    log.info(f"Found {len(simple_submissions)} unique respondents")

    # Now separate them into those who said they are still active and those who
    # said they are not
    active   = dict()
    inactive = dict()
    for fid, row in simple_submissions.items():
        is_active = True if row['active'] == 'Yes' else False
        if not is_active or fid not in families:
            inactive[fid] = row
        else:
            active[fid] = row

    return active, inactive, complete_submissions, wb, ws

###########################################################################

def setup_args():
    # These options are for Google Authentication
    tools.argparser.add_argument('--app-id',
                                 default=constants.gapp_id,
                                 help='Filename containing Google application credentials.  Only necessary if sending an email that contains a {*_reminder} tag.')
    tools.argparser.add_argument('--user-credentials',
                                 default=constants.guser_cred_file,
                                 help='Filename containing Google user credentials.  Only necessary if sending an email that contains a {*_reminder} tag.')

    args = tools.argparser.parse_args()

    return args

###########################################################################

def main():
    args = setup_args()
    log = ECC.setup_logging(debug=False)

    # Read in all the PDS data
    log.info("Reading PDS data...")
    (pds, families,
     members) = PDSChurch.load_families_and_members(filename='pdschurch.sqlite3',
                                                    parishioners_only=True,
                                                    log=log)

    active, inactive, complete, wb_complete, ws_complete = read_census_google_spreadsheet(args, families, log=log)

    #family_no_longer_active(inactive, families, log)
    #family_census2021_keywords(active, families, log)
    #family_address_updates(active, families, log)
    #family_mass_updates(active, families, log)
    #family_phone_updates(active, families, members, log)
    #family_emergency_contacts(active, families, log)

    #survey(wb_complete, ws_complete, families, log)

    member_submissions = member_extract_submissions(ws_complete, members, log)
    #member_name_updates(member_submissions, members, log)
    #member_birthday_updates(member_submissions, members, log)
    member_cell_updates(member_submissions, members, log)
    member_email_updates(member_submissions, members, log)

    #member_parish_wide_email(member_submissions, members, log)
    #member_obit_email(member_submissions, members, log)
    #member_business_logistics_email(member_submissions, members, log)
    #member_weekday_mass_email(member_submissions, members, log)
    family_annual_donations_email(member_submissions, members, families, log)

    member_sex_updates(member_submissions, members, log)
    member_marital_status_updates(member_submissions, members, log)
    member_marriage_date_updates(member_submissions, members, log)

    member_student_updates(member_submissions, members, log)
    member_employment_updates(member_submissions, members, log)

    member_language_updates(member_submissions, members, log)

    member_find_yes_email_but_no_address(member_submissions, members, log)

    #member_attendance(member_submissions, members, log)

    #comments(complete, families, log)

main()
