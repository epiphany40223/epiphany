#!/usr/bin/env python3

import traceback
import openpyxl
import datetime
import argparse
import sys
import csv
import os

# Load the ECC python modules.  There will be a sym link off this directory.
moddir = os.path.join(os.path.dirname(sys.argv[0]), 'ecc-python-modules')
if not os.path.exists(moddir):
    print("ERROR: Could not find the ecc-python-modules directory.")
    print("ERROR: Please make a ecc-python-modules sym link and run again.")
    exit(1)

sys.path.insert(0, moddir)

import ECC
import PDSChurch
import Google
import GoogleAuth

##############################################################################

import helpers
import constants

from oauth2client import tools

#--------------------------------------------------------------------------

# JMS Kinda yukky that "args" is global
args = None

# Need to make these global for the lambda functions
members  = 1
families = 1

###########################################################################

def setup_google(args, log):
    apis = {
        'drive' : { 'scope'       : Google.scopes['drive'],
                    'api_name'    : 'drive',
                    'api_version' : 'v3', },
    }
    services = GoogleAuth.service_oauth_login(apis,
                                              app_json=args.app_id,
                                              user_json=args.user_credentials,
                                              log=log)
    google = services['drive']

    #---------------------------------------------------------------------

    def _read_jotform_gsheet(google, gfile_id):
        response = google.files().export(fileId=gfile_id,
                                          mimeType=Google.mime_types['xlsx']).execute()

        # There almost certainly a more efficient way to do this
        tmpfile = 'results-google.xlsx'
        with open(tmpfile, "wb") as fp:
            fp.write(response)

        wb = openpyxl.load_workbook(tmpfile)
        # Assume there's 1 sheet
        name = wb.sheetnames[0]
        ws = wb[name]

        os.unlink(tmpfile)

        return wb, ws

    log.info("Loading Jotform submissions Google sheet")
    jotform_xlsx_wb, jotform_xlsx_ws = _read_jotform_gsheet(google, constants.jotform_gsheet_gfile_id)

    return jotform_xlsx_wb, jotform_xlsx_ws

###########################################################################

def add_envid_column(wb, ws, families, log):
    # Insert a column *before* column 4
    ws.insert_cols(4)
    ws['D1'] = 'Env ID'

    row = 2
    while True:
        fid = ws[f'E{row}'].value
        log.info(f"Read fid: {fid}")
        if fid is None or fid == '':
            break
        fid = int(fid)
        if fid in families:
            cell =  f'D{row}'
            family = families[fid]
            ws[cell] = f"'{family['ParKey']}"

        row += 1

    filename = 'jotform-results-with-envid.xlsx'
    wb.save(filename=filename)
    log.info(f"Wrote {filename}")

###########################################################################

def setup_args():
    # These options control which emails are sent.
    # You can only use one of these options at a time.
    group = tools.argparser.add_mutually_exclusive_group()

    tools.argparser.add_argument('--logfile',
                                 default='logfile.txt',
                                 help='Output filename for logfile')

    # These options are for Google Authentication
    tools.argparser.add_argument('--app-id',
                                 default=constants.gapp_id,
                                 help='Filename containing Google application credentials.  Only necessary if sending an email that contains a {*_reminder} tag.')
    tools.argparser.add_argument('--user-credentials',
                                 default=constants.guser_cred_file,
                                 help='Filename containing Google user credentials.  Only necessary if sending an email that contains a {*_reminder} tag.')

    args = tools.argparser.parse_args()

    return args

###########################################################################

def main():
    global families, members

    # JMS Kinda yukky that "args" is global
    global args
    args = setup_args()

    try:
        os.unlink(args.logfile)
    except:
        pass
    log = ECC.setup_logging(debug=False, logfile=args.logfile)

    wb, ws = setup_google(args, log=log)
    # Read in all the PDS data
    log.info("Reading PDS data...")
    (pds, families,
     members) = PDSChurch.load_families_and_members(filename='pdschurch.sqlite3',
                                                    parishioners_only=True,
                                                    log=log)

    add_envid_column(wb, ws, families, log)

main()
