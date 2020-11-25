#!/usr/bin/env python3

# Make sure to pip install everything in requirements.txt.

import sys
sys.path.insert(0, '../../python')

import datetime
import argparse
import os

import ECC
import PDSChurch

import constants

from datetime import date
from datetime import datetime
from datetime import timedelta

import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

start_of_interest = date(month=1, day=1, year=constants.stewardship_year - 1)

##############################################################################

gapp_id         = 'client_id.json'
guser_cred_file = 'user-credentials.json'

##############################################################################

def write_xlsx(ministry_name, interested, log):
    # Make a worksheet
    filename = f"{ministry_name}-interested.xlsx"
    filename = filename.replace('/', '-')
    log.debug(f"Writing {filename}")

    wb = Workbook()
    ws = wb.active

    #---------------------------------------------------------------------

    # Set narrow margins so that we can get as much in as possible if people print it out
    ws.page_margins.left  = 0.4
    ws.page_margins.right = ws.page_margins.left

    ws.page_setup.orientation = ws.ORIENTATION_LANDSCAPE

    #---------------------------------------------------------------------

    title_font    = Font(name='Arial', size=12, bold=True)
    thin_side     = Side(style='thin')
    bottom_border = Border(bottom=thin_side)

    def _make_title(cell, value):
        ws[cell]        = value
        ws[cell].font   = title_font
        ws[cell].border = bottom_border

    _make_title('A1', 'Ministry')
    _make_title('B1', '')
    _make_title('C1', ministry_name)
    _make_title('A2', 'Ministry Chair')
    _make_title('B2', '')
    _make_title('C2', '')

    #---------------------------------------------------------------------

    outcome_font  = Font(name='Arial', size=12, bold=True)
    outcome_align = Alignment(horizontal='center', wrap_text=True)
    box_border    = Border(top=thin_side, bottom=thin_side, left=thin_side, right=thin_side)

    row = 3
    ws.merge_cells(f'G{row}:I{row}')
    cell               = f'G{row}'
    ws[cell]           = 'OUTCOME'
    ws[cell].font      = outcome_font
    ws[cell].alignment = outcome_align
    ws[cell].border    = box_border

    # The "Ouycom" box is merged, so we have to set the top cell border on all the cells
    ws[f'H{row}'].border = Border(top=thin_side, bottom=thin_side)
    ws[f'I{row}'].border = Border(top=thin_side, bottom=thin_side, right=thin_side)

    #---------------------------------------------------------------------

    def _make_heading(cell, value, width, fill=None):
        ws[cell]           = value
        ws[cell].font      = heading_font
        ws[cell].alignment = heading_align
        if fill:
            ws[cell].fill  = fill
        ws[cell].border    = box_border

        ws.column_dimensions[cell[0]].width = width

    heading_font  = Font(name='Arial', size=10)
    heading_align = outcome_align

    join_fill     = PatternFill(fgColor='D9E9D4', fill_type='solid')
    not_int_fill  = PatternFill(fgColor='FFF1CE', fill_type='solid')
    no_resp_fill  = PatternFill(fgColor='FEC7CE', fill_type='solid')

    row += 1
    _make_heading(f'A{row}', 'Name', 20)
    _make_heading(f'B{row}', 'Envelope ID', 9)
    _make_heading(f'C{row}', 'Email addresses', 35)
    _make_heading(f'D{row}', 'Date of email', 10)
    _make_heading(f'E{row}', 'Phone numbers', 25)
    _make_heading(f'F{row}', 'Date of phone call', 10)
    join_col = 'G'
    _make_heading(f'{join_col}{row}', 'Join ministry', 10, join_fill)
    not_int_col = 'H'
    _make_heading(f'{not_int_col}{row}', 'No longer interested', 10, not_int_fill)
    no_resp_col = 'I'
    _make_heading(f'{no_resp_col}{row}', 'No response / followup', 10, no_resp_fill)

    #---------------------------------------------------------------------

    ws.freeze_panes = ws[f'A{row+1}']

    #---------------------------------------------------------------------

    def _set(cell, value):
        ws[cell]           = value
        ws[cell].font      = value_font
        ws[cell].alignment = value_align
        ws[cell].border    = box_border

    def _set_color(cell, fill):
        ws[cell].fill      = fill
        ws[cell].alignment = outcome_align

    def _find_phones(entity, phones):
        key = 'phones'
        if key not in entity:
            return

        for phone in entity[key]:
            if 'emergency' not in phone['type'].lower():
                item = {
                    'number' : phone['number'],
                    'string' : phone['number'],
                    'type'   : '',
                }
                if phone['type']:
                    item['string'] += f" ({phone['type']})"
                    item['type']    = phone['type']

                phones[phone['number']] = item

    value_font  = Font(name='Arial', size=10)
    value_align = Alignment(wrap_text=True)

    for member_name in sorted(interested):
        member = interested[member_name]
        log.debug(f"- Member: {member['email_name']} (MID: {member['MemRecNum']})")

        row += 1

        _set(f'A{row}', member['email_name'])
        _set(f'B{row}', member['family']['ParKey'])

        value = ''
        num_emails = 0
        for email in PDSChurch.find_any_email(member):
            if value:
                value += '\r\n'
            value += email
            num_emails += 1
        _set(f'C{row}', value)

        # Take non-Emergency phone numbers
        phones = dict()
        _find_phones(member['family'], phones)
        _find_phones(member, phones)
        num_phones = 0
        if len(phones) > 0:
            value = ''
            for num in sorted(phones):
                phone = phones[num]
                if value:
                    value += '\r\n'
                value += phone['string']
                num_phones += 1
            _set(f'E{row}', value)

        # Add cell borders for cells that have no values
        for column in ['D', 'F', 'G', 'H', 'I']:
            ws[f"{column}{row}"].border = box_border

        # Set colors
        _set_color(f'{join_col}{row}', join_fill)
        _set_color(f'{not_int_col}{row}', not_int_fill)
        _set_color(f'{no_resp_col}{row}', no_resp_fill)

        # Set the row height
        num_lines = num_emails if num_emails > num_phones else num_phones
        num_lines = 1 if num_emails == 0 else num_lines
        actual_height = 15 * num_lines
        if actual_height < 30:
            ws.row_dimensions[row].height = 30

    #---------------------------------------------------------------------

    wb.save(filename)
    log.info(f'Wrote {filename}')

##############################################################################

def write_all_interested(all_interested, log):
    for ministry_name in sorted(all_interested):
        write_xlsx(ministry_name, all_interested[ministry_name], log)

##############################################################################

def find_interested(members, log):
    out = dict()

    for mid, member in members.items():
        log.debug(f"Checking member: {member['email_name']}")
        key = 'inactive_ministries'
        if key not in member:
            continue

        # We're looking for Members who:
        # - have status "Interested"
        # - have a start date >= 2020
        # - do not have an end date
        for ministry in member[key]:
            log.debug(f"  Checking ministry {ministry['Description']}")
            if ministry['status'] != 'Interested':
                log.debug(f"    Status is {ministry['status']} -- skipping")
                continue
            elif ministry['end'] != PDSChurch.date_never:
                log.debug(f"    End date is {ministry['end']} -- skipping")
                continue
            elif ministry['start'] < start_of_interest:
                log.debug(f"    Start date is {ministry['start']} -- skipping")
                continue

            log.debug("    Interested!")
            key = ministry['Description']
            if key not in out:
                out[key] = dict()

            # Stored by ministry, then by Member name.
            # Doing it this way (vs. via appending to lists) so that we
            # can sort by both keys and have repeatable, reproducible output.
            out[key][member['Name']] = member

    return out

##############################################################################

def setup_args():
    parser = argparse.ArgumentParser()
    parser.add_argument('--debug',
                                action='store_true')

    global gapp_id
    parser.add_argument('--app-id',
                                 default=gapp_id,
                                 help='Filename containing Google application credentials')
    global guser_cred_file
    parser.add_argument('--user-credentials',
                                 default=guser_cred_file,
                                 help='Filename containing Google user credentials')

    args = parser.parse_args()

    return args

##############################################################################

def main():
    args = setup_args()
    log = ECC.setup_logging(debug=args.debug)

    #---------------------------------------------------------------

    log.info("Reading PDS data...")
    (pds, families,
     members) = PDSChurch.load_families_and_members(filename='pdschurch.sqlite3',
                                                    parishioners_only=True,
                                                    log=log)

    #---------------------------------------------------------------

    interested = find_interested(members, log)
    write_all_interested(interested, log)

main()
