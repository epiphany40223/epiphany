#!/usr/bin/env python3

# Make sure to pip install everything in requirements.txt.

import sys
sys.path.insert(0, '../../python')

import collections
import traceback
import datetime
import argparse
import smtplib
import csv
import os
import re

import ECC
import Google
import PDSChurch
import GoogleAuth

import helpers

from datetime import date
from datetime import datetime
from datetime import timedelta

from oauth2client import tools
from apiclient.http import MediaFileUpload
from email.message import EmailMessage

import matplotlib
import matplotlib.dates as mdates
import matplotlib.pyplot as plt

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment

#------------------------------------------------------------------------------

from constants import already_submitted_fam_status

from constants import stewardship_begin_date
from constants import stewardship_end_date

from constants import gapp_id
from constants import guser_cred_file

from constants import jotform_gsheet_gfile_id
from constants import jotform_gsheet_columns
from constants import upload_team_drive_folder_id
from constants import gsheet_editors

from constants import stewardship_year
from constants import title

from constants import smtp_server
from constants import smtp_from

from constants import jotform

from constants import MAX_PDS_FAMILY_MEMBER_NUM

##############################################################################

ecc = '@epiphanycatholicchurch.org'

# Comments report email
comments_email_to = f'angie{ecc},mary{ecc},jeff@squyres.com'
comments_email_subject = 'Comments report'

# Statistics report email
statistics_email_to = f'sdreiss71@gmail.com,angie{ecc},mary{ecc},jeff@squyres.com'
statistics_email_subject = 'Statistics report'

fid_participation_email_to = f'mary{ecc},jeff@squyres.com'

pledge_email_to = fid_participation_email_to
pledge_email_subject = 'Pledge PDS CSV import file'

# JMS for debugging/testing
#statistics_email_to = 'jsquyres@gmail.com'
#comments_email_to = 'jsquyres@gmail.com'
#fid_participation_email_to = 'jsquyres@gmail.com'
#pledge_email_to = 'jsquyres@gmail.com'


##############################################################################

def _upload_to_gsheet(google, google_folder_id, google_filename, mime_type, local_filename, remove_local, log):
    try:
        log.info(f'Uploading file to google "{local_filename}" -> "{google_filename}"')
        metadata = {
            'name'     : google_filename,
            'mimeType' : Google.mime_types['sheet'],
            'parents'  : [ google_folder_id ],
            'supportsTeamDrives' : True,
            }
        media = MediaFileUpload(local_filename,
                                mimetype=Google.mime_types[mime_type],
                                resumable=True)
        file = google.files().create(body=metadata,
                                     media_body=media,
                                     supportsTeamDrives=True,
                                     fields='id').execute()
        log.debug(f'Successfully uploaded file: "{google_filename}" (ID: {file["id"]})')

    except:
        log.error('Google upload failed for some reason:')
        log.error(traceback.format_exc())
        exit(1)

    # Set permissions on the GSheet to allow the
    # workers group to edit the file (if you are view-only, you
    # can't even adjust the column widths, which will be
    # problematic for the comments report!).
    try:
        perm = {
            'type': 'group',
            'role': 'writer',
            'emailAddress': gsheet_editors,
        }
        out = google.permissions().create(fileId=file['id'],
                                          supportsTeamDrives=True,
                                          sendNotificationEmail=False,
                                          body=perm,
                                          fields='id').execute()
        log.debug(f"Set Google permission for file: {id}")
    except:
        log.error('Google set permission failed for some reason:')
        log.error(traceback.format_exc())
        exit(1)

    # Remove the temp file when we're done
    if remove_local:
        try:
            os.remove(local_filename)
        except:
            pass

    return file['id']

#-----------------------------------------------------------------------------

def _make_filenames(filename, extension):
    extension = f".{extension}"
    if filename.endswith(extension):
        local_filename  = filename
        google_filename = filename[:-len(extension)]
    else:
        local_filename  = f"{filename}{extension}"
        google_fielname = filename

    return local_filename, google_filename

#-----------------------------------------------------------------------------

def upload_csv_to_gsheet(google, google_folder_id, filename, fieldnames, csv_rows, remove_local, log):
    if csv_rows is None or len(csv_rows) == 0:
        return None, None

    # First, write out a CSV file
    csv_filename, google_filename = _make_filenames(filename, 'csv')
    try:
        os.remove(csv_filename)
    except:
        pass

    csvfile = open(csv_filename, 'w')
    writer  = csv.DictWriter(csvfile, fieldnames=fieldnames,
                            quoting=csv.QUOTE_ALL)
    writer.writeheader()
    for row in csv_rows:
        writer.writerow(row)
    csvfile.close()

    # Now upload that file to Google Drive
    id = _upload_to_gsheet(google,
                        google_folder_id=google_folder_id,
                        google_filename=google_filename,
                        mime_type='csv',
                        local_filename=csv_filename,
                        remove_local=remove_local,
                        log=log)

    return id, None if remove_local else csv_filename

#-----------------------------------------------------------------------------

def upload_xlsx_to_gsheet(google, google_folder_id, filename, workbook, remove_local, log):
    # First, write out the XLSX file
    xlsx_filename, google_filename = _make_filenames(filename, 'xlsx')
    try:
        os.remove(xlsx_filename)
    except:
        pass
    workbook.save(xlsx_filename)

    # Now upload that file to Google Drive
    id = _upload_to_gsheet(google,
                        google_folder_id=google_folder_id,
                        # For some reason, when we upload an XLSX and want
                        # it to convert to a Google Sheet, the google filename
                        # must end in .xlsx.  Sigh.
                        google_filename=xlsx_filename,
                        mime_type='sheet',
                        local_filename=xlsx_filename,
                        remove_local=remove_local,
                        log=log)

    return id, None if remove_local else xlsx_filename

##############################################################################

def _change(label, old_value, new_value, message):
    return {
        'label'     : label,
        'old_value' : old_value,
        'new_value' : new_value,
        'message'   : message,
    }

def _compare(changes, label, jot_value, pds_value):
    if jot_value is None:
        jot_value = ''
    if pds_value is None:
        pds_value = ''

    if jot_value.strip() == pds_value.strip():
        return

    message = ('{label}: {new_value}'
               .format(label=label, new_value=jot_value))
    changes.append(_change(label=label,
                           old_value=pds_value,
                           new_value=jot_value,
                           message=message))

##############################################################################

def comments_to_xlsx(google, jotform_data, id_field, emails_field, name_field, env_field, workbook, log):
    sheet = workbook.active

    comments_label    = "Comments"
    pledge_last_label = f'CY{stewardship_year-1} pledge'
    pledge_cur_label  = f'CY{stewardship_year} pledge'
    amount_label      = f'CY{stewardship_year-1} amount'

    # Setup the title row
    # Title rows + set column widths
    title_font   = Font(color='FFFF00')
    title_fill   = PatternFill(fgColor='0000FF', fill_type='solid')
    title_align  = Alignment(horizontal='center', wrap_text=True)

    wrap_align   = Alignment(horizontal='general', wrap_text=True)
    right_align  = Alignment(horizontal='right')

    money_format = "$###,###,###"

    xlsx_cols = dict();
    def _add_col(name, width=10):
        col             = len(xlsx_cols) + 1
        xlsx_cols[name] = {'name' : name, 'column' : col, 'width' : width }
    _add_col('Date', width=20)
    _add_col('FID')
    _add_col('Envelope')
    _add_col('Family names', width=30)
    _add_col('Emails', width=50)
    _add_col(pledge_last_label)
    _add_col(pledge_cur_label)
    _add_col(amount_label)
    _add_col('Comments', width=100)

    for data in xlsx_cols.values():
        col            = data['column']
        cell           = sheet.cell(row=1, column=col, value=data['name'])
        cell.fill      = title_fill
        cell.font      = title_font
        cell.alignment = title_align

        col_char = chr(ord('A') - 1 + col)
        sheet.column_dimensions[col_char].width = data['width']

    sheet.freeze_panes = sheet['A2']

    #-------------------------------------------------------------------

    def _extract_money_string(val):
        if val.startswith('$'):
            val = val[1:]
        val = int(val.replace(',', ''))

        if val != 0:
            return int(val), money_format
        else:
            return 0, None

    #-------------------------------------------------------------------

    # Now fill in all the data rows
    xlsx_row = 2
    for row in jotform_data:
        # Skip if the comments are empty
        if comments_label not in row:
            continue
        if row[comments_label] == '':
            continue

        if row[pledge_last_label]:
            pledge_last, pledge_last_format = _extract_money_string(row[pledge_last_label])
        else:
            pledge_last = 0
            pledge_last_format = None

        if row[pledge_cur_label]:
            pledge_cur = int(row[pledge_cur_label])
            pledge_cur_format = money_format
        else:
            pledge_cur = 0
            pledge_cur_format = None

        if row[amount_label]:
            amount, amount_format = _extract_money_string(row[amount_label])
        else:
            amount = 0
            amount_format = None

        def _fill(col_name, value, align=None, format=None):
            col_data = xlsx_cols[col_name]
            cell     = sheet.cell(row=xlsx_row, column=col_data['column'], value=value)
            if align:
                cell.alignment = align
            if format:
                cell.number_format = format

        _fill('Date', row['SubmitDate'])
        _fill('FID', int(row[id_field]))
        # Do NOT convert the envelope ID to an int -- the leading zeros
        # are significant.
        _fill('Envelope', row[env_field])
        _fill('Family names', row[name_field])
        _fill('Emails', row[emails_field])
        _fill(pledge_last_label, pledge_last, format=pledge_last_format)
        _fill(pledge_cur_label, pledge_cur, format=pledge_cur_format)
        _fill(amount_label, amount, format=amount_format)
        _fill('Comments', row[comments_label], align=wrap_align)

        xlsx_row += 1

    # Return the number of comments we found
    return xlsx_row - 2

def reorder_rows_by_date(jotform_data):
    data = dict()
    for row in jotform_data:
        data[row['SubmitDate']] = row

    out = list()
    for row in sorted(data):
        out.append(data[row])

    return out

def comments_report(args, google, start, end, time_period, jotform_data, log):
    log.info("Composing comments report...")

    # jotform_data is a list of rows, in order by FID.  Re-order them to be
    # ordered by submission date.
    ordered_data = reorder_rows_by_date(jotform_data)

    # Examine the jotform data and see if there are any comments that
    # need to be reported
    workbook = Workbook()
    num_comments = comments_to_xlsx(google, jotform_data=ordered_data,
                     id_field='fid', name_field='Family names',
                     env_field='EnvId', emails_field='Emails to reply to',
                     workbook=workbook, log=log)

    # If we have any comments, upload them to a Gsheet
    gsheet_id = None
    sheet = workbook.active
    if num_comments > 0:
        filename   = f'Comments {time_period}.xlsx'
        gsheet_id, _ = upload_xlsx_to_gsheet(google,
                                        google_folder_id=upload_team_drive_folder_id,
                                        filename=filename,
                                        workbook=workbook,
                                        remove_local=True,
                                        log=log)

    # Send the comments report email
    body = list()
    body.append(f"""<html>
<body>
<h2>{title} comments report</h2>

<h3>Time period: {time_period}</h3>""")

    if num_comments == 0:
        body.append("<p>No comments submitted during this time period.</p>")
    else:
        url = 'https://docs.google.com/spreadsheets/d/{id}'.format(id=gsheet_id)
        body.append(f"""<p><a href="{url}">Link to Google sheet containing comments for this timeframe</a>.</p>
<p>There are {num_comments} comments in this report.</p>""")

    body.append("""
</body>
</html>""")

    to = comments_email_to
    subject = '{subj} ({t})'.format(subj=comments_email_subject, t=time_period)
    try:
        log.info('Sending "{subject}" email to {to}'
                 .format(subject=subject, to=to))
        with smtplib.SMTP_SSL(host=smtp_server,
                              local_hostname='epiphanycatholicchurch.org') as smtp:
            msg = EmailMessage()
            msg['Subject'] = subject
            msg['From'] = smtp_from
            msg['To'] = to
            msg.set_content('\n'.join(body))
            msg.replace_header('Content-Type', 'text/html')

            # This assumes that the file has a single line in the format of username:password.
            with open(args.smtp_auth_file) as f:
                line = f.read()
                smtp_username, smtp_password = line.split(':')

            # Login; we can't rely on being IP whitelisted.
            try:
                smtp.login(smtp_username, smtp_password)
            except Exception as e:
                log.error(f'Error: failed to SMTP login: {e}')
                exit(1)

            smtp.send_message(msg)
    except:
        print("==== Error with {email}".format(email=to))
        print(traceback.format_exc())

##############################################################################

def statistics_compute(pds_families, unique_fid_jotform, log):
    ret = dict()

    # Total number of active Families in the parish
    ret['num_active'] = len(pds_families)

    #---------------------------------------------------------------

    # Number of active Families who are eligible for electronic stewardship
    # (i.e., we have an email address for the spouse and/or HoH)

    eligible = dict()
    for fid, family in pds_families.items():
        for member in family['members']:
            if helpers.member_is_hoh_or_spouse(member):
                em = PDSChurch.find_any_email(member)
                if len(em) == 0:
                    continue
                eligible[fid] = True
                continue

    ret['num_eligible'] = len(eligible)

    #-----------------------------------------------------------

    # The unique_fid_jotform dictionary we have will have, at most, 1 entry per
    # FID.  So we can just take the length of it to know how many (unique)
    # families have submitted electronically.
    ret['num_electronic'] = len(unique_fid_jotform)

    #-----------------------------------------------------------

    # Build a cross-reference of which families have submitted electronically
    fids_electronic = dict()
    for row in unique_fid_jotform:
        fid = int(row['fid'])
        fids_electronic[fid] = True

    fids_paper_or_electronic = fids_electronic.copy()

    # - Count how many submitted paper
    # - Count how many submitted paper and electronic
    # - Count how many submitted paper or electronic
    fids_paper = dict()
    fids_paper_and_electronic = dict()
    for fid, family in pds_families.items():
        if 'status' in family and family['status'] == already_submitted_fam_status:
            fids_paper[fid] = True
            fids_paper_or_electronic[fid] = True

            if fid in fids_electronic:
                fids_paper_and_electronic[fid] = True

    ret['num_paper']                = len(fids_paper)
    ret['num_paper_or_electronic']  = len(fids_paper_or_electronic)
    ret['num_paper_and_electronic'] = len(fids_paper_and_electronic)

    #-----------------------------------------------------------

    return ret

#------------------------------------------------------------------------

def statistics_graph(pds_members, pds_families, jotform, log):
    def _find_range(data, earliest, latest):
        for row in data:
            submitted = row['SubmitDate']
            submitted_dt = helpers.jotform_date_to_datetime(submitted)
            if submitted_dt < earliest:
                earliest = submitted_dt
            if submitted_dt > latest:
                latest = submitted_dt

        return earliest, latest

    #------------------------------------------------------------------------

    def _compute(start, end, pds_members, pds_families,
                 jotform, log):
        # Compute these values just in the date range:
        # - How many unique family submissions total?

        family_submitted = dict()

        # Check electronic submissions
        for row in jotform:
            if row['fid'] == 'fid':
                continue # Skip title row

            # Is this row in our date range?
            dt = helpers.jotform_date_to_datetime(row['SubmitDate'])
            if dt < start or dt > end:
                continue

            fid = int(row['fid'])
            log.debug(f"Found submission in our date window: {fid} on {dt}")

            # Make sure the family hasn't been deleted
            if fid not in pds_families:
                continue

            family_submitted[fid] = True

        return len(family_submitted)

    #------------------------------------------------------------------------

    one_day  = timedelta(days=1)
    earliest = datetime(year=9999, month=12, day=31)
    latest   = datetime(year=1971, month=1,  day=1)
    earliest, latest = _find_range(jotform, earliest, latest)

    earliest = datetime(year=earliest.year, month=earliest.month, day=earliest.day)
    latest   = datetime(year=latest.year, month=latest.month, day=latest.day)
    latest   += one_day - timedelta(seconds=1)

    log.info(f"Earliest: {earliest}")
    log.info(f"Latest:   {latest}")

    day             = earliest
    dates           = list()
    data_per_day    = list()
    data_cumulative = list()

    # Make lists that we can give to matplotlib for plotting
    while day <= latest:
        log.debug(f"Get per-day stats for {day.date()}")
        per_day    = _compute(day, day + one_day,
                              pds_members, pds_families,
                              jotform, log)
        log.debug(f"Get cumulative stats for {earliest} - {day + one_day}")
        cumulative = _compute(earliest, day + one_day,
                              pds_members, pds_families,
                              jotform, log)

        log.debug(f"Date: {day}: per day {per_day}, cumulative {cumulative}")

        dates.append(day.date())
        data_per_day.append(per_day)
        data_cumulative.append(cumulative)

        day += one_day

    # Make the plot
    fig, ax = plt.subplots()

    # The X label of "days" is obvious, there's no good Y label since we have
    # multiple lines
    n = datetime.now()
    hour = n.hour if n.hour > 0 else 12
    ampm = 'am' if n.hour < 12 else 'pm'
    plt.title("As of {month} {day}, {year} at {hour}:{min:02}{ampm}"
            .format(month=n.strftime("%B"), day=n.day, year=n.year,
                    hour=hour, min=n.minute, ampm=ampm))
    plt.suptitle(title + " electronic submissions statistics")

    ax.plot(dates, data_per_day, label='Unique family electronic submissions per day')
    ax.plot(dates, data_cumulative, label='Cumulative unique family electronic submissions')

    ax.get_xaxis().set_major_locator(mdates.DayLocator(interval=1))
    ax.get_xaxis().set_major_formatter(mdates.DateFormatter("%a %b %d"))
    plt.setp(ax.get_xticklabels(), rotation=20, ha="right",
             rotation_mode="anchor")
    ax.grid()

    plt.legend(loc="upper left")

    # Make a filename based on the time that this was run
    filename = ('statistics-as-of-{year:04}-{month:02}-{day:02}-{hour:02}{min:02}{sec:02}.pdf'
                .format(year=n.year, month=n.month, day=n.day,
                        hour=n.hour, min=n.minute, sec=n.second))
    fig.savefig(filename)

    plt.close(fig)

    return filename

#-----------------------------------------------------------------------------

def statistics_report(args, end, pds_members, pds_families, jotform, log):
    log.info("Composing statistics report...")

    #---------------------------------------------------------------

    graph_filename = statistics_graph(pds_members, pds_families, jotform, log)
    data = statistics_compute(pds_families, jotform, log)

    #---------------------------------------------------------------------

    electronic_percentage           = data['num_electronic'] / data['num_eligible'] * 100
    paper_percentage                = data['num_paper_and_electronic'] / data['num_active'] * 100
    paper_and_electronic_percentage = data['num_paper_and_electronic'] / data['num_active'] * 100
    paper_or_electronic_percentage  = data['num_paper_or_electronic'] / data['num_active'] * 100

    # Send the statistics report email
    body = list()
    body.append(f"""<html>
<body>
<h2>{title} statistics update</h2>

<h3>Time period: through {end}</h3>
<ul>
<li> Total number of active PDS Families in the parish: {data['num_active']:,d}</li>
<br>
<li> Number of active PDS Families eligible for electronic stewardship: {data['num_eligible']:,d}
    <ul>
    <li>This means that we have an email address in PDS for the Head of Household and/or the Spouse of a given Family</li>
    </ul></li>
<br>
<li> Number of electronic-stewardship-eligible Families who have electronically submitted: {data['num_electronic']:,d} (out of {data['num_eligible']:,d}, or {electronic_percentage:.1f}%)
    <ul>
    <li>This is the number of families who submitted electronically.</li>
    </ul></li>
<br>
<li> Number of active PDS Families with "{already_submitted_fam_status}" status: {data['num_paper']:,d} (out of {data['num_active']:,d}, or {paper_percentage:.1f}%)
    <ul>
    <li>This is the number of families who submitted a paper card.</li>
    </ul></li>
<br>
<li> Number of active PDS Families who have electronically completed their submission <em>and</em> have "{already_submitted_fam_status}" status: {data['num_paper_and_electronic']:,d} (out of {data['num_active']:,d}, or {paper_and_electronic_percentage:.1f}%)
    <ul>
    <li>These are people who submitted twice -- paper and electronically!</li>
    </ul></li>
<br>
<li> Number of active PDS Families who have electronically completed their submission <em>or</em> have "{already_submitted_fam_status}" status: {data['num_paper_or_electronic']:,d} (out of {data['num_active']:,d}, or {paper_or_electronic_percentage:.1f}%)
    <ul>
    <li>This is the total number of active Families who have submitted.</li>
    </ul></li>
</ul>
</body>
</html>""")

    to = statistics_email_to
    time_period = "through {end}".format(end=end)
    subject = '{subj} ({t})'.format(subj=statistics_email_subject, t=time_period)
    try:
        log.info('Sending "{subject}" email to {to}'
                 .format(subject=subject, to=to))
        with smtplib.SMTP_SSL(host=smtp_server,
                              local_hostname='epiphanycatholicchurch.org') as smtp:
            msg = EmailMessage()
            msg['Subject'] = subject
            msg['From'] = smtp_from
            msg['To'] = to
            msg.set_content('\n'.join(body))
            msg.replace_header('Content-Type', 'text/html')

            # This assumes that the file has a single line in the format of username:password.
            with open(args.smtp_auth_file) as f:
                line = f.read()
                smtp_username, smtp_password = line.split(':')

            # Login; we can't rely on being IP whitelisted.
            try:
                smtp.login(smtp_username, smtp_password)
            except Exception as e:
                log.error(f'Error: failed to SMTP login: {e}')
                exit(1)

            with open(graph_filename, "rb") as f:
                csv_data = f.read()
            msg.add_attachment(csv_data, maintype='application', subtype='pdf',
                            filename=graph_filename)
            os.unlink(graph_filename)

            smtp.send_message(msg)
    except:
        print("==== Error with {email}".format(email=to))
        print(traceback.format_exc())

##############################################################################

# Columns that I need
# - fid
# - Recurring Charge Name:
# - Terms / Frequency: Weekly, Biweekly, Monthly, Bimonthly, Semimonthly, Quarterly, Semiannually, Annually
# - Begin Date: 01/01/2020
# - End Date: 12/31/2020
# - Rate
# - Total pledge
def convert_pledges_to_pds_import(pds_families, jotform, log):
    def _map_to_freqrate(pledge):
        freq = pledge[f'CY{stewardship_year} frequency']
        amount = float(pledge[f'CY{stewardship_year} pledge'])

        if freq == 'Weekly donations':
            return 'Weekly', amount / 52
        elif freq == 'Monthly donations':
            return 'Monthly', amount / 12
        elif freq == 'Quarterly donations':
            return 'Quarterly', amount / 4
        elif freq == 'One annual donation':
            return 'Annually', amount
        else:
            return None, None

    #-------------------------------------------------------------------------

    out = list()

    for pledge in jotform:
        # Skip the title row
        fid = pledge['fid']
        if 'fid' in fid:
            continue

        # Here's something that can happen: A family may be deleted from PDS
        # even though they submitted.  In this case, skip them.
        fid = int(fid)
        if fid not in pds_families:
            log.warn("WARNING: Family FID {fid} / {names} submitted a pledge, but is no longer in PDS"
                .format(fid=pledge['fid'], names=pledge['Names']))
            continue

        # If there is a $0 pledge, Per Lynne's comment, we'll
        # transform this into a $1 annual pledge -- just so that this
        # person is on the books, so to speak.
        pledge_field  = f'CY{stewardship_year} pledge'
        pledge_amount = pledge[pledge_field]
        if not pledge_amount or pledge_amount == '' or int(pledge_amount) == 0:
            pledge_amount = 1
            pledge[pledge_field] = pledge_amount
            pledge[f'CY{stewardship_year} frequency'] = 'One annual donation'

        frequency, rate = _map_to_freqrate(pledge)

        # Round pledge value and rate to 2 decimal points, max
        rate  = float(int(rate * 100)) / 100.0
        total = float(int(pledge_amount) * 100) / 100.0

        # Use an OrderedDict to keep the fields in order
        row = collections.OrderedDict()
        row['fid']           = pledge['fid']
        row['RecChargeName'] = 'Due/Contributions'
        row['Frequency']     = frequency
        row['BeginDate']     = stewardship_begin_date.strftime('%m/%d/%Y')
        row['EndDate']       = stewardship_end_date.strftime('%m/%d/%Y')
        row['PledgeRate']    = rate
        row['TotalPledge']   = total
        row['SubmitDate']    = pledge['SubmitDate']
        row['Names']         = pledge['Family names']
        row['Envelope ID']   = helpers.pkey_url(pledge['EnvId'])

        # Calculate family pledge values for last CY
        family = pds_families[fid]
        year = f'{stewardship_year - 2000 - 2:02}'
        helpers.calculate_family_values(family, year, log)

        row[f'CY{stewardship_year - 2} YTD gifts'] = family['calculated']['contributed']

        # Calculate family pledge values for this CY
        year = f'{stewardship_year - 2000 - 1:02}'
        helpers.calculate_family_values(family, year, log)

        row[f'CY{stewardship_year - 1} YTD gifts'] = family['calculated']['contributed']
        row[f'CY{stewardship_year - 1} pledge']    = family['calculated']['pledged']

        # Add column for how they want to fullfill their pledge
        row[f'CY{stewardship_year} frequency'] = pledge[f'CY{stewardship_year} frequency']
        row[f'CY{stewardship_year} mechanism'] = pledge[f'CY{stewardship_year} mechanisms']

        # Add a column for this Family's "Envelope user" value
        row['PDS Envelope User'] = family['EnvelopeUser']

        # Add a column for whether the Family selected the "offeratory envelopes" option on the Jotform.  A handful of people also typed something that included the word "offeratory", so include those, too.
        val = False
        field = f'CY{stewardship_year} mechanisms'
        if re.search(pledge[field], 'Offeratory'):
            val = True
        row['Jotform asked for Envelopes'] = val

        out.append(row)

    return out

#-----------------------------------------------------------------------------

def family_pledge_csv_report(args, google, pds_families, jotform, log):
    pledges = convert_pledges_to_pds_import(pds_families, jotform, log)

    # If we have pledges, upload them to a Google sheet
    gsheet_id = None
    if len(pledges) > 0:
        filename = 'Family Pledge PDS import.csv'
        gsheet_id, csv_filename = upload_csv_to_gsheet(google,
                                        google_folder_id=upload_team_drive_folder_id,
                                        filename=filename,
                                        fieldnames=pledges[0].keys(),
                                        csv_rows=pledges,
                                        remove_local=False,
                                        log=log)

    # Send the statistics report email
    body = list()
    body.append(f"""<html>
<body>
<h2>{title} pledge update</h2>

""")

    if gsheet_id:
        url = 'https://docs.google.com/spreadsheets/d/{id}'.format(id=gsheet_id)
        body.append(f'<p><a href="{url}">Link to Google sheet containing pledge updates for this timeframe</a>.</p>')
        body.append("<p>See the attachment for a CSV to import directly into PDS.</p>")
    else:
        body.append("<p>There were no pledge submissions during this timeframe.<p>")

    body.append("""</body>
</html>""")

    try:
        to = pledge_email_to
        log.info(f'Sending "{pledge_email_subject}" email to {to}')
        with smtplib.SMTP_SSL(host=smtp_server,
                              local_hostname='epiphanycatholicchurch.org') as smtp:

            # This assumes that the file has a single line in the format of username:password.
            with open(args.smtp_auth_file) as f:
                line = f.read()
                smtp_username, smtp_password = line.split(':')

            # Login; we can't rely on being IP whitelisted.
            try:
                smtp.login(smtp_username, smtp_password)
            except Exception as e:
                log.error(f'Error: failed to SMTP login: {e}')
                exit(1)

            msg = EmailMessage()
            msg['Subject'] = pledge_email_subject
            msg['From'] = smtp_from
            msg['To'] = to
            msg.set_content('\n'.join(body))
            msg.replace_header('Content-Type', 'text/html')

            # If there were results, attach the CSV
            if gsheet_id:
                with open(filename, "rb") as f:
                    csv_data = f.read()
                msg.add_attachment(csv_data, maintype='text', subtype='csv', filename=filename)

            smtp.send_message(msg)

            if gsheet_id:
                os.unlink(filename)
    except:
        print("==== Error with {email}".format(email=to))
        print(traceback.format_exc())

##############################################################################

MINISTRY_PARTICIPATE = 1
MINISTRY_INTERESTED  = 2
MINISTRY_INACTIVE    = 3

def _convert_jotform_ministry_status(jotform_status):
    if jotform_status is None:
        return MINISTRY_INACTIVE

    if 'PARTICIPATE' in jotform_status:
        return MINISTRY_PARTICIPATE
    elif 'INTERESTED' in jotform_status:
        return MINISTRY_INTERESTED
    elif 'NO LONGER' in jotform_status:
        return MINISTRY_INACTIVE
    else:
        # Catch-all
        return MINISTRY_INACTIVE

def _status_to_str(status):
    if status == MINISTRY_PARTICIPATE:
        return 'I already participate'
    elif status == MINISTRY_INTERESTED:
        return 'I am interested'
    else:
        return 'I do not participate'

#-----------------------------------------------------------------------------

# Fields I need:
# - MID
# - Ministry
# - Status
# - Start date (from constants.py)
# - End date (from constants.py)
# Other fields requested by the staff:
# - Member nickname + last name
# - Member email addresses
# - Member phones
# - Member status on ministry
#
# Will generate several outputs:
# - Likely can be directly imported
#   - NOT PARTICIPATE -> INTERESTED
# - Likely cannot be directly imported (still need to test)
#   - PARTICPATE -> NOT PARTICIPATE
# - Needs to be examined by a human
#   - NOT PARTICIPATE -> ALREADY PARTICIPATE
#   - PARTICIPATE -> INTERESTED
#
# Output will be a dictionary:
# - PDS ministry name (in the case of multi-names, like lectors, use the jotform name): dictionary of
#   - interested: list of members
#   - no_longer_interested: list of members
#   - needs_human: list of members
def analyze_member_ministry_submissions(pds_members, pds_families, jotform_csv, log):
    def _get_pds_member_ministry_status(member, ministry_names):
        for ministry in member['active_ministries']:
            if ministry['Description'] in ministry_names:
                if ministry['active']:
                    return MINISTRY_PARTICIPATE, ministry['status']
                else:
                    return MINISTRY_INACTIVE, ministry['status']

        # If the member is not currently active in this ministry, see if they
        # were ever previously a member of this ministry
        results = dict()
        for ministry in member['inactive_ministries']:
            if ministry['Description'] in ministry_names:
                results[ministry['start']] = ministry['status']
                results[ministry['end']]   = ministry['status']

        if len(results) > 0:
            dates = sorted(results.keys())
            end   = dates[-1]
            result_str = f'Inactive, last status on {end}: {results[end]}'
        else:
            result_str = 'Never been a member of this ministry'

        return MINISTRY_INACTIVE, result_str

    #-------------------------------------------------------------------------
    def _status_to_pds(status):
        if status == MINISTRY_PARTICIPATE:
            return 'Actively Involved'
        elif status == MINISTRY_INACTIVE:
            return 'No Longer Involved'
        elif status == MINISTRY_INTERESTED:
            return 'Interested'

    #-------------------------------------------------------------------------

    output = dict()

    # Each row is a family
    for jrow in jotform_csv:
        fid = int(jrow['fid'])

        # Make sure the family is still active
        if fid not in pds_families:
            log.warn(f"WARNING: Family {fid} submitted, but cannot be found -- skipping")
            continue

        family = pds_families[fid]
        log.info(f"Processing Jotform Family submission: {family['Name']} (FID {fid})")

        # Check the members in this row
        for member_num in range(MAX_PDS_FAMILY_MEMBER_NUM):
            column_names = jotform_gsheet_columns['members'][member_num]
            # The 0th entry in each Member is the MID
            mid = jrow[column_names[0]]

            # If there's no MID for this Member, then there's no Member.
            # We're done with the loop.
            if mid == '':
                break

            # Here's something that can happen: a MID was submitted, but is no
            # longer an active member in PDS.  In that case, ignore the submission.
            mid = int(mid)
            if mid not in pds_members:
                log.warn(f"WARNING: Member {mid} submitted, but cannot be found -- skipping")
                continue

            member = pds_members[mid]
            log.info(f"  Processing Jotform member {member['email_name']} (MID {mid})")

            # JMS Debug
            #if not m['Name'].startswith('Squyres,Jeff'):
            #    continue

            # Go through the list of ministry grids from the jotform
            for grid in jotform.ministry_grids:
                # Go through the rows in the jotform ministry grid
                for mrow in grid.rows:
                    # Each row has its PDS ministry name
                    ministry_entry = mrow['pds_ministry']

                    # Some ministry rows are lists because we want to treat them equivalently
                    if type(ministry_entry) is list:
                        output_key = mrow['row_heading']
                        ministries = ministry_entry
                    else:
                        output_key = ministry_entry
                        ministries = [ ministry_entry ]

                    # Get the Member's status in this Ministry from the jotform
                    jotform_column_name = mrow['jotform_columns'][member_num]
                    jotform_status_str  = jrow[jotform_column_name]
                    jotform_status      = _convert_jotform_ministry_status(jotform_status_str)

                    # Get their status from PDS
                    pds_status, pds_status_string = _get_pds_member_ministry_status(member, ministries)
                    key = 'jotform'
                    if key not in member:
                        member[key] = dict()
                    member[key][output_key] = pds_status_string

                    # If they're the same, nothing to do
                    if jotform_status == pds_status:
                        continue

                    if output_key not in output:
                        output[output_key] = dict()

                    # If PDS INACTIVE -> INTERESTED
                    if (pds_status == MINISTRY_INACTIVE and
                          jotform_status == MINISTRY_INTERESTED):
                        key = 'Interested'
                        if key not in output[output_key]:
                            output[output_key][key] = list()
                        output[output_key][key].append(member)

                    elif (pds_status == MINISTRY_PARTICIPATE and
                          jotform_status == MINISTRY_INACTIVE):
                        key = 'No longer interested'
                        if key not in output[output_key]:
                            output[output_key][key] = list()
                        output[output_key][key].append(member)

                    elif (pds_status == MINISTRY_INACTIVE and
                          jotform_status == MINISTRY_PARTICIPATE):
                        key = 'Needs human: PDS=inactive, but Jotform=active'
                        if key not in output[output_key]:
                            output[output_key][key] = list()
                        output[output_key][key].append(member)

                    elif (pds_status == MINISTRY_PARTICIPATE and
                          jotform_status == MINISTRY_INTERESTED):
                        key = 'Needs human: PDS=active, but Jotform=not interested'
                        if key not in output[output_key]:
                            output[output_key][key] = list()
                        output[output_key][key].append(member)

    return output

#-----------------------------------------------------------------------------

def member_ministry_csv_report(args, google, start, end, time_period, pds_members, pds_families, jotform_csv, log):
    def _find_all_phones(member):
        found = list()
        key   = 'phones'
        key2  = 'unlisted'
        if key in member:
            # Is this the old member format?  Seems to be a dict of
            # phone_id / phone_data, and no "Unlisted".  :-(
            for p in member[key]:
                # Skip emergency contacts
                if 'Emergency' in p['type']:
                    continue

                text = f"{p['number']} {p['type']}"
                if key2 in p:
                    if p[key2]:
                        text += ' UNLISTED'
                found.append(text)

        # When used with XLSX word wrapping alignment, this will
        # across put each phone number on a separate line, but all
        # within a single cell.
        return '\r\n'.join(found)

    #--------------------------------------------------------------------

    def _find_family_home_phone(member):
        family = member['family']
        key = 'phones'
        key2 = 'unlisted'
        if key in family:
            for p in family[key]:
                if 'Home' in p['type']:
                    text = f"{p['number']} {p['type']}"
                    if key2 in p:
                        if p[key2]:
                            text += ' UNLISTED'
                    return text

        return ""

    #--------------------------------------------------------------------

    output      = analyze_member_ministry_submissions(pds_members, pds_families, jotform_csv, log)
    today       = date.today()

    #--------------------------------------------------------------------

    def _setup_new_workbook():
        workbook    = Workbook()
        sheet       = workbook.active

        for data in xlsx_cols.values():
            col            = data['column']
            cell           = sheet.cell(row=1, column=col, value=data['name'])
            cell.fill      = title_fill
            cell.font      = title_font
            cell.alignment = title_align

            col_char = chr(ord('A') - 1 + col)
            sheet.column_dimensions[col_char].width = data['width']

        sheet.freeze_panes = sheet['A2']

        return workbook

    #--------------------------------------------------------------------

    def _fill(col_name, value, align=None, format=None):
        col_data = xlsx_cols[col_name]
        cell     = sheet.cell(row=xlsx_row, column=col_data['column'], value=value)
        if align:
            cell.alignment = align
        if format:
            cell.number_format = format

    #--------------------------------------------------------------------

    title_font  = Font(color='FFFF00')
    title_fill  = PatternFill(fgColor='0000FF', fill_type='solid')
    title_align = Alignment(horizontal='center', wrap_text=True)

    wrap_align  = Alignment(horizontal='general', wrap_text=True)
    right_align = Alignment(horizontal='right')

    xlsx_cols = dict();
    def _add_col(name, width=10):
        col             = len(xlsx_cols) + 1
        xlsx_cols[name] = {'name' : name, 'column' : col, 'width' : width }
    _add_col('Full Name', width=20)
    _add_col('First')
    _add_col('Last')
    _add_col('Age')
    _add_col('Envelope ID')
    _add_col('Email', width=30)
    _add_col('Member phones', width=20)
    _add_col('Family home phone', width=20)
    _add_col('Category', width=25)
    _add_col('Current ministry status', width=50)
    _add_col('MID')

    for ministry_name in sorted(output.keys()):
        workbook = _setup_new_workbook()
        sheet    = workbook.active

        data = output[ministry_name]
        xlsx_row = 2
        for category in sorted(data.keys()):

            for member in data[category]:
                family = member['family']

                _fill('Full Name', member['email_name'])
                _fill('First', member['first'])
                _fill('Last', member['last'])
                if member['date_of_birth']:
                    age = today - member['date_of_birth']
                    _fill('Age', int(age.days / 365))
                _fill('Envelope ID', family['ParKey'])
                emails = PDSChurch.find_any_email(member)
                if emails:
                    _fill('Email', emails[0])
                _fill('Member phones', _find_all_phones(member), align=wrap_align)
                _fill('Family home phone', _find_family_home_phone(member))
                _fill('Category', category.capitalize(), align=wrap_align)
                _fill('Current ministry status', member['jotform'][ministry_name], align=wrap_align)
                _fill('MID', member['MemRecNum'])

                xlsx_row += 1

        # Write out the XLSX with the results
        filename = f'{ministry_name} jotform results.xlsx'.replace('/', '-')
        if os.path.exists(filename):
            os.unlink(filename)
        workbook.save(filename)
        log.info(f"Wrote to filename: {filename}")

##############################################################################

def family_status_csv_report(args, google, pds_families, jotform, log):
    # Simple report: FID, Family name, and constants.already_submitted_fam_status

    # Did we find anything?
    if len(jotform) == 0:
        log.info("No submissions -- no statuses to update")
        return

    # Make a dictionary of the final CSV data
    csv_data = list()
    for row in jotform:
        fid = int(row['fid'])
        if fid not in pds_families:
            continue

        last_name = pds_families[fid]['Name'].split(',')[0]
        csv_data.append({
            'fid' : fid,
            'Family Name' : pds_families[fid]['Name'],
            'Envelope ID' : helpers.pkey_url(pds_families[fid]['ParKey']),
            'Last Name'   : last_name,
            'Status'      : already_submitted_fam_status,
        })

    filename = ('Family Status Update.csv')
    #ef upload_csv_to_gsheet(google, google_folder_id, filename, fieldnames, csv_rows, remove_local, log):

    gsheet_id, csv_filename = upload_csv_to_gsheet(google,
                                google_folder_id=upload_team_drive_folder_id,
                                filename=filename,
                                fieldnames=csv_data[0].keys(),
                                csv_rows=csv_data,
                                remove_local=False,
                                log=log)
    url = f'https://docs.google.com/spreadsheets/d/{gsheet_id}'

    #------------------------------------------------------------------------

    body = list()
    body.append(f"""<html>
<body>
<h2>Family Status data update</h2>

<p> See attached spreadsheet of FIDs that have submitted anything at all in this time period.
The same spreadsheet <a href="{url}">is also available as a Google Sheet</a>.</p>

<p> Total of {len(csv_data)} families.</p>
</body>
</html>""")

    to = fid_participation_email_to
    subject = f'{title} Family Status updates'
    try:
        log.info(f'Sending "{subject}" email to {to}')
        with smtplib.SMTP_SSL(host=smtp_server,
                              local_hostname='epiphanycatholicchurch.org') as smtp:
            msg = EmailMessage()
            msg['Subject'] = subject
            msg['From'] = smtp_from
            msg['To'] = to
            msg.set_content('\n'.join(body))
            msg.replace_header('Content-Type', 'text/html')

            # This assumes that the file has a single line in the format of username:password.
            with open(args.smtp_auth_file) as f:
                line = f.read()
                smtp_username, smtp_password = line.split(':')

            # Login; we can't rely on being IP whitelisted.
            try:
                smtp.login(smtp_username, smtp_password)
            except Exception as e:
                log.error(f'Error: failed to SMTP login: {e}')
                exit(1)

            # If there were results, attach CSV files
            with open(csv_filename, 'rb') as f:
                csv_data = f.read()
            msg.add_attachment(csv_data, maintype='text', subtype='csv',
                            filename=csv_filename)

            smtp.send_message(msg)

            os.unlink(csv_filename)
    except:
        print(f"==== Error with {to}")
        print(traceback.format_exc())

##############################################################################

def _export_gsheet_to_csv(service, start, end, google_sheet_id, fieldnames):
    response = service.files().export(fileId=google_sheet_id,
                                      mimeType=Google.mime_types['csv']).execute()

    # Some of the field names will be lists.  In those cases, use the first field name in the list.
    final_fieldnames = list()
    final_fieldnames.extend(fieldnames['prelude'])
    for member in fieldnames['members']:
        final_fieldnames.extend(member)
    final_fieldnames.extend(fieldnames['family'])
    final_fieldnames.extend(fieldnames['epilog'])

    csvreader = csv.DictReader(response.decode('utf-8').splitlines(),
                               fieldnames=final_fieldnames)

    rows = list()
    for row in csvreader:
        # Skip title row
        if 'Submission' in row['SubmitDate']:
            continue

        # Is this submission between start and end?
        submit_date = helpers.jotform_date_to_datetime(row['SubmitDate'])
        if submit_date < start or submit_date > end:
            continue

        rows.append(row)

    return rows

#-----------------------------------------------------------------------------

def read_jotform_gsheet(google, start, end, gfile_id):
    csv_data = _export_gsheet_to_csv(google, start, end, gfile_id,
                                      jotform_gsheet_columns)

    # Deduplicate: save the last row number for any given FID
    # (we only really care about the *last* entry that someone makes)
    deduplicated = dict()
    for row in csv_data:
        fid = row['fid']

        # Skip the title row
        if fid == 'fid':
            continue

        deduplicated[fid] = row

    # Turn this dictionary into a list of rows
    out = list()
    for fid in sorted(deduplicated):
        out.append(deduplicated[fid])

    return out

##############################################################################

def setup_args():
    tools.argparser.add_argument('--gdrive-folder-id',
                                 help='If specified, upload a Google Sheet containing the results to this Team Drive folder')

    tools.argparser.add_argument('--all',
                                 action='store_const',
                                 const=True,
                                 help='If specified, run the comparison for all time (vs. running for the previous time period')

    tools.argparser.add_argument('--smtp-auth-file',
                                 required=True,
                                 help='File containing SMTP AUTH username:password')

    global gapp_id
    tools.argparser.add_argument('--app-id',
                                 default=gapp_id,
                                 help='Filename containing Google application credentials')
    global guser_cred_file
    tools.argparser.add_argument('--user-credentials',
                                 default=guser_cred_file,
                                 help='Filename containing Google user credentials')

    args = tools.argparser.parse_args()

    return args

##############################################################################

def main():
    global families, members

    args = setup_args()
    log = ECC.setup_logging(debug=False)

    #---------------------------------------------------------------

    # Calculate the start and end of when we are analyzing in the
    # source data
    epoch = datetime(year=1971, month=1, day=1)
    end = datetime.now()
    if args.all:
        start = epoch
    else:
        # If not supplied on the command line:
        # Sun: skip
        # Mon: everything from last 3 days (Sat+Sun)
        # Tue-Fri: everything from yesterday
        # Sat: skip
        today = end.strftime('%a')
        if today == 'Sat' or today == 'Sun':
            print("It's the weekend.  Nothing to do!")
            exit(0)
        elif today == 'Mon':
            start = end - timedelta(days=3)
        else:
            start = end - timedelta(days=1)

    # No one wants to see the microseconds
    start = start - timedelta(microseconds=start.microsecond)
    end   = end   - timedelta(microseconds=end.microsecond)

    if args.all:
        time_period = 'all results to date'
    else:
        time_period = '{start} - {end}'.format(start=start, end=end)

    log.info("Comments for: {tp}".format(tp=time_period))

    #---------------------------------------------------------------

    log.info("Reading PDS data...")
    (pds, pds_families,
     pds_members) = PDSChurch.load_families_and_members(filename='pdschurch.sqlite3',
                                                        parishioners_only=True,
                                                        log=log)

    #---------------------------------------------------------------

    apis = {
        'drive' : { 'scope'       : Google.scopes['drive'],
                    'api_name'    : 'drive',
                    'api_version' : 'v3', },
    }
    services = GoogleAuth.service_oauth_login(apis,
                                              app_json=args.app_id,
                                              user_json=args.user_credentials,
                                              log=log)
    google = services['drive']

    #---------------------------------------------------------------

    # Load all the results
    log.info("Downloading Jotform raw data...")
    jotform_all = read_jotform_gsheet(google, epoch, end, jotform_gsheet_gfile_id)

    # Load a range of results
    if start == epoch:
        jotform_range = jotform_all.copy()
    else:
        jotform_range = read_jotform_gsheet(google, start, end, jotform_gsheet_gfile_id)

    #---------------------------------------------------------------

    # Do the individual reports

    # These two reports were run via cron at 12:07am on Mon-Fri
    # mornings.
    #comments_report(args, google, start, end, time_period,
    #                jotform_range, log)
    statistics_report(args, end, pds_members, pds_families,
                      jotform_all, log)

    # These reports were uncommented and run by hand upon demand.
    #family_pledge_csv_report(args, google, pds_families, jotform_all, log)
    #family_status_csv_report(args, google, pds_families, jotform_all, log)
    #member_ministry_csv_report(args, google, start, end, time_period,
    #                           pds_members, pds_families, jotform_range, log)

main()
