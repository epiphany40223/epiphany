#!/usr/bin/env python3

# This script is really just for debugging / reference.  It didn't
# play a part in the sending of emails, etc.  It was edited and run on
# demand just as a help for writing / debugging the other scripts.

import sys
sys.path.insert(0, '../../python')

import ECC
import PDSChurch

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment

import helpers

from pprint import pprint
from pprint import pformat

families_per_spreadsheet = 20

##############################################################################

def write_spreadsheet(sorted_families, prefix=None, log=None):
    first     = sorted_families[0]['last_name']
    last      = sorted_families[-1]['last_name']
    filename  = f"ECC Covid wellness calls: {first} through {last}.xlsx"
    if prefix:
        filename = f"{prefix}-{filename}"

    wb = Workbook()
    ws = wb.active

    # Title rows + set column widths
    title_font   = Font(color='FFFF00')
    title_fill   = PatternFill(fgColor='0000FF', fill_type='solid')
    center_align = Alignment(horizontal='center', wrap_text=True)
    left_align   = Alignment(horizontal='left', wrap_text=True)

    def _title(title, width, align=center_align):
        return {
            'name'  : title,
            'width' : width,
            'align' : align,
        }

    # Title row
    titles = [
        _title('Household name', 35),
        _title('Phone number', 15),
        _title('Phone type', 12),

        _title('Head of household', 30),
        _title('Phone number', 15),
        _title('Phone type', 12),

        _title('Spouse', 30),
        _title('Phone number', 15),
        _title('Phone type', 12),

        _title('Number in household', 10),

        _title('Date called', 12),
        _title('Result\n(left message, spoke to someone)', 20),
        _title('Help Requested\n(none, priest, errands, financial, friendly check in, other)', 30),
        _title('Notes', 100, left_align),
    ]

    row = 1
    for index, entry in enumerate(titles):
        column             = chr(ord('A') + index)
        cell               = f'{column}{row}'
        ws[cell]           = entry['name']
        ws[cell].fill      = title_fill
        ws[cell].font      = title_font
        ws[cell].alignment = entry['align']
        ws.column_dimensions[column].width = entry['width']

    # Freeze title row
    row = row + 1
    cell = f'A{row}'
    ws.freeze_panes = ws[cell]

    # Helper to add phone number/type
    def _add_phones(start_column, start_row, entity):
        max_row = start_row
        key     = 'phones'
        if key in entity:
            for index, ph in enumerate(entity[key]):
                ph_row   = start_row + index

                # Phone number
                this_col = chr(ord(start_column) + 0)
                cell     = f'{this_col}{ph_row}'
                ws[cell] = ph['number']

                # Phone type
                this_col = chr(ord(start_column) + 1)
                cell     = f'{this_col}{ph_row}'
                ws[cell] = ph['type']

                max_row = ph_row

        return max_row

    # Write out each family member
    for family_data in sorted_families:
        pds_family = family_data['pds_family']

        # Household name
        cell     = f'A{row}'
        ws[cell] = family_data['family_summary_name']

        # Household phones
        max_row = _add_phones('B', row, pds_family)

        # Head of household
        for index, key in enumerate(['Head of Household', 'Spouse']):
            # Each entry is 3 columns wide (name, phone number, phone type)
            column = chr(ord('D') + index * 3)

            if key in family_data:
                pds_member = family_data[key]
                cell       = f'{column}{row}'
                ws[cell]   = pds_member['full_name']
                lowest_row = _add_phones(chr(ord(column) + 1), row, pds_member)
                max_row    = max(lowest_row, max_row)

        # Total number of household members
        column   = 'J'
        cell     = f'{column}{row}'
        ws[cell] = len(pds_family['members'])
        ws[cell].alignment = center_align

        # Leave a blank row before the next Family
        row    = max_row + 2

    # Write out the filename
    wb.save(filename)
    log.info(f"Wrote {filename}")

#-----------------------------------------------------------------------------

def write_spreadsheets(families, prefix=None, log=None):
    # Important to go in sorted order, so we can't use simple foo[0:19]
    # kind of syntax (because the keys/family names are not necessarily
    # in order in the dictionary keys).
    subset = list()
    for id in sorted(families):
        f = families[id]
        subset.append(f)
        if len(subset) >= families_per_spreadsheet:
            write_spreadsheet(subset, prefix=prefix, log=log)
            subset = list()

    if len(subset) > 0:
        write_spreadsheet(subset, prefix=prefix, log=log)

##############################################################################

def process_families(families, log):
    target_families = dict()
    target_families_all = dict()

    fids = sorted(families)
    for fid in fids:
        f = families[fid]

        family_summary_name = f['Name']
        parts               = f['Name'].split(',')
        family_last_name    = parts[0]
        log.info(f"=== Family: {family_last_name}")

        this_family = {
            'last_name'           : family_last_name,
            'family_summary_name' : family_summary_name,
            'pds_family'          : f,
            'fid'                 : fid,
        }

        for m in f['members']:
            if helpers.member_is_hoh_or_spouse(m):
                this_family[m['type']] = m

        target_families[family_last_name] = this_family
        # Make sure to make the key unique (but still sortable!) by appending the FID
        target_families_all[f"{family_last_name} {fid}"] = this_family

    #pprint(target_families)
    return target_families, target_families_all

# Household name
# Home phone (if exists)
# Adults (name, cell phone number)
# Number of additional household members
#
# Date called
# Result: left message, spoke to someone
# Request (priest, groceries, financial help, friendly call, other)
# Other Notes
#
# Name the spreadsheet something like "Calls: <first name> through <last name>"20 Families per sheet

##############################################################################

def main():
    log = ECC.setup_logging(debug=False)

    (pds, families,
     members) = PDSChurch.load_families_and_members(filename='pdschurch.sqlite3',
                                                    log=log)

    # Remove non-parishioner families
    families = helpers.filter_parishioner_families_only(families, log=log)

    print(f"Total number of parishioner families: {len(families)}")

    # Shoot.  We initially stored by last name only, which effectively removed
    # all families with duplicate last names.
    target_families, target_families_all = process_families(families, log)
    # Find all the families that were accidentally excluded from the initial list
    num_deleted = 0
    for this_family in target_families.values():
        fid = this_family['fid']
        for key, value in target_families_all.items():
            if fid == value['fid']:
                del target_families_all[key]
                num_deleted += 1
                break

    print(f"Number of families deleted from dup list: {num_deleted}")
    print(f"Number of target families in initial list: {len(target_families)}")
    print(f"Number of families excluded from initial list: {len(target_families_all)}")

    write_spreadsheets(target_families, log=log)
    write_spreadsheets(target_families_all, prefix="missed", log=log)

main()
