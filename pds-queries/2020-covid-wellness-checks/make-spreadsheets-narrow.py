#!/usr/bin/env python3

# This script is really just for debugging / reference.  It didn't
# play a part in the sending of emails, etc.  It was edited and run on
# demand just as a help for writing / debugging the other scripts.

import sys
sys.path.insert(0, '../../python')

import ECC
import PDSChurch

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

import helpers

from pprint import pprint
from pprint import pformat

##############################################################################

def write_spreadsheet(sorted_families, prefix=None, log=None):
    first     = sorted_families[0]['last_name']
    last      = sorted_families[-1]['last_name']
    filename  = f"ECC Covid wellness calls: {first} through {last}.xlsx"
    if prefix:
        filename = f"{prefix}-{filename}"

    #---------------------------------------------------------------------

    wb = Workbook()
    ws = wb.active

    # Set narrow margins so that we can get as much in as possible if people print it out
    ws.page_margins.left   = 0.25
    ws.page_margins.right  = ws.page_margins.left
    ws.page_margins.top    = 0.4
    ws.page_margins.bottom = ws.page_margins.top

    ws.page_setup.orientation = ws.ORIENTATION_LANDSCAPE

    thin_side  = Side(style='thin')
    box_border = Border(top=thin_side, bottom=thin_side, left=thin_side, right=thin_side)

    #---------------------------------------------------------------------

    title_font   = Font(color='000000', bold=True)
    title_fill   = PatternFill(fgColor='EFEFEF', fill_type='solid')
    center_align = Alignment(horizontal='center', wrap_text=True)
    left_align   = Alignment(horizontal='left', wrap_text=True)

    def _make_title(cell, value, width):
        ws[cell]           = value
        ws[cell].font      = title_font
        ws[cell].fill      = title_fill
        ws[cell].alignment = center_align
        ws[cell].border    = box_border
        ws.column_dimensions[cell[0]].width = width

    _make_title('A1', 'Household names', 25)
    _make_title('B1', 'Phone numbers', 22)
    _make_title('C1', 'Dates messages left', 10)
    _make_title('D1', 'Date reached', 10)
    _make_title('E1', 'Help Requested\n(none, priest, errands, financial, other)', 20)
    _make_title('F1', 'Notes', 47)
    _make_title('G1', 'Call again?', 10)

    #---------------------------------------------------------------------

    # Freeze title row
    ws.freeze_panes = ws['A2']

    #---------------------------------------------------------------------

    def _find_phones(entity, phones):
        key = 'phones'
        if key not in entity:
            return

        for phone in entity[key]:
            if 'emergency' not in phone['type'].lower():
                string = phone['number']
                type   = phone['type']
                if len(type) > 0:
                    string += f' ({type})'
                phones[phone['number']] = string

    def _set(cell, value):
        ws[cell]           = value
        ws[cell].font      = value_font
        ws[cell].alignment = value_align
        ws[cell].border    = box_border

    # Starting row
    row = 1

    value_font  = Font(name='Arial', size=10)
    value_align = Alignment(wrap_text=True)

    # Write out each family member
    for family_data in sorted_families:
        pds_family = family_data['pds_family']

        # Household name
        row += 1
        _set(f'A{row}', f"Household: {family_data['last_name']}")
        num = len(pds_family['members'])
        s   = 's' if num > 1 else ''
        _set(f'B{row}', f'{num} household member{s}')

        for type in ['Head of Household', 'Spouse']:
            if type in family_data:
                row += 1
                pds_member = family_data[type]

                _set(f'A{row}', pds_member['full_name'])

                phones = dict()
                _find_phones(pds_member, phones)
                _find_phones(pds_family, phones)
                string = ''
                for type in phones:
                    if len(string) > 0:
                        string += '\n'
                    string += phones[type]
                if len(string) == 0:
                    string = 'None on record'
                _set(f'B{row}', string)

                # Set blank values for the rest of the columns so that we get the box borders
                for col in ['C', 'D', 'E', 'F', 'G']:
                    _set(f'{col}{row}', '')

                # Set the row height
                num_phones = len(phones)
                num_phones = 1 if num_phones == 0 else num_phones
                actual_height = 15 * num_phones
                if actual_height < 30:
                    ws.row_dimensions[row].height = 30

        # Leave a blank row before the next Family
        row += 1

    # Write out the filename
    wb.save(filename)
    log.info(f"Wrote {filename}")

#-----------------------------------------------------------------------------

def write_spreadsheets(families, num_sheets, prefix=None, log=None):
    def _get_fps():
        return base_fps if remainder <= 0 else (base_fps+1)

    # Base number of families per spreadsheet
    base_fps = int(len(families) / num_sheets)
    # Unless that divided evenly, some spreadsheets will need to have one more family
    remainder = len(families) - (base_fps * num_sheets)

    # Important to go in sorted order, so we can't use simple foo[0:19]
    # kind of syntax (because the keys/family names are not necessarily
    # in order in the dictionary keys).
    subset = list()
    fps = _get_fps()
    for id in sorted(families):
        f = families[id]
        subset.append(f)
        if len(subset) >= fps:
            write_spreadsheet(subset, prefix=prefix, log=log)

            subset = list()
            remainder -= 1
            fps = _get_fps()

    if len(subset) > 0:
        write_spreadsheet(subset, prefix=prefix, log=log)

##############################################################################

def process_families(families, log):
    spanish_families = dict()
    english_families = dict()

    for fid in sorted(families):
        f = families[fid]

        family_summary_name = f['Name']
        parts               = f['Name'].split(',')
        family_last_name    = parts[0]
        log.info(f"=== Family: {family_last_name}")

        this_family = {
            'last_name'           : family_last_name,
            'family_summary_name' : family_summary_name,
            'pds_family'          : f,
            'fid'                 : fid,
        }

        languages = dict()
        for m in f['members']:
            if helpers.member_is_hoh_or_spouse(m):
                this_family[m['type']] = m

                key = 'language'
                if key in m:
                    primary = m[key].split('/')[0].lower()
                    languages[primary] = True

        if len(languages) == 1 and 'spanish' in languages:
            spanish_families[f"{family_last_name} {fid}"] = this_family
        else:
            english_families[f"{family_last_name} {fid}"] = this_family

    return spanish_families, english_families

##############################################################################

def main():
    log = ECC.setup_logging(debug=False)

    (pds, families,
     members) = PDSChurch.load_families_and_members(filename='pdschurch.sqlite3',
                                                    log=log)

    # Remove non-parishioner families
    families = helpers.filter_parishioner_families_only(families, log=log)

    print(f"Total number of parishioner families: {len(families)}")

    spanish_families, english_families = process_families(families, log)

    print(f"Number of spanish families in list: {len(spanish_families)}")
    print(f"Number of english families in list: {len(english_families)}")

    # We have 70 english volunteers and 1 spanish volunteer
    write_spreadsheets(families=english_families, num_sheets=70, prefix='English', log=log)
    write_spreadsheets(families=spanish_families, num_sheets=1, prefix='Spanish', log=log)

main()
