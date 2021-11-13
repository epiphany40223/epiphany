#!/usr/bin/env python3

# Make sure to pip install everything in requirements.txt.

import sys

import os
import datetime
import argparse

# We assume that there is a "ecc-python-modules" sym link in this
# directory that points to the directory with ECC.py and friends.
moddir = os.path.join(os.getcwd(), 'ecc-python-modules')
if not os.path.exists(moddir):
    print("ERROR: Could not find the ecc-python-modules directory.")
    print("ERROR: Please make a ecc-python-modules sym link and run again.")
    exit(1)

sys.path.insert(0, moddir)

import ECC
import Google
import GoogleAuth

import helpers

from datetime import date
from datetime import datetime
from datetime import timedelta

from oauth2client import tools
from googleapiclient.http import MediaFileUpload

import googleapiclient
from google.api_core import retry

import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment

from copy import copy

##############################################################################

gapp_id         = 'client_id.json'
guser_cred_file = 'user-credentials.json'

source_google_shared_drive = '0AHQEqKijqXcFUk9PVA'
source_google_drive_folder = '15gISGMwGS8F3qziEaWPeT5YhW6oHKH6y'

##############################################################################

def _upload_to_gsheet(google, google_folder_id, google_filename, mime_type, local_filename, remove_local, log):
    @retry.Retry(predicate=Google.retry_errors)
    def _upload():
        log.info(f'Uploading file to google "{local_filename}" -> "{google_filename}"')
        metadata = {
            'name'     : google_filename,
            'mimeType' : Google.mime_types['sheet'],
            'parents'  : [ google_folder_id ],
            'supportsTeamDrives' : True,
            }
        media = MediaFileUpload(local_filename,
                                mimetype=Google.mime_types[mime_type],
                                resumable=True)
        file = google.files().create(body=metadata,
                                     media_body=media,
                                     supportsTeamDrives=True,
                                     fields='id').execute()
        log.debug(f'Successfully uploaded file: "{google_filename}" (ID: {file["id"]})')

        return file

    # Set permissions on the GSheet to allow the
    # workers group to edit the file (if you are view-only, you
    # can't even adjust the column widths, which will be
    # problematic for the comments report!).
    @retry.Retry(predicate=Google.retry_errors)
    def _set_perms():
        log.debug(f"Setting Google permission for file: {id}")
        perm = {
            'type': 'group',
            'role': 'writer',
            'emailAddress': gsheet_editors,
        }
        out = google.permissions().create(fileId=file['id'],
                                          supportsTeamDrives=True,
                                          sendNotificationEmail=False,
                                          body=perm,
                                          fields='id').execute()

    file = _upload()
    out = _set_perms(file)

    # Remove the temp file when we're done
    if remove_local:
        try:
            os.remove(local_filename)
        except:
            pass

    return file['id']

##############################################################################

def _download_google_sheet(google, gfile, log):
    @retry.Retry(predicate=Google.retry_errors)
    def _download():
        log.info(f"Downloading Sheet {gfile['name']}...")
        return google.files().export(fileId=gfile['id'],
                                     mimeType=Google.mime_types['xlsx']).execute()

    xlsx_content = _download()

    # Write the downloaded XLSX content into a file and then use
    # Openpyxl to load it.
    tmpfile = 'foo.xlsx'
    if os.path.exists(tmpfile):
        os.unlink(tmpfile)
    with open(tmpfile, "wb") as f:
        f.write(xlsx_content)

    workbook = openpyxl.load_workbook(tmpfile)
    os.unlink(tmpfile)

    return workbook

#-----------------------------------------------------------------------------

def download_google_sheets(google, args, log):
    @retry.Retry(predicate=Google.retry_errors)
    def _get_file_list():
        log.info(f"Finding Google Sheets in Team Drive {args.shared_drive_id}, folder {args.gdrive_folder_id}")
        mime = Google.mime_types['sheet']
        q = f'"{args.gdrive_folder_id}" in parents and mimeType="{mime}" and trashed=false'
        response = google.files().list(corpora='drive',
                                       driveId=args.shared_drive_id,
                                       includeTeamDriveItems=True,
                                       supportsAllDrives=True,
                                       q=q,
                                       fields='files(name,id)').execute()
        return response

    response = _get_file_list()
    gfiles = response.get('files', [])

    # Make a sorted list of files
    sortable_gfiles = dict()
    for gfile in gfiles:
        sortable_gfiles[gfile['name']] = gfile

    workbooks = list()
    for name in sorted(sortable_gfiles):
        gfile = sortable_gfiles[name]
        workbooks.append(_download_google_sheet(google, gfile, log))

    return workbooks

##############################################################################

def squish(workbooks, log):
    def _copy_row(src_row, dest_sheet, row_number):
        for cell in src_row:
            val = cell.value
            if val == '':
                continue

            # A few common mistakes people make
            if val == 'inactive':
                cell.value = 'Inactive'
            elif type(val) is str and val.lower() == 'active':
                cell.value = 'Actively Involved'
            elif val == 'interested':
                cell.value = 'Interested'
            elif type(val) is str and val.lower() == 'no longer interested':
                cell.value = 'No longer interested'

            log.debug(f"Copying cell: row={row_number}, col={cell.col_idx}, value={cell.value}")
            new_cell = dest_sheet.cell(row=row_number,
                                       column=cell.col_idx,
                                       value=cell.value)

            if cell.has_style:
                new_cell.font = copy(cell.font)
                #new_cell.border = copy(cell.border)
                new_cell.fill = copy(cell.fill)
                new_cell.number_format = copy(cell.number_format)
                new_cell.protection = copy(cell.protection)
                new_cell.alignment = copy(cell.alignment)

    #-----------------------------------------------

    output_wb = Workbook()
    output_ws = output_wb.active

    first = True
    row_number = 1

    for src_workbook in workbooks:
        src_worksheet = src_workbook.active
        for src_row in src_worksheet.rows:
            if src_row[0].value == 'Full Name':
                # First time through, we copy the title row and column
                # widths
                if first:
                    first = False
                    _copy_row(src_row, output_ws, row_number)

                    widths = list()
                    for col, width in src_worksheet.column_dimensions.items():
                        output_ws.column_dimensions[col] = width

                    row_number += 1
                continue

            # Otherwise, just copy the row
            _copy_row(src_row, output_ws, row_number)
            row_number += 1

    return output_wb

##############################################################################

def upload_workbook(google, workbook, args, log):
    # Sometimes the output that it writes is corrupted -- i.e., Excel
    # complains about bad data in it when you open the XLSX.  ...but
    # the file seems to load just fine.  So if you save it again to
    # get rid of whatever was bad and then upload that to Google Drive
    # -- good to go!
    filename = 'worksheet-rollup.xlsx'
    workbook.save(filename)
    log.info(f"Wrote {filename} -- did NOT upload to Google!")

##############################################################################

def setup_args():
    tools.argparser.add_argument('--shared-drive-id',
                                 default=source_google_shared_drive,
                                 help='Google Shared Drive ID')

    tools.argparser.add_argument('--gdrive-folder-id',
                                 default=source_google_drive_folder,
                                 help='Google Drive folder from which to download/load the Sheets (must be in the indicated shared drive)')

    global gapp_id
    tools.argparser.add_argument('--app-id',
                                 default=gapp_id,
                                 help='Filename containing Google application credentials')
    global guser_cred_file
    tools.argparser.add_argument('--user-credentials',
                                 default=guser_cred_file,
                                 help='Filename containing Google user credentials')

    args = tools.argparser.parse_args()

    return args

##############################################################################

def main():
    args = setup_args()
    log = ECC.setup_logging(debug=False)

    #---------------------------------------------------------------

    apis = {
        'drive' : { 'scope'       : Google.scopes['drive'],
                    'api_name'    : 'drive',
                    'api_version' : 'v3', },
    }
    services = GoogleAuth.service_oauth_login(apis,
                                              app_json=args.app_id,
                                              user_json=args.user_credentials,
                                              log=log)
    google = services['drive']

    #---------------------------------------------------------------

    # Load all the sheets
    workbooks = download_google_sheets(google, args, log)
    workbook = squish(workbooks, log)
    upload_workbook(google, workbook, args, log)

main()
