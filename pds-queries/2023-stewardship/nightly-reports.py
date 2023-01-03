#!/usr/bin/env python3

# Make sure to pip install everything in requirements.txt.

import sys

import collections
import traceback
import datetime
import argparse
import smtplib
import csv
import os
import re

# We assume that there is a "ecc-python-modules" sym link in this
# directory that points to the directory with ECC.py and friends.
moddir = os.path.join(os.getcwd(), 'ecc-python-modules')
if not os.path.exists(moddir):
    print("ERROR: Could not find the ecc-python-modules directory.")
    print("ERROR: Please make a ecc-python-modules sym link and run again.")
    exit(1)

sys.path.insert(0, moddir)

import ECC
import Google
import PDSChurch
import GoogleAuth

import helpers

from datetime import date
from datetime import datetime
from datetime import timedelta

from oauth2client import tools
from googleapiclient.http import MediaFileUpload
from email.message import EmailMessage

import matplotlib
import matplotlib.dates as mdates
import matplotlib.pyplot as plt

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment

#------------------------------------------------------------------------------

from constants import already_submitted_fam_status
from constants import already_submitted_fam_keyword

from constants import stewardship_begin_date
from constants import stewardship_end_date

from constants import gapp_id
from constants import guser_cred_file

from constants import jotform_gsheet_gfile_id
from constants import jotform_gsheet_columns
from constants import upload_team_drive_folder_id
from constants import gsheet_editors

from constants import stewardship_year
from constants import title

from constants import smtp_server
from constants import smtp_from

from constants import jotform

from constants import MAX_PDS_FAMILY_MEMBER_NUM

from constants import jotform_gsheet_prev_year_gfile_id
from constants import jotform_gsheet_prev_year_columns

##############################################################################

ecc = '@epiphanycatholicchurch.org'

# Comments and pledge analysis report email
reports_email_to = f'angie{ecc},mary{ecc},jeff@squyres.com'
reports_email_subject = 'Comments and pledge analysis reports'

# Statistics report email
statistics_email_to = f'sdreiss71@gmail.com,angie{ecc},mary{ecc},jeff@squyres.com'
statistics_email_subject = 'Statistics report'

fid_participation_email_to = f'mary{ecc},jeff@squyres.com'

pledge_email_to = fid_participation_email_to
pledge_email_subject = 'Pledge PDS CSV import file'

# JMS for debugging/testing
#statistics_email_to = 'jeff@squyres.com'
#reports_email_to = statistics_email_to
#fid_participation_email_to = statistics_email_to
#pledge_email_to = statistics_email_to


##############################################################################

def _upload_to_gsheet(google, google_folder_id, google_filename, mime_type, local_filename, remove_local, log):
    try:
        log.info(f'Uploading file to google "{local_filename}" -> "{google_filename}"')
        metadata = {
            'name'     : google_filename,
            'mimeType' : Google.mime_types['sheet'],
            'parents'  : [ google_folder_id ],
            'supportsTeamDrives' : True,
            }
        media = MediaFileUpload(local_filename,
                                mimetype=Google.mime_types[mime_type],
                                resumable=True)
        file = google.files().create(body=metadata,
                                     media_body=media,
                                     supportsTeamDrives=True,
                                     fields='id').execute()
        log.debug(f'Successfully uploaded file: "{google_filename}" (ID: {file["id"]})')

    except:
        log.error('Google upload failed for some reason:')
        log.error(traceback.format_exc())
        exit(1)

    # Set permissions on the GSheet to allow the
    # workers group to edit the file (if you are view-only, you
    # can't even adjust the column widths, which will be
    # problematic for the comments report!).
    # JMS Fix me
    if False:
        try:
            perm = {
                'type': 'group',
                'role': 'writer',
                'emailAddress': gsheet_editors,
            }
            out = google.permissions().create(fileId=file['id'],
                                              supportsTeamDrives=True,
                                              sendNotificationEmail=False,
                                              body=perm,
                                              fields='id').execute()
            log.debug(f"Set Google permission for file: {id}")
        except:
            log.error('Google set permission failed for some reason:')
            log.error(traceback.format_exc())
            exit(1)

    # Remove the temp file when we're done
    if remove_local:
        try:
            os.remove(local_filename)
        except:
            pass

    return file['id']

#-----------------------------------------------------------------------------

def _make_filenames(filename, extension):
    extension = f".{extension}"
    if filename.endswith(extension):
        local_filename  = filename
        google_filename = filename[:-len(extension)]
    else:
        local_filename  = f"{filename}{extension}"
        google_fielname = filename

    return local_filename, google_filename

#-----------------------------------------------------------------------------

def upload_csv_to_gsheet(google, google_folder_id, filename, fieldnames, csv_rows, remove_local, log):
    if csv_rows is None or len(csv_rows) == 0:
        return None, None

    # First, write out a CSV file
    csv_filename, google_filename = _make_filenames(filename, 'csv')
    try:
        os.remove(csv_filename)
    except:
        pass

    csvfile = open(csv_filename, 'w')
    writer  = csv.DictWriter(csvfile, fieldnames=fieldnames,
                            quoting=csv.QUOTE_ALL)
    writer.writeheader()
    for row in csv_rows:
        writer.writerow(row)
    csvfile.close()

    # Now upload that file to Google Drive
    id = _upload_to_gsheet(google,
                        google_folder_id=google_folder_id,
                        google_filename=google_filename,
                        mime_type='csv',
                        local_filename=csv_filename,
                        remove_local=remove_local,
                        log=log)

    return id, None if remove_local else csv_filename

#-----------------------------------------------------------------------------

def upload_xlsx_to_gsheet(google, google_folder_id, filename, workbook, remove_local, log):
    # First, write out the XLSX file
    xlsx_filename, google_filename = _make_filenames(filename, 'xlsx')
    try:
        os.remove(xlsx_filename)
    except:
        pass
    workbook.save(xlsx_filename)

    # Now upload that file to Google Drive
    id = _upload_to_gsheet(google,
                        google_folder_id=google_folder_id,
                        # For some reason, when we upload an XLSX and want
                        # it to convert to a Google Sheet, the google filename
                        # must end in .xlsx.  Sigh.
                        google_filename=xlsx_filename,
                        mime_type='sheet',
                        local_filename=xlsx_filename,
                        remove_local=remove_local,
                        log=log)

    return id, None if remove_local else xlsx_filename

##############################################################################

def _change(label, old_value, new_value, message):
    return {
        'label'     : label,
        'old_value' : old_value,
        'new_value' : new_value,
        'message'   : message,
    }

def _compare(changes, label, jot_value, pds_value):
    if jot_value is None:
        jot_value = ''
    if pds_value is None:
        pds_value = ''

    if jot_value.strip() == pds_value.strip():
        return

    message = ('{label}: {new_value}'
               .format(label=label, new_value=jot_value))
    changes.append(_change(label=label,
                           old_value=pds_value,
                           new_value=jot_value,
                           message=message))

##############################################################################

def comments_to_xlsx(google, jotform_data, id_field, emails_field, name_field, env_field, workbook, log):
    sheet = workbook.active

    comments_label    = "Comments"
    pledge_last_label = f'CY{stewardship_year-1} pledge'
    pledge_cur_label  = f'CY{stewardship_year} pledge'
    amount_label      = f'CY{stewardship_year-1} gifts'

    # Setup the title row
    # Title rows + set column widths
    title_font   = Font(color='FFFF00')
    title_fill   = PatternFill(fgColor='0000FF', fill_type='solid')
    title_align  = Alignment(horizontal='center', wrap_text=True)

    wrap_align   = Alignment(horizontal='general', wrap_text=True)
    right_align  = Alignment(horizontal='right')

    money_format = "$###,###,###"

    xlsx_cols = dict();
    def _add_col(name, width=10):
        col             = len(xlsx_cols) + 1
        xlsx_cols[name] = {'name' : name, 'column' : col, 'width' : width }
    _add_col('Date', width=20)
    _add_col('FID')
    _add_col('Envelope')
    _add_col('Family names', width=30)
    _add_col('Emails', width=50)
    _add_col(pledge_last_label)
    _add_col(pledge_cur_label)
    _add_col(amount_label)
    _add_col('Comments', width=100)

    for data in xlsx_cols.values():
        col            = data['column']
        cell           = sheet.cell(row=1, column=col, value=data['name'])
        cell.fill      = title_fill
        cell.font      = title_font
        cell.alignment = title_align

        col_char = chr(ord('A') - 1 + col)
        sheet.column_dimensions[col_char].width = data['width']

    sheet.freeze_panes = sheet['A2']

    #-------------------------------------------------------------------

    def _extract_money_string(val):
        if val.startswith('$'):
            val = val[1:]
        val = int(float(val.replace(',', '').strip()))

        if val != 0:
            return int(val), money_format
        else:
            return 0, None

    #-------------------------------------------------------------------

    # Now fill in all the data rows
    xlsx_row = 2
    log.info(f"Checking {len(jotform_data)} rows for comments")
    for row in jotform_data:
        # Skip if the comments are empty
        if comments_label not in row:
            continue
        if row[comments_label] == '':
            continue

        if row[pledge_last_label]:
            pledge_last, pledge_last_format = _extract_money_string(row[pledge_last_label])
        else:
            pledge_last = 0
            pledge_last_format = None

        if row[pledge_cur_label]:
            pledge_cur = helpers.jotform_text_to_int(row[pledge_cur_label])
            pledge_cur_format = money_format
        else:
            pledge_cur = 0
            pledge_cur_format = None

        if row[amount_label]:
            amount, amount_format = _extract_money_string(row[amount_label])
        else:
            amount = 0
            amount_format = None

        def _fill(col_name, value, align=None, format=None):
            col_data = xlsx_cols[col_name]
            cell     = sheet.cell(row=xlsx_row, column=col_data['column'], value=value)
            if align:
                cell.alignment = align
            if format:
                cell.number_format = format

        _fill('Date', row['SubmitDate'])
        _fill('FID', int(row[id_field]))
        # Do NOT convert the envelope ID to an int -- the leading zeros
        # are significant.
        _fill('Envelope', row[env_field])
        _fill('Family names', row[name_field])
        _fill('Emails', row[emails_field])
        _fill(pledge_last_label, pledge_last, format=pledge_last_format)
        _fill(pledge_cur_label, pledge_cur, format=pledge_cur_format)
        _fill(amount_label, amount, format=amount_format)
        _fill('Comments', row[comments_label], align=wrap_align)

        xlsx_row += 1

    # Return the number of comments we found
    return xlsx_row - 2

def reorder_rows_by_date(jotform_data):
    data = dict()
    for row in jotform_data:
        data[row['SubmitDate']] = row

    out = list()
    for row in sorted(data):
        out.append(data[row])

    return out

def comments_report(args, google, start, end, time_period, jotform_data, log):
    log.info("Composing comments report...")

    # jotform_data is a list of rows, in order by FID.  Re-order them to be
    # ordered by submission date.
    ordered_data = reorder_rows_by_date(jotform_data)

    # Examine the jotform data and see if there are any comments that
    # need to be reported
    workbook = Workbook()
    num_comments = comments_to_xlsx(google, jotform_data=ordered_data,
                     id_field='fid', name_field='Family names',
                     env_field='EnvId', emails_field='Emails to reply to',
                     workbook=workbook, log=log)

    # If we have any comments, upload them to a Gsheet
    gsheet_id = None
    sheet = workbook.active
    if num_comments > 0:
        filename   = f'Comments {time_period}.xlsx'
        gsheet_id, _ = upload_xlsx_to_gsheet(google,
                                        google_folder_id=upload_team_drive_folder_id,
                                        filename=filename,
                                        workbook=workbook,
                                        remove_local=True,
                                        log=log)

    return gsheet_id

##############################################################################

def statistics_compute(pds_families, unique_fid_jotform, log):
    ret = dict()

    # Total number of active Families in the parish
    ret['num_active'] = len(pds_families)

    #---------------------------------------------------------------

    # Number of active Families who are eligible for electronic stewardship
    # (i.e., we have an email address for the spouse and/or HoH)

    eligible = dict()
    for fid, family in pds_families.items():
        for member in family['members']:
            if PDSChurch.is_member_hoh_or_spouse(member):
                em = PDSChurch.find_any_email(member)
                if len(em) == 0:
                    continue
                eligible[fid] = True
                continue

    ret['num_eligible'] = len(eligible)

    #-----------------------------------------------------------

    # The unique_fid_jotform dictionary we have will have, at most, 1 entry per
    # FID.  So we can just take the length of it to know how many (unique)
    # families have submitted electronically.
    ret['num_electronic'] = len(unique_fid_jotform)

    #-----------------------------------------------------------

    # Build a cross-reference of which families have submitted electronically
    fids_electronic = dict()
    for row in unique_fid_jotform:
        fid = int(row['fid'])
        fids_electronic[fid] = True

    fids_paper_or_electronic = fids_electronic.copy()

    # - Count how many submitted paper
    # - Count how many submitted paper and electronic
    # - Count how many submitted paper or electronic
    fids_paper = dict()
    fids_paper_and_electronic = dict()
    for fid, family in pds_families.items():
        if 'status' in family and family['status'] == already_submitted_fam_status:
            fids_paper[fid] = True
            fids_paper_or_electronic[fid] = True

            if fid in fids_electronic:
                fids_paper_and_electronic[fid] = True

    ret['num_paper']                = len(fids_paper)
    ret['num_paper_or_electronic']  = len(fids_paper_or_electronic)
    ret['num_paper_and_electronic'] = len(fids_paper_and_electronic)

    #-----------------------------------------------------------

    return ret

#------------------------------------------------------------------------

def statistics_graph(pds_members, pds_families, jotform, log):
    def _find_range(data, earliest, latest):
        for row in data:
            submitted = row['SubmitDate']
            submitted_dt = helpers.jotform_date_to_datetime(submitted)
            if submitted_dt < earliest:
                earliest = submitted_dt
            if submitted_dt > latest:
                latest = submitted_dt

        return earliest, latest

    #------------------------------------------------------------------------

    def _compute(start, end, pds_members, pds_families, jotform, log):
        # Compute these values just in the date range:
        # - How many unique family submissions total?

        family_submitted = dict()

        # Check electronic submissions
        for row in jotform:
            if row['fid'] == 'fid':
                continue # Skip title row

            # Is this row in our date range?
            dt = helpers.jotform_date_to_datetime(row['SubmitDate'])
            if dt < start or dt > end:
                continue

            fid = int(row['fid'])
            log.debug(f"Found submission in our date window: {fid} on {dt}")

            # Make sure the family hasn't been deleted
            if fid not in pds_families:
                continue

            family_submitted[fid] = True

        return len(family_submitted)

    #------------------------------------------------------------------------

    one_day  = timedelta(days=1)
    earliest = datetime(year=9999, month=12, day=31)
    latest   = datetime(year=1971, month=1,  day=1)
    earliest, latest = _find_range(jotform, earliest, latest)

    earliest = datetime(year=earliest.year, month=earliest.month, day=earliest.day)
    latest   = datetime(year=latest.year, month=latest.month, day=latest.day)
    latest   += one_day - timedelta(seconds=1)

    log.info(f"Earliest: {earliest}")
    log.info(f"Latest:   {latest}")

    day             = earliest
    dates           = list()
    data_per_day    = list()
    data_cumulative = list()

    # Make lists that we can give to matplotlib for plotting
    while day <= latest:
        log.debug(f"Get per-day stats for {day.date()}")
        per_day    = _compute(day, day + one_day,
                              pds_members, pds_families,
                              jotform, log)
        log.debug(f"Get cumulative stats for {earliest} - {day + one_day}")
        cumulative = _compute(earliest, day + one_day,
                              pds_members, pds_families,
                              jotform, log)

        log.debug(f"Date: {day}: per day {per_day}, cumulative {cumulative}")

        dates.append(day.date())
        data_per_day.append(per_day)
        data_cumulative.append(cumulative)

        day += one_day

    # Make the plot
    fig, ax = plt.subplots()

    # The X label of "days" is obvious, there's no good Y label since we have
    # multiple lines
    n = datetime.now()
    hour = n.hour if n.hour > 0 else 12
    ampm = 'am' if n.hour < 12 else 'pm'
    plt.title("As of {month} {day}, {year} at {hour}:{min:02}{ampm}"
            .format(month=n.strftime("%B"), day=n.day, year=n.year,
                    hour=hour, min=n.minute, ampm=ampm))
    plt.suptitle(title + " electronic submissions statistics")

    ax.plot(dates, data_per_day, label='Unique family electronic submissions per day')
    ax.plot(dates, data_cumulative, label='Cumulative unique family electronic submissions')

    ax.get_xaxis().set_major_locator(mdates.DayLocator(interval=1))
    ax.get_xaxis().set_major_formatter(mdates.DateFormatter("%a %b %d"))
    plt.setp(ax.get_xticklabels(), rotation=20, ha="right",
             rotation_mode="anchor")
    ax.grid()

    plt.legend(loc="upper left")

    # Make a filename based on the time that this was run
    filename = ('statistics-as-of-{year:04}-{month:02}-{day:02}-{hour:02}{min:02}{sec:02}.pdf'
                .format(year=n.year, month=n.month, day=n.day,
                        hour=n.hour, min=n.minute, sec=n.second))
    fig.savefig(filename)

    plt.close(fig)

    return filename

#-----------------------------------------------------------------------------

def statistics_report(args, time_period, pds_members, pds_families, jotform, log):
    log.info("Composing statistics report...")

    #---------------------------------------------------------------

    graph_filename = statistics_graph(pds_members, pds_families, jotform, log)
    data = statistics_compute(pds_families, jotform, log)

    #---------------------------------------------------------------------

    electronic_percentage           = data['num_electronic'] / data['num_eligible'] * 100
    paper_percentage                = data['num_paper'] / data['num_active'] * 100
    paper_and_electronic_percentage = data['num_paper_and_electronic'] / data['num_active'] * 100
    paper_or_electronic_percentage  = data['num_paper_or_electronic'] / data['num_active'] * 100

    # Send the statistics report email
    body = list()
    body.append(f"""<html>
<body>
<h2>{title} statistics update</h2>

<h3>Time period: {time_period}</h3>
<ul>
<li> Total number of active PDS Families in the parish: {data['num_active']:,d}</li>
<br>
<li> Number of active PDS Families eligible for electronic stewardship: {data['num_eligible']:,d}
    <ul>
    <li>This means that we have an email address in PDS for the Head of Household and/or the Spouse of a given Family</li>
    </ul></li>
<br>
<li> Number of electronic-stewardship-eligible Families who have electronically submitted: {data['num_electronic']:,d} (out of {data['num_eligible']:,d}, or {electronic_percentage:.1f}%)
    <ul>
    <li>This is the number of families who submitted electronically.</li>
    </ul></li>
<br>
<li> Number of active PDS Families with "{already_submitted_fam_status}" status: {data['num_paper']:,d} (out of {data['num_active']:,d}, or {paper_percentage:.1f}%)
    <ul>
    <li>This is the number of families who submitted a paper card.</li>
    </ul></li>
<br>
<li> Number of active PDS Families who have electronically completed their submission <em>and</em> have "{already_submitted_fam_status}" status: {data['num_paper_and_electronic']:,d} (out of {data['num_active']:,d}, or {paper_and_electronic_percentage:.1f}%)
    <ul>
    <li>These are people who submitted twice -- paper and electronically!</li>
    </ul></li>
<br>
<li> Number of active PDS Families who have electronically completed their submission <em>or</em> have "{already_submitted_fam_status}" status: {data['num_paper_or_electronic']:,d} (out of {data['num_active']:,d}, or {paper_or_electronic_percentage:.1f}%)
    <ul>
    <li>This is the total number of active Families who have submitted.</li>
    </ul></li>
</ul>
</body>
</html>""")

    to = statistics_email_to
    subject = f'{statistics_email_subject} ({time_period})'
    try:
        log.info(f'Sending "{subject}" email to {to}')
        with smtplib.SMTP_SSL(host=smtp_server,
                              local_hostname='epiphanycatholicchurch.org') as smtp:
            msg = EmailMessage()
            msg['Subject'] = subject
            msg['From'] = smtp_from
            msg['To'] = to
            msg.set_content('\n'.join(body))
            msg.replace_header('Content-Type', 'text/html')

            # This assumes that the file has a single line in the format of username:password.
            with open(args.smtp_auth_file) as f:
                line = f.read()
                smtp_username, smtp_password = line.split(':')

            # Login; we can't rely on being IP whitelisted.
            try:
                smtp.login(smtp_username, smtp_password)
            except Exception as e:
                log.error(f'Error: failed to SMTP login: {e}')
                exit(1)

            with open(graph_filename, "rb") as f:
                csv_data = f.read()
            msg.add_attachment(csv_data, maintype='application', subtype='pdf',
                            filename=graph_filename)
            os.unlink(graph_filename)

            smtp.send_message(msg)
    except:
        print("==== Error with {email}".format(email=to))
        print(traceback.format_exc())

##############################################################################

def pledge_comparison_report(google, jotform_this_year, jotform_last_year, log):
    # Extract just 2 fields from the jotform data and return it in a
    # dict indexed by fid.
    def _simplify_jotform(jotform_data, participation_fieldname,
                          this_year_pledge_fieldname,
                          last_year_pledge_fieldname, log):
        out = dict()
        for fid, data in jotform_data.items():
            participate = True
            if participation_fieldname and data[participation_fieldname].startswith("Because"):
                participate = False

            current_pledge = 0
            if participate:
                current_pledge = helpers.jotform_text_to_int(data[this_year_pledge_fieldname])
            previous_pledge = helpers.jotform_text_to_int(data[last_year_pledge_fieldname])
            out[fid] = {
                'participate'    : participate,
                'current pledge' : current_pledge,
                'previous pledge': previous_pledge,
            }

        return out

    # ------------------------------------------------------------------------

    # Compares the dictionaries of pledges from this year to that of last year,
    # and outputs a CSV showing a few statistics relating to which category the
    # pledges falls into (Can't, Reduced, No Change, New, Increased, No Pledge
    # Both Years) and some relevant info / analysis on totals and percentages.
    def _compare(this_year_data, log):
        out = dict()
        for fid in this_year_data:
            current_pledge = this_year_data[fid]["current pledge"]
            previous_pledge = this_year_data[fid]["previous pledge"]
            if this_year_data[fid]["participate"] == False:
                category = "Cannot pledge"
            # If the family is recorded as having a pledge of 0 OR 1 last year,
            # then they did not have a pledge last year.
            elif previous_pledge == 0 or previous_pledge == 1:
                # Set their previous pledge to zero to accurately reflect how
                # much they *actually* pledged last year.
                previous_pledge = 0
                if current_pledge == 0:
                    category = "No pledge both years"
                # If the family didn't pledge last year, but pledged this year,
                # it's a new pledge.
                elif current_pledge > 0:
                    category = "NEW pledge"
            elif current_pledge == previous_pledge:
                category = "No change"
            elif current_pledge > previous_pledge:
                category = "Increased pledge"
            elif current_pledge < previous_pledge:
                category = "Reduced pledge"

            dollar_impact = current_pledge - previous_pledge

            if category not in out:
                out[category] = {
                    'households'       : 0,
                    'dollar impact'    : 0,
                    'total of pledges' : 0,
                }

            out[category]["households"] += 1
            out[category]["dollar impact"] += dollar_impact
            out[category]["total of pledges"] += current_pledge

        return out

    # ------------------------------------------------------------------------

    def _make_xlsx(comparison, log):
        workbook = Workbook()
        sheet = workbook.active

        comments_label    = "Comments"
        pledge_last_label = f'CY{stewardship_year-1} pledge'
        pledge_cur_label  = f'CY{stewardship_year} pledge'
        amount_label      = f'CY{stewardship_year-1} gifts'

        # Setup the title rows
        # Title rows + set column widths
        title_font   = Font(color='FFFF00')
        title_fill   = PatternFill(fgColor='0000FF', fill_type='solid')
        title_align  = Alignment(horizontal='center', wrap_text=True)

        wrap_align   = Alignment(horizontal='general', wrap_text=True)
        right_align  = Alignment(horizontal='right')

        money_format = "$##,###,###,###"
        percentage_format = "##.#"

        xlsx_cols = dict();
        def _add_col(name, width=15, format=None):
            col             = len(xlsx_cols) + 1
            xlsx_cols[name] = {'name' : name, 'format' : format,
                               'column' : col, 'width' : width }

        _add_col('Category', width=20)
        _add_col('Number of Households')
        _add_col('%-age of Total Submitted Households')
        _add_col('Dollar Impact')
        _add_col('Total of Pledges Submitted')
        _add_col('%-age of Total Pledges Submitted')

        # Make 2 rows of merged cells for wide titles
        def _make_merged_title_row(row, value):
            cell           = sheet.cell(row=row, column=1, value=value)
            cell.fill      = title_fill
            cell.font      = title_font
            cell.alignment = title_align
            end_col_char   = chr(ord('A') - 1 + len(xlsx_cols))
            sheet.merge_cells(f'A{row}:{end_col_char}{row}')

        _make_merged_title_row(row=1, value='eStewardship Pledge Analysis')
        _make_merged_title_row(row=2, value='')

        # Now add all the column titles
        for data in xlsx_cols.values():
            col            = data['column']
            cell           = sheet.cell(row=3, column=col, value=data['name'])
            cell.fill      = title_fill
            cell.font      = title_font
            cell.alignment = title_align

            col_char = chr(ord('A') - 1 + col)
            sheet.column_dimensions[col_char].width = data['width']

        # Finally, fill in all the data rows.
        # First, compute totals so that we can compute percentages.
        total_households = 0
        total_pledges    = 0
        total_impact     = 0
        for data in comparison.values():
            total_households += data['households']
            total_pledges    += data['total of pledges']
            total_impact     += data['dollar impact']

        def _fill(column, value, align=None, format=None):
            cell = sheet.cell(row=xlsx_row, column=column, value=value)

            want_format = True
            if (type(value) is int or type(value) is float) and value == 0:
                    want_format = False

            if want_format:
                if align:
                    cell.alignment = align
                if format:
                    cell.number_format = format

        xlsx_row = 4
        for category, data in comparison.items():
            _fill(1, category)
            _fill(2, data['households'])
            _fill(3, data['households'] / total_households * 100.0,
                format=percentage_format)
            _fill(4, data['dollar impact'],
                format=money_format)
            _fill(5, data['total of pledges'],
                format=money_format)
            _fill(6, data['total of pledges'] / total_pledges * 100.0,
                format=percentage_format)

            xlsx_row += 1

        _fill(1, 'Totals', align=right_align)
        _fill(2, total_households)
        _fill(4, total_impact, format=money_format)
        _fill(5, total_pledges, format=money_format)

        return workbook

    # ------------------------------------------------------------------------

    # Make simplified data structures from the full jotform data.  These will be
    # easier to compare. We use this year's jotform for last year's data because
    # it (this year's jotform) has pre-filled info on how much the family gave
    # last year.
    this_year_data = _simplify_jotform(jotform_this_year,
                                       'CY2022 participation', 'CY2022 pledge', 'CY2021 pledge', log)

    # Do the comparison
    comparison = _compare(this_year_data, log)

    # Make an XLSX report
    workbook = _make_xlsx(comparison, log)

    # Upload the XLS to google
    now      = datetime.now()
    filename = f'{now.year:04}-{now.month:02}-{now.day:02} Pledge analysis.xlsx'
    gsheet_id, _ = upload_xlsx_to_gsheet(google,
                    google_folder_id=upload_team_drive_folder_id,
                    filename=filename,
                    workbook=workbook,
                    remove_local=True,
                    log=log)

    return gsheet_id

##############################################################################

def send_reports_email(time_period, comments_gfile, pledge_analysis_gfile, args, log):
    # Send the comments report email
    body = list()
    body.append(f"""<html>
<body>
<h2>{title} comments and pledge analysis reports</h2>

<h3>Time period: {time_period}</h3>""")

    if comments_gfile is None:
        body.append("<p>No comments submitted during this time period.</p>")
    else:
        url = 'https://docs.google.com/spreadsheets/d/{id}'.format(id=comments_gfile)
        body.append(f'<p><a href="{url}">Link to Google sheet containing comments for this timeframe</a>.</p>')

    url = 'https://docs.google.com/spreadsheets/d/{id}'.format(id=pledge_analysis_gfile)
    body.append(f'''<p><a href="{url}">Link to Google sheet containing pledge analysis.</a>.</p>
</body>
</html>''')

    to = reports_email_to
    subject = '{subj} ({t})'.format(subj=reports_email_subject, t=time_period)
    try:
        log.info('Sending "{subject}" email to {to}'
                 .format(subject=subject, to=to))
        with smtplib.SMTP_SSL(host=smtp_server,
                              local_hostname='epiphanycatholicchurch.org') as smtp:
            msg = EmailMessage()
            msg['Subject'] = subject
            msg['From'] = smtp_from
            msg['To'] = to
            msg.set_content('\n'.join(body))
            msg.replace_header('Content-Type', 'text/html')

            # This assumes that the file has a single line in the format of username:password.
            with open(args.smtp_auth_file) as f:
                line = f.read()
                smtp_username, smtp_password = line.split(':')

            # Login; we can't rely on being IP whitelisted.
            try:
                smtp.login(smtp_username, smtp_password)
            except Exception as e:
                log.error(f'Error: failed to SMTP login: {e}')
                exit(1)

            smtp.send_message(msg)
    except:
        print("==== Error with {email}".format(email=to))
        print(traceback.format_exc())

##############################################################################

def _family_comparison_reports(pds_families, jotform, log):
    census_keyword = f'Active: Census {stewardship_year - 1}'
    key = 'keywords'

    stewardship_fids = list()
    census_fids = list()

    # Find everyone who did census
    for fid, family in pds_families.items():
        if key not in family:
            continue

        for keyword in family[key]:
            if keyword == census_keyword:
                census_fids.append(int(fid))
                break

    # Index everyone who did estewardship
    # Jotform
    for row in jotform:
        stewardship_fids.append(int(row['fid']))

    # Also look at the PDS keywords to catch people who submitted
    # outside of the jotform
    stewardship_keyword = f'Active: Stewardship {stewardship_year}'
    for fid, family in pds_families.items():
        if key not in family:
            continue

        for keyword in family[key]:
            if keyword == stewardship_keyword:
                stewardship_fids.append(int(fid))
                break

    # Compare
    census_not_stewardship_fids = list()
    stewardship_not_census_fids = list()

    for fid in census_fids:
        if fid not in stewardship_fids:
            census_not_stewardship_fids.append(fid)

    for fid in stewardship_fids:
        if fid not in census_fids:
            # First, ensure that the FID is still in PDS
            if fid not in pds_families:
                continue

            stewardship_not_census_fids.append(fid)


    # Make CSV data
    def _doit(rows):
        out = list()
        for fid in rows:
            family = pds_families[fid]
            last_name = pds_families[fid]['Name'].split(',')[0]
            item = {
                'fid' : fid,
                'Family Name' : family['Name'],
                'Envelope ID' : helpers.pkey_url(family['ParKey']),
                'Last Name' : last_name,
            }
            out.append(item)

        return out

    census_not_stewardship = _doit(census_not_stewardship_fids)
    stewardship_not_census = _doit(stewardship_not_census_fids)

    log.info(f"Found {len(census_not_stewardship)} families who submitted census but not stewardship")
    log.info(f"Found {len(stewardship_not_census)} families who submitted stewardship but not census")

    return census_not_stewardship, stewardship_not_census

#-----------------------------------------------------------------------------

# Find who did census earlier this year, but did not submit
# stewardship in this campaign.
#
# Also find who did stewardship in this compaign but did not submit
# census earlier this year.
def family_comparison_reports(args, google, pds_families, jotform, log):
    def _doit(filename, rows):
        if len(rows) == 0:
            return

        gsheet_id = None
        gsheet_id, csv_filename = upload_csv_to_gsheet(google,
                                        google_folder_id=upload_team_drive_folder_id,
                                        filename=filename,
                                        fieldnames=rows[0].keys(),
                                        csv_rows=rows,
                                        remove_local=False,
                                        log=log)

        url = 'https://docs.google.com/spreadsheets/d/{id}'.format(id=gsheet_id)
        log.info(f"Uploaded to {url}")

    #-------------------------------------------------------------------------

    census_not_stewardship, stewardship_not_census = \
        _family_comparison_reports(pds_families, jotform, log)

    # If we have pledges, upload them to a Google sheet
    _doit('Families who submitted census but not stewardship.csv',
          census_not_stewardship)
    _doit('Families who submitted stewardship but not census.csv',
          stewardship_not_census)

##############################################################################

# Columns that I need
# - fid
# - Recurring Charge Name:
# - Terms / Frequency: Weekly, Biweekly, Monthly, Bimonthly, Semimonthly, Quarterly, Semiannually, Annually
# - Begin Date: 01/01/2020
# - End Date: 12/31/2020
# - Rate
# - Total pledge
def convert_pledges_to_pds_import(pds_families, jotform, log):
    def _map_to_freqrate(pledge):
        freq = pledge[f'CY{stewardship_year} frequency']
        amount = helpers.jotform_text_to_int(pledge[f'CY{stewardship_year} pledge'])

        if freq == 'Weekly donations':
            return 'Weekly', amount / 52
        elif freq == 'Monthly donations':
            return 'Monthly', amount / 12
        elif freq == 'Quarterly donations':
            return 'Quarterly', amount / 4
        elif freq == 'One annual donation':
            return 'Annually', amount
        else:
            return None, None

    #-------------------------------------------------------------------------

    out = list()

    for pledge in jotform:
        # Skip the title row
        fid = pledge['fid']
        if 'fid' in fid:
            continue

        # Here's something that can happen: A family may be deleted from PDS
        # even though they submitted.  In this case, skip them.
        fid = int(fid)
        if fid not in pds_families:
            log.warning(f"WARNING: Family FID {pledge['fid']} / {pledge['Family names']} submitted a pledge, but is no longer in PDS")
            continue

        # If there is a $0 pledge, Per Lynne's comment, we'll
        # transform this into a $1 annual pledge -- just so that this
        # person is on the books, so to speak.
        pledge_field  = f'CY{stewardship_year} pledge'
        pledge_amount = helpers.jotform_text_to_int(pledge[pledge_field])
        if not pledge_amount or pledge_amount == '' or pledge_amount == 0:
            pledge_amount = 1
            pledge[pledge_field] = pledge_amount
            pledge[f'CY{stewardship_year} frequency'] = 'One annual donation'

        frequency, rate = _map_to_freqrate(pledge)

        # Round pledge value and rate to 2 decimal points, max
        rate  = float(int(rate * 100)) / 100.0
        total = float(int(pledge_amount) * 100) / 100.0

        # Use an OrderedDict to keep the fields in order
        row = collections.OrderedDict()
        row['fid']           = pledge['fid']
        row['RecChargeName'] = 'Due/Contributions'
        row['Frequency']     = frequency
        row['BeginDate']     = stewardship_begin_date.strftime('%m/%d/%Y')
        row['EndDate']       = stewardship_end_date.strftime('%m/%d/%Y')
        row['PledgeRate']    = rate
        row['TotalPledge']   = total
        row['SubmitDate']    = pledge['SubmitDate']
        row['Names']         = pledge['Family names']
        row['Envelope ID']   = helpers.pkey_url(pledge['EnvId'])

        # Calculate family pledge values for last CY
        family = pds_families[fid]
        helpers.calculate_family_values(family, stewardship_year - 2, log)

        row[f'CY{stewardship_year - 2} YTD gifts'] = family['calculated']['gifts']

        # Calculate family pledge values for this CY
        helpers.calculate_family_values(family, stewardship_year - 1, log)

        row[f'CY{stewardship_year - 1} YTD gifts'] = family['calculated']['gifts']
        row[f'CY{stewardship_year - 1} pledge']    = family['calculated']['pledged']

        # Add column for how they want to fullfill their pledge
        row[f'CY{stewardship_year} frequency'] = pledge[f'CY{stewardship_year} frequency']
        row[f'CY{stewardship_year} mechanism'] = pledge[f'CY{stewardship_year} mechanisms']

        # Add a column for this Family's "Envelope user" value
        row['PDS Envelope User'] = family['EnvelopeUser']

        # Add a column for whether the Family selected the "offeratory
        # envelopes" option on the Jotform.
        val = False
        if 'Offertory' in pledge[f'CY{stewardship_year} mechanisms']:
            val = True
        row['Jotform asked for Envelopes'] = val

        out.append(row)

    return out

#-----------------------------------------------------------------------------

def family_pledge_csv_report(args, google, pds_families, jotform, log):
    pledges = convert_pledges_to_pds_import(pds_families, jotform, log)

    # If we have pledges, upload them to a Google sheet
    gsheet_id = None
    if len(pledges) > 0:
        filename = 'Family Pledge PDS import.csv'
        gsheet_id, csv_filename = upload_csv_to_gsheet(google,
                                        google_folder_id=upload_team_drive_folder_id,
                                        filename=filename,
                                        fieldnames=pledges[0].keys(),
                                        csv_rows=pledges,
                                        remove_local=False,
                                        log=log)

    # Send the statistics report email
    body = list()
    body.append(f"""<html>
<body>
<h2>{title} pledge update</h2>

""")

    if gsheet_id:
        url = 'https://docs.google.com/spreadsheets/d/{id}'.format(id=gsheet_id)
        body.append(f'<p><a href="{url}">Link to Google sheet containing pledge updates for this timeframe</a>.</p>')
        body.append("<p>See the attachment for a CSV to import directly into PDS.</p>")
    else:
        body.append("<p>There were no pledge submissions during this timeframe.<p>")

    body.append("""</body>
</html>""")

    try:
        to = pledge_email_to
        log.info(f'Sending "{pledge_email_subject}" email to {to}')
        with smtplib.SMTP_SSL(host=smtp_server,
                              local_hostname='epiphanycatholicchurch.org') as smtp:

            # This assumes that the file has a single line in the format of username:password.
            with open(args.smtp_auth_file) as f:
                line = f.read()
                smtp_username, smtp_password = line.split(':')

            # Login; we can't rely on being IP whitelisted.
            try:
                smtp.login(smtp_username, smtp_password)
            except Exception as e:
                log.error(f'Error: failed to SMTP login: {e}')
                exit(1)

            msg = EmailMessage()
            msg['Subject'] = pledge_email_subject
            msg['From'] = smtp_from
            msg['To'] = to
            msg.set_content('\n'.join(body))
            msg.replace_header('Content-Type', 'text/html')

            # If there were results, attach the CSV
            if gsheet_id:
                with open(filename, "rb") as f:
                    csv_data = f.read()
                msg.add_attachment(csv_data, maintype='text', subtype='csv', filename=filename)

            smtp.send_message(msg)

            if gsheet_id:
                os.unlink(filename)
    except:
        print("==== Error with {email}".format(email=to))
        print(traceback.format_exc())

##############################################################################

MINISTRY_PARTICIPATE = 1
MINISTRY_INTERESTED  = 2
MINISTRY_INACTIVE    = 3

def _convert_jotform_ministry_status(jotform_status):
    if jotform_status is None:
        return MINISTRY_INACTIVE

    if 'PARTICIPATE' in jotform_status:
        return MINISTRY_PARTICIPATE
    elif 'INTERESTED' in jotform_status:
        return MINISTRY_INTERESTED
    elif 'NO LONGER' in jotform_status:
        return MINISTRY_INACTIVE
    else:
        # Catch-all
        return MINISTRY_INACTIVE

def _status_to_str(status):
    if status == MINISTRY_PARTICIPATE:
        return 'I already participate'
    elif status == MINISTRY_INTERESTED:
        return 'I am interested'
    else:
        return 'I do not participate'

#-----------------------------------------------------------------------------

# Fields I need:
# - MID
# - Ministry
# - Status
# - Start date (from constants.py)
# - End date (from constants.py)
# Other fields requested by the staff:
# - Member nickname + last name
# - Member email addresses
# - Member phones
# - Member status on ministry
#
# Will generate several outputs:
# - Likely can be directly imported
#   - NOT PARTICIPATE -> INTERESTED
# - Likely cannot be directly imported (still need to test)
#   - PARTICPATE -> NOT PARTICIPATE
# - Needs to be examined by a human
#   - NOT PARTICIPATE -> ALREADY PARTICIPATE
#   - PARTICIPATE -> INTERESTED
#
# Output will be a dictionary:
# - PDS ministry name (in the case of multi-names, like lectors, use the jotform name): dictionary of
#   - interested: list of members
#   - no_longer_interested: list of members
#   - needs_human: list of members
def analyze_member_ministry_submissions(pds_members, pds_families, jotform_csv, log):
    def _get_pds_member_ministry_status(member, ministry_names):
        for ministry in member['active_ministries']:
            if ministry['Description'] in ministry_names:
                if ministry['active']:
                    return MINISTRY_PARTICIPATE, ministry['status']
                else:
                    return MINISTRY_INACTIVE, ministry['status']

        # If the member is not currently active in this ministry, see if they
        # were ever previously a member of this ministry
        results = dict()
        for ministry in member['inactive_ministries']:
            if ministry['Description'] in ministry_names:
                results[ministry['start']] = ministry['status']
                results[ministry['end']]   = ministry['status']

        if len(results) > 0:
            dates = sorted(results.keys())
            end   = dates[-1]
            result_str = f'Inactive, last status on {end}: {results[end]}'
        else:
            result_str = 'Never been a member of this ministry'

        return MINISTRY_INACTIVE, result_str

    #-------------------------------------------------------------------------
    def _status_to_pds(status):
        if status == MINISTRY_PARTICIPATE:
            return 'Actively Involved'
        elif status == MINISTRY_INACTIVE:
            return 'No Longer Involved'
        elif status == MINISTRY_INTERESTED:
            return 'Interested'

    #-------------------------------------------------------------------------

    output = dict()

    # Some stats numbers
    num_members_interested = 0
    num_members_no_longer_interested = 0
    num_members_needs_human = 0

    # Each row is a family
    for jrow in jotform_csv:
        fid = int(jrow['fid'])

        # Make sure the family is still active
        if fid not in pds_families:
            log.warn(f"WARNING: Family {fid} submitted, but cannot be found -- skipping")
            continue

        family = pds_families[fid]
        log.info(f"Processing Jotform Family submission: {family['Name']} (FID {fid})")

        # Check the members in this row
        for member_num in range(MAX_PDS_FAMILY_MEMBER_NUM):
            column_names = jotform_gsheet_columns['members'][member_num]
            # The 0th entry in each Member is the MID
            mid = jrow[column_names[0]]

            # If there's no MID for this Member, then there's no Member.
            # We're done with the loop.
            if mid == '':
                break

            # Here's something that can happen: a MID was submitted, but is no
            # longer an active member in PDS.  In that case, ignore the submission.
            mid = int(mid)
            if mid not in pds_members:
                log.warn(f"WARNING: Member {mid} submitted, but cannot be found -- skipping")
                continue

            member = pds_members[mid]
            log.info(f"  Processing Jotform member {member['email_name']} (MID {mid})")

            # JMS Debug
            #if not m['Name'].startswith('Squyres,Jeff'):
            #    continue

            # For generating stats, below
            member_interested = False
            member_no_longer_interested = False
            member_needs_human = False

            # Go through the list of ministry grids from the jotform
            for grid in jotform.ministry_grids:
                # Go through the rows in the jotform ministry grid
                for mrow in grid.rows:
                    # Each row has its PDS ministry name
                    ministry_entry = mrow['pds_ministry']

                    # Some ministry rows are lists because we want to treat them equivalently
                    if type(ministry_entry) is list:
                        output_key = mrow['row_heading']
                        ministries = ministry_entry
                    else:
                        output_key = ministry_entry
                        ministries = [ ministry_entry ]

                    # Get the Member's status in this Ministry from the jotform
                    jotform_column_name = mrow['jotform_columns'][member_num]
                    jotform_status_str  = jrow[jotform_column_name]
                    jotform_status      = _convert_jotform_ministry_status(jotform_status_str)

                    # Get their status from PDS
                    pds_status, pds_status_string = _get_pds_member_ministry_status(member, ministries)
                    key = 'jotform'
                    if key not in member:
                        member[key] = dict()
                    member[key][output_key] = pds_status_string

                    # If they're the same, nothing to do
                    if jotform_status == pds_status:
                        continue

                    if output_key not in output:
                        output[output_key] = dict()

                    # If PDS INACTIVE -> INTERESTED
                    if (pds_status == MINISTRY_INACTIVE and
                          jotform_status == MINISTRY_INTERESTED):
                        key = 'Interested'
                        if key not in output[output_key]:
                            output[output_key][key] = list()
                        output[output_key][key].append(member)
                        member_interested = True

                    elif (pds_status == MINISTRY_PARTICIPATE and
                          jotform_status == MINISTRY_INACTIVE):
                        key = 'No longer interested'
                        if key not in output[output_key]:
                            output[output_key][key] = list()
                        output[output_key][key].append(member)
                        member_no_longer_interested = True

                    elif (pds_status == MINISTRY_INACTIVE and
                          jotform_status == MINISTRY_PARTICIPATE):
                        key = 'Needs human: PDS=inactive, but Jotform=active'
                        if key not in output[output_key]:
                            output[output_key][key] = list()
                        output[output_key][key].append(member)
                        member_needs_human = True

                    elif (pds_status == MINISTRY_PARTICIPATE and
                          jotform_status == MINISTRY_INTERESTED):
                        key = 'Needs human: PDS=active, but Jotform=interested'
                        if key not in output[output_key]:
                            output[output_key][key] = list()
                        output[output_key][key].append(member)
                        member_needs_human = True

            # For generating stats, below
            if member_interested:
                num_members_interested += 1
            if member_no_longer_interested:
                num_members_no_longer_interested += 1
            if member_needs_human:
                num_members_needs_human += 1

    log.info(f"Total number of Members who were interested in a new ministry: {num_members_interested}")
    log.info(f"Total number of Members who were no longer interested in an existing ministry: {num_members_no_longer_interested}")
    log.info(f"Total number of Members who submitted an ambiguous response: {num_members_needs_human}")

    return output

#-----------------------------------------------------------------------------

def member_ministry_csv_report(args, google, start, end, time_period, pds_members, pds_families, jotform_csv, log):
    def _find_all_phones(member):
        found = list()
        key   = 'phones'
        key2  = 'unlisted'
        if key in member:
            # Is this the old member format?  Seems to be a dict of
            # phone_id / phone_data, and no "Unlisted".  :-(
            for p in member[key]:
                # Skip emergency contacts
                if 'Emergency' in p['type']:
                    continue

                text = f"{p['number']} {p['type']}"
                if key2 in p:
                    if p[key2]:
                        text += ' UNLISTED'
                found.append(text)

        # When used with XLSX word wrapping alignment, this will
        # across put each phone number on a separate line, but all
        # within a single cell.
        return '\r\n'.join(found)

    #--------------------------------------------------------------------

    def _find_family_home_phone(member):
        family = member['family']
        key = 'phones'
        key2 = 'unlisted'
        if key in family:
            for p in family[key]:
                if 'Home' in p['type']:
                    text = f"{p['number']} {p['type']}"
                    if key2 in p:
                        if p[key2]:
                            text += ' UNLISTED'
                    return text

        return ""

    #--------------------------------------------------------------------

    output      = analyze_member_ministry_submissions(pds_members, pds_families, jotform_csv, log)
    today       = date.today()

    #--------------------------------------------------------------------

    def _setup_new_workbook():
        workbook    = Workbook()
        sheet       = workbook.active

        for data in xlsx_cols.values():
            col            = data['column']
            cell           = sheet.cell(row=1, column=col, value=data['name'])
            cell.fill      = title_fill
            cell.font      = title_font
            cell.alignment = title_align

            col_char = chr(ord('A') - 1 + col)
            sheet.column_dimensions[col_char].width = data['width']

        sheet.freeze_panes = sheet['A2']

        return workbook

    #--------------------------------------------------------------------

    def _fill(col_name, value, align=None, format=None):
        col_data = xlsx_cols[col_name]
        cell     = sheet.cell(row=xlsx_row, column=col_data['column'], value=value)
        if align:
            cell.alignment = align
        if format:
            cell.number_format = format

    #--------------------------------------------------------------------

    title_font  = Font(color='FFFF00')
    title_fill  = PatternFill(fgColor='0000FF', fill_type='solid')
    title_align = Alignment(horizontal='center', wrap_text=True)

    wrap_align  = Alignment(horizontal='general', wrap_text=True)
    right_align = Alignment(horizontal='right')

    xlsx_cols = dict();
    def _add_col(name, width=10):
        col             = len(xlsx_cols) + 1
        xlsx_cols[name] = {'name' : name, 'column' : col, 'width' : width }
    _add_col('Full Name', width=20)
    _add_col('First')
    _add_col('Last')
    _add_col('Age')
    _add_col('Envelope ID')
    _add_col('Email', width=30)
    _add_col('Member phones', width=20)
    _add_col('Family home phone', width=20)
    _add_col('Category', width=25)
    _add_col('Current ministry status', width=50)
    _add_col('MID')
    _add_col('PDS ministry name', width=50)

    for ministry_name in sorted(output.keys()):
        workbook = _setup_new_workbook()
        sheet    = workbook.active

        data = output[ministry_name]
        xlsx_row = 2
        for category in sorted(data.keys()):

            for member in data[category]:
                family = member['family']

                _fill('Full Name', member['email_name'])
                _fill('First', member['first'])
                _fill('Last', member['last'])
                if member['date_of_birth']:
                    age = today - member['date_of_birth']
                    _fill('Age', int(age.days / 365))
                _fill('Envelope ID', family['ParKey'])
                emails = PDSChurch.find_any_email(member)
                if emails:
                    _fill('Email', emails[0])
                _fill('Member phones', _find_all_phones(member), align=wrap_align)
                _fill('Family home phone', _find_family_home_phone(member))
                _fill('Category', category.capitalize(), align=wrap_align)
                _fill('Current ministry status', member['jotform'][ministry_name], align=wrap_align)
                _fill('MID', member['MemRecNum'])
                _fill('PDS ministry name', ministry_name)

                xlsx_row += 1

        # Write out the XLSX with the results
        filename = f'{ministry_name} jotform results.xlsx'.replace('/', '-')
        if os.path.exists(filename):
            os.unlink(filename)
        workbook.save(filename)
        log.info(f"Wrote to filename: {filename}")

##############################################################################

def family_status_csv_report(args, google, pds_families, jotform, log):
    # Simple report: FID, Family name, and constants.already_submitted_fam_status

    # Did we find anything?
    if len(jotform) == 0:
        log.info("No submissions -- no statuses to update")
        return

    # Make a dictionary of the final CSV data
    csv_data = list()
    for row in jotform:
        fid = int(row['fid'])
        if fid not in pds_families:
            continue

        last_name = pds_families[fid]['Name'].split(',')[0]
        csv_data.append({
            'fid' : fid,
            'Family Name' : pds_families[fid]['Name'],
            'Envelope ID' : helpers.pkey_url(pds_families[fid]['ParKey']),
            'Last Name'   : last_name,
            'Status'      : already_submitted_fam_status,
            'Keyword'     : already_submitted_fam_keyword,
        })

    filename = ('Family Status and Keyword Update.csv')
    #ef upload_csv_to_gsheet(google, google_folder_id, filename, fieldnames, csv_rows, remove_local, log):

    gsheet_id, csv_filename = upload_csv_to_gsheet(google,
                                google_folder_id=upload_team_drive_folder_id,
                                filename=filename,
                                fieldnames=csv_data[0].keys(),
                                csv_rows=csv_data,
                                remove_local=False,
                                log=log)
    url = f'https://docs.google.com/spreadsheets/d/{gsheet_id}'

    #------------------------------------------------------------------------

    body = list()
    body.append(f"""<html>
<body>
<h2>Family Status data update</h2>

<p> See attached spreadsheet of FIDs that have submitted anything at all in this time period.
The same spreadsheet <a href="{url}">is also available as a Google Sheet</a>.</p>

<p> Total of {len(csv_data)} families.</p>
</body>
</html>""")

    to = fid_participation_email_to
    subject = f'{title} Family Status updates'
    try:
        log.info(f'Sending "{subject}" email to {to}')
        with smtplib.SMTP_SSL(host=smtp_server,
                              local_hostname='epiphanycatholicchurch.org') as smtp:
            msg = EmailMessage()
            msg['Subject'] = subject
            msg['From'] = smtp_from
            msg['To'] = to
            msg.set_content('\n'.join(body))
            msg.replace_header('Content-Type', 'text/html')

            # This assumes that the file has a single line in the format of username:password.
            with open(args.smtp_auth_file) as f:
                line = f.read()
                smtp_username, smtp_password = line.split(':')

            # Login; we can't rely on being IP whitelisted.
            try:
                smtp.login(smtp_username, smtp_password)
            except Exception as e:
                log.error(f'Error: failed to SMTP login: {e}')
                exit(1)

            # If there were results, attach CSV files
            with open(csv_filename, 'rb') as f:
                csv_data = f.read()
            msg.add_attachment(csv_data, maintype='text', subtype='csv',
                            filename=csv_filename)

            smtp.send_message(msg)

            os.unlink(csv_filename)
    except:
        print(f"==== Error with {to}")
        print(traceback.format_exc())

##############################################################################

def _export_gsheet_to_csv(service, start, end, google_sheet_id, fieldnames, log):
    response = service.files().export(fileId=google_sheet_id,
                                      mimeType=Google.mime_types['csv']).execute()
    csvreader = csv.DictReader(response.decode('utf-8').splitlines(),
                               fieldnames=fieldnames)

    rows = list()
    for row in csvreader:
        # Skip title row
        if 'Submission' in row['SubmitDate']:
            continue
        if row['fid'] == '':
            continue

        # As of Sep 2021, Google Sheets CSV export sucks. :-(
        # The value of the "Edit Submission" field from Jotform is something
        # like:
        #
        # =HYPERLINK("https://www.jotform.com/edit/50719736733810","Edit Submission")
        #
        # Google Sheet CSV export splits this into 2 fields.  The first one
        # has a column heading of "Edit Submission" (which is what the
        # Jotform-created sheet column heading it) and contains the long number
        # in the URL.  The 2nd one has no column heading, and is just the words
        # "Edit Submission".  :-(  CSV.DictReader therefore puts a value of
        # "Edit Submission" in a dict entry of "None" (because it has no column
        # heading).
        #
        # For our purposes here, just delete the "None" entry from the
        # DictReader.
        if None in row and row[None] == ['Edit Submission']:
            del row[None]

        # Is this submission between start and end?
        if start is not None and end is not None:
            submit_date = helpers.jotform_date_to_datetime(row['SubmitDate'])
            if submit_date < start or submit_date > end:
                continue

        rows.append(row)

    return rows

#-----------------------------------------------------------------------------

def read_jotform_gsheet(google, start, end, fieldnames, gfile_id, log):
    log.info(f"Downloading Jotform raw data ({gfile_id})...")

    # Some of the field names will be lists.  In those cases, use the first
    # field name in the list.
    final_fieldnames = list()
    final_fieldnames.extend(fieldnames['prelude'])
    for member in fieldnames['members']:
        final_fieldnames.extend(member)
    final_fieldnames.extend(fieldnames['family'])
    final_fieldnames.extend(fieldnames['epilog'])

    csv_data = _export_gsheet_to_csv(google, start, end, gfile_id,
                                      final_fieldnames, log)

    # Deduplicate: save the last row number for any given FID
    # (we only really care about the *last* entry that someone makes)
    out_dict = dict()
    for row in csv_data:
        fid = row['fid']

        # Skip the title row
        if fid == 'fid':
            continue

        out_dict[fid] = row

    # Turn this dictionary into a list of rows
    out_list = [ out_dict[fid] for fid in sorted(out_dict) ]

    return out_list, out_dict

##############################################################################

def setup_args():
    tools.argparser.add_argument('--gdrive-folder-id',
                                 help='If specified, upload a Google Sheet containing the results to this Team Drive folder')

    tools.argparser.add_argument('--all',
                                 action='store_const',
                                 const=True,
                                 help='If specified, run the comparison for all time (vs. running for the previous time period')

    tools.argparser.add_argument('--smtp-auth-file',
                                 required=True,
                                 help='File containing SMTP AUTH username:password')

    global gapp_id
    tools.argparser.add_argument('--app-id',
                                 default=gapp_id,
                                 help='Filename containing Google application credentials')
    global guser_cred_file
    tools.argparser.add_argument('--user-credentials',
                                 default=guser_cred_file,
                                 help='Filename containing Google user credentials')

    args = tools.argparser.parse_args()

    return args

##############################################################################

def main():
    global families, members

    args = setup_args()
    log = ECC.setup_logging(debug=False)

    #---------------------------------------------------------------

    # Calculate the start and end of when we are analyzing in the
    # source data
    start = None
    end = None
    if args.all:
        time_period = 'all results to date'

    else:
        # If not supplied on the command line:
        # Sun: skip
        # Mon: everything from last 3 days (Sat+Sun)
        # Tue-Fri: everything from yesterday
        # Sat: skip
        today = end.strftime('%a')
        if today == 'Sat' or today == 'Sun':
            print("It's the weekend.  Nothing to do!")
            exit(0)
        elif today == 'Mon':
            start = end - timedelta(days=3)
        else:
            start = end - timedelta(days=1)

        # No one wants to see the microseconds
        start = start - timedelta(microseconds=start.microsecond)
        end   = end   - timedelta(microseconds=end.microsecond)

        time_period = '{start} - {end}'.format(start=start, end=end)

    log.info("Time period: {tp}".format(tp=time_period))

    #---------------------------------------------------------------

    log.info("Reading PDS data...")
    (pds, pds_families,
     pds_members) = PDSChurch.load_families_and_members(filename='pdschurch.sqlite3',
                                                        parishioners_only=True,
                                                        log=log)

    #---------------------------------------------------------------

    apis = {
        'drive' : { 'scope'       : Google.scopes['drive'],
                    'api_name'    : 'drive',
                    'api_version' : 'v3', },
    }
    services = GoogleAuth.service_oauth_login(apis,
                                              app_json=args.app_id,
                                              user_json=args.user_credentials,
                                              log=log)
    google = services['drive']

    #---------------------------------------------------------------

    # Load all the results
    log.info(f"Reading Jotform {stewardship_year} data...")
    jotform_all_list, jotform_all_dict = read_jotform_gsheet(google,
                                    start=None, end=None,
                                    fieldnames=jotform_gsheet_columns,
                                    gfile_id=jotform_gsheet_gfile_id,
                                    log=log)

    # Load a range of results
    if start is None:
        jotform_range_list = jotform_all_list.copy()
    else:
        jotform_range_list, jotform_range_dict = read_jotform_gsheet(google,
                                    start=start, end=end,
                                    fieldnames=jotform_gsheet_columns,
                                    gfile_id=jotform_gsheet_gfile_id,
                                    log=log)

    #---------------------------------------------------------------

    # These are periodic reports that are run during the campaign
    if False:
        # Stats of how many families have submitted, etc.
        statistics_report(args, time_period, pds_members, pds_families,
                          jotform_all_list, log)

        # A collection of all the random text comments that people
        # submitted (so that staff members can act on them).
        comments_gfile = None
        comments_gfile = comments_report(args, google, start, end, time_period,
                                         jotform_range_list, log)

        # A comparison of this year's pledges vs. last year's pledges.
        #pledge_gfile = None
        #pledge_gfile = pledge_comparison_report(google, jotform_all_dict,
        #                                        jotform_last_year, log)

        #send_reports_email(time_period, comments_gfile, pledge_gfile, args, log)

    if True:
        family_comparison_reports(args, google, pds_families,
                                  jotform_all_list, log)

    # These reports are generally run after the campaign
    if False:
        # Raw list of pledges (I think this is importable to PDS...?)
        family_pledge_csv_report(args, google, pds_families, jotform_all_list, log)

        # Raw list of families who submitted (this may be importable to PDS...?)
        family_status_csv_report(args, google, pds_families, jotform_all_list, log)

        # Per-ministry CSVs showing member status changes (given to
        # staff members to review, and ultimately to make phone calls
        # to followup).
        member_ministry_csv_report(args, google, start, end, time_period,
                                   pds_members, pds_families, jotform_range_list, log)

main()
