#!/usr/bin/env python3

import sys
import os

import logging.handlers
import logging

# We assume that there is a "ecc-python-modules" sym link in this
# directory that points to the directory with ECC.py and friends.
moddir = os.path.join(os.getcwd(), 'ecc-python-modules')
if not os.path.exists(moddir):
    print("ERROR: Could not find the ecc-python-modules directory.")
    print("ERROR: Please make a ecc-python-modules sym link and run again.")
    exit(1)

sys.path.insert(0, moddir)

import ECC
import Google
import ParishSoftv2 as ParishSoft
import GoogleAuth
import googleapiclient
from google.api_core import retry

from datetime import datetime
from datetime import timedelta

from oauth2client import tools
from googleapiclient.http import MediaFileUpload

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment

from pprint import pprint
from pprint import pformat

# Globals

gapp_id         = 'client_id.json'
guser_cred_file = 'user-credentials.json'

ministry_sheets = [
    {
        'ministry' : '100-Parish Pastoral Council',
        'gsheet_id' : '1aIoStpSOsup8XL5eNd8nhpJwM-IqN2gTkwVf_Qvlylc',
        'birthday' : False,
    },
    {
        'ministry' : '102-Finance Advisory Council',
        'gsheet_id' : '1oGkjyLDexQyb-z53n2luFpE9vU7Gxv0rX6XirtxSjA0',
        'birthday' : False,
    },
    {
        'ministry' : '103-Worship Committee',
        'gsheet_id' : '1h_ZvhkYlnebIu0Tk7h1ldJo-VKnJsJGe1jEzY34mcd0',
        'birthday' : False,
    },
    {
        'ministry' : '104-Stewardship & E Committee',
        'gsheet_id' : '1avaHrl-sHOWOc541GclHZL3ajG30HozISR0HDoswKnM',
        'birthday' : False,
    },
    {
        'ministry' : '107-Social Resp Steering Comm',
        'gsheet_id' : '1Am3v0Pv4D9zubkGYgFbUd8e92PZnBPrbcwKrMrs8AnI',
        'birthday' : False,
    },
    {
        'ministry' : '108-Faith Formation Team',
        'gsheet_id' : '1rWy81dH5ZSdjfXGyFN6nF9DIkVk1NS5G7RmHZDn5FTY',
        'birthday' : False,
    },
    {
        'ministry' : '110-Ten Percent Committee',
        'gsheet_id' : '18BIrnBWf_4LS9XeC9tordSD1SBgJz67a0I9Ouj6ZcEc',
        'birthday' : False,
    },
    {
        'ministry' : '111-Hispanic Ministry Team',
        'gsheet_id' : '1FuUzKKmezkfBcBJdGiX1gbw4PHmqR_nivwAjoFTLjK8',
        'birthday' : False,
    },
    {
        'ministry' : '113-Media Comms Planning Comm.',
        'gsheet_id' : '1erlqkK_F-ol8Bmo5Pg6SoAe355tOxwU3D4pU179foCU',
        'birthday' : False,
    },
    {
        'ministry' : '115-Community Care Committee',
        'gsheet_id' : '1bKiNztcE20jrK24ccoCzOJDuJsA-nAR-P-6BYstIDog',
        'birthday' : False,
    },
    {
        'ministry' : '116-Youth Council',
        'gsheet_id' : '1dM3gNsnTJyQDLeRSmP1iRhjWlXV0aHiOCmdLtG3KnKc',
        'birthday' : False,
    },
    {
        'ministry' : '200-Audit Committee',
        'gsheet_id' : '1OJ3z9tS9qoZMjr4EfSh5zeKy_PKmaRja9htIthjUPpI',
        'birthday' : False,
    },
    {
        'ministry' : '201-Collection Counter',
        'gsheet_id' : '1w-PQ-_jnzWWRB3id-Mh4pNHTHJzf9Al1gc_p7cS0koM',
        'birthday' : False,
    },
    {
        'ministry' : '202-Facility Mgmt & Planning',
        'gsheet_id' : '1kcna05U9Oiy0dyvinjOjLi8tXSwQ3sHBC8N8LqjDsjU',
        'birthday' : False,
    },
    {
        'ministry' : '203-Garden & Grounds',
        'gsheet_id' : '1ZrJCRCrrClQuEq9KPXiRVj2s2u78L6R4-P7hbMpYqzs',
        'birthday' : False,
    },
    {
        'ministry' : '204-Parish Office Volunteers',
        'gsheet_id' : '1yhGeBNaUTuVkEGTRkCh5SC6hQLOpdaEVNpXaq2WmMc4',
        'birthday' : False,
    },
    {
        'ministry' : '206-Space Arrangers',
        'gsheet_id' : '1ZIAn7MdeDVcgvHwR98dXs_ewIaU2uWbl6mvriLSk5V0',
        'birthday' : False,
    },
    {
        'ministry' : '207-Technology Committee',
        'gsheet_id' : '1Gn2m2VMabPkWJWg_NTs6XeGPf_Qi7qPELLxyOx9Q0vU',
        'birthday' : False,
    },
    {
        'ministry' : '208-Weekend Closer',
        'gsheet_id' : '1pj2r_0Xjog22g-kiSmSGkO6BHIIZqW7uXcE3hUs6Xy8',
        'birthday' : False,
    },
    {
        'ministry' : '300-Art & Environment',
        'gsheet_id' : '1uXy2zTQeeH_YtBAR46lDV5X8ADZRgPnNROtzDU_iqv0',
        'birthday' : False,
    },
    {
        'ministry' : '301-Audio/Visual/Light Minstry',
        'gsheet_id' : '1LtsNJc-9KYZkQqy2BITQ4Xgd_ns4Uo7z3YSQdnQrzN8',
        'birthday' : False,
    },
    {
        'ministry' : '303-Linens/Vestments Ministry',
        'gsheet_id' : '1ul5doCFZx4Y_L8ZNmDarVdgLy_dQo2MkE0mdo5kbrY0',
        'birthday' : False,
    },
    {
        'ministry' : '304-LiturgicalPlanningDscrnmnt',
        'gsheet_id' : '11A59wLOy58-ADKm60YuVGDG0N02E8a9iY-VoCjQX9tY',
        'birthday' : False,
    },
    {
        'ministry' : '305-Movers Ministry',
        'gsheet_id' : '1q0rNZQ0Od3cCoFG2qLgKZep10GzQ3nmvVIIJuUOWBgY',
        'birthday' : False,
    },
    {
        'ministry' : '306-Music Support for Children',
        'gsheet_id' : '1ODBCOAUY5g93-ZDrgAx9-Fv5wT4VSeMYxHgpFeb2F4s',
        'birthday' : False,
    },
    {
        'ministry' : '307-Wedding Assistant',
        'gsheet_id' : '1UBzQOEpdYmb1afAnrPDlQcz0hoouvd6mkpEV6rYDMxg',
        'birthday' : False,
    },
    {
        'ministry' : '308-Worship&Music Support Team',
        'gsheet_id' : '1kHRnohAEeaC_2-yTTcgHe36yRVJ4EhWhmbBCZ0P_vI8',
        'birthday' : False,
    },
    {
        'ministry' : '309-Acolytes',
        'gsheet_id' : '1zXfAxnuQCATWQ7nU7QYn2Mvy1Ql1Huy7Ve1DW60nPKA',
        'birthday' : False,
    },
    {
        'ministry' : '310-Adult Choir',
        'gsheet_id' : '1ku8Aq9dXm_mrOq421MWVk7hAqV2Am5FFSgUACOYs2WU',
        'birthday' : False,
    },
    {
        'ministry' : '311-Bell Choir',
        'gsheet_id' : '1UTzXgO9ZLBHB0w-zAW8-u57cgWLbkWWGanJgPC9gboE',
        'birthday' : True,
    },
    {
        'ministry' : '312-Children\'s Music Ministry',
        'gsheet_id' : '11H5SbDti8Jm2HMqdo3mZ3_5iTBky4OmZGdB_WeCOZYI',
        'birthday' : False,
    },
    {
        'ministry' : '313-Eucharistic Ministers',
        'gsheet_id' : '10Aq9XtZHL3v0m0Erm71f8BSUYr7CiEDMwzDWTSgKPJ4',
        'birthday' : False,
    },
    {
        'ministry' : '315-Funeral Mass Ministry',
        'gsheet_id' : '107rbGV_WbEz6m6m8XMliLWIMFKKMz8pQv8pdjgLFvLI',
        'birthday' : False,
    },
    {
        'ministry' : '316-Greeters',
        'gsheet_id' : '1vd7sltcU8MVwIBad__PyXXMN_cUEFvdXpL5RkYiBWJ0',
        'birthday' : False,
    },
    {
        'ministry' : '317-Instrumentalists & Cantors',
        'gsheet_id' : '1YP3sC4dcOWH9Li1rJV8D5FI9mef50xvxqOf6K1K54_U',
        'birthday' : True,
    },
    {
        'ministry' : '318-Lectors',
        'gsheet_id' : '1X796X7_wFZmYoKMzGnj2BFFCOeoncIEILv1cmq_CJB8',
        'birthday' : False,
    },
    {
        'ministry' : '319-Liturgical Dance Ministry',
        'gsheet_id' : '14s4gtJMjK0qiHvN-eQ1zaD76PoSGVd9qcfIPxTDGhrc',
        'birthday' : False,
    },
    {
        'ministry' : '401-Epiphany Companions',
        'gsheet_id' : '1voFdTbY3RMs3R_X-pO6-hbjBfC9Prm80ON-lLDTl2UI',
        'birthday' : False,
    },
    {
        'ministry' : '404-Welcome Desk',
        'gsheet_id' : '1pJwD2UXiDAKng-DFwdWMp0AbZQFSJoiOsJaQgJmfvCE',
        'birthday' : False,
    },
    {
        'ministry' : '406-Evangelization Team',
        'gsheet_id' : '1YLK4dsW2-whQeKPR0g4F8q3ZuximhwLiN1gS_ps4JgQ',
        'birthday' : False,
    },
    {
        'ministry' : '407-Stewardship Team',
        'gsheet_id' : '1T2FCBImLj4akVnX1hpfymi-JKV5jPf3G0ucknzQz8Io',
        'birthday' : False,
    },
    {
        'ministry' : '408-Engagement Team',
        'gsheet_id' : '1vGOdc2YtiNj1VnzUMy0ftrjuwVxqSJz_8XFeTxnQ8l8',
        'birthday' : False,
    },
    {
        'ministry' : '409-Sunday Morning Coffee',
        'gsheet_id' : '1zaVGPdnHp5zZu3zbWfGckIaxbCeCSSTYyoe4pCIlQjs',
        'birthday' : False,
    },
    {
        'ministry' : '411-Men of Epiphany',
        'gsheet_id' : '1Fq11QLU7OQ_yHYPeDPXT8JaA31ackl3eJ2xOvecG_go',
        'birthday' : False,
    },
    {
        'ministry' : '412-Sages (for 50 yrs. +)',
        'gsheet_id' : '11LCDr-Vc3jyeKh5nrd49irscdvTv3TDXhpOoFWlohgs',
        'birthday' : False,
    },
    {
        'ministry' : '413-Singles Explore Life (SEL)',
        'gsheet_id' : '1-uvQO5RRf0K6NJlR_4Mijygn4XGk0zhvowdflKLoEUc',
        'birthday' : False,
    },
    {
        'ministry' : '452-Media Communications',
        'gsheet_id' : '1QLcjLidJppCfhqaNBn10h1Toe81DEy48PwwJAhwDNUs',
        'birthday' : False,
    },
    {
        'ministry' : '500-Bereavement Receptions',
        'gsheet_id' : '17QXoqgreLu8sUQpZfooNkto4NV06wdoE3yA6Q5vgiAw',
        'birthday' : False,
        'role sheets' : [
            {
                'name' : '500-Bereavement Receptions: Team 1',
                'roles' : [ 'Staff', 'Chairperson', 'Team 1 Leader', 'Team 1 Member', 'Team Member: Both' ],
                'gsheet_id' : '1UKk8-zsi4m1i271-SbM4UhkREs0dYcZSpmwGsV7m9rQ',
            },
            {
                'name' : '500-Bereavement Receptions: Team 2',
                'roles' : [ 'Staff', 'Chairperson', 'Team 2 Leader', 'Team 2 Member', 'Team Member: Both' ],
                'gsheet_id' : '1XGvOa98piwKa05yYUIWXIXGXyMHNYSYPMAyvwwSl28A',
            },
        ],
    },
    {
        'ministry' : '501-Eucharist to Sick&Homebnd',
        'gsheet_id' : '1KkCF2V4JIK65b9QsfSG-lGMCASXPavxIxim-CfzxBYk',
        'birthday' : False,
    },
    {
        'ministry' : '505-Healing Blanket Ministry',
        'gsheet_id' : '1fKXe-NuoObjgYoFRcbuJQwIOjFIAfT7TQB87eccgfng',
        'birthday' : False,
    },
    {
        'ministry' : '508-Messages of Hope Ministry',
        'gsheet_id' : '1NOijo2AHwftLBGDyCiyGMEIp4Qmfw-0jVjHG9Ygz3S8',
        'birthday' : False,
    },
    {
        'ministry' : '509-HOPE Support Groups',
        'gsheet_id' : '1Ua7pm-4Av7vPhU8quL5sAEKfowo8O_jfEF34vD1s3Wg',
        'birthday' : False,
    },
    {
        'ministry' : '510-Flower Delivery to SHB',
        'gsheet_id' : '160CPnzR_Q-kph_36LLCPFMC_pLRdMFbUQ2fkQ90LEw0',
        'birthday' : False,
    },
    {
        'ministry' : '700-Advocates for Common Good',
        'gsheet_id' : '1r6FNXGn-T5anj9X7HGlB2EpZ9oXfy5dQXD6z-OQEGgE',
        'birthday' : False,
    },
    {
        'ministry' : '701-CLOUT',
        'gsheet_id' : '1k_hH1tEWBGuERCmFvhZxKOfAsBkqy0uZ16LAd0_jMDg',
        'birthday' : False,
    },
    {
        'ministry' : '703-Eyeglass Ministry',
        'gsheet_id' : '1Iz8hz7NAhh9-dVMiC7mL8yYFi_qmM_ayB5IXhJU0uPw',
        'birthday' : False,
    },
    {
        'ministry' : '704-Habitat for Humanity',
        'gsheet_id' : '1gBQXnTgxodILkjXvBrfJ-sztVj6NZaZQI1kbddAESNE',
        'birthday' : False,
    },
    {
        'ministry' : '705-Hunger & Poverty Ministry',
        'gsheet_id' : '1i3EBKO3Lj3lIprhnoc4VeabEIPAuce5JeV99y1zrUAc',
        'birthday' : False,
    },
    {
        'ministry' : '706-Prison Ministry',
        'gsheet_id' : '1HacDJsMK-oLKjuPjvrABhbYn-b0joox3GMY9uEA--yg',
        'birthday' : False,
    },
    {
        'ministry' : '707-St. Vincent De Paul',
        'gsheet_id' : '1m0Cp7k0XyeJvZ8Z0IZXrudD3oALfeLr0sWvvnEEgn0o',
        'birthday' : False,
    },
    {
        'ministry' : '709-Twinning Committee:Chiapas',
        'gsheet_id' : '1lXkeHsyHNqHYH4zs_HiQdaq8PK8vh5W2koQE5sCq_7U',
        'birthday' : False,
    },
    {
        'ministry' : '710-Creation Care Team',
        'gsheet_id' : '1tFmzvu53v-oBF3P59fl4zv46VHt4PRu066YyceuYNl4',
        'birthday' : False,
    },
    {
        'ministry' : '800-Catechists for Children',
        'gsheet_id' : '1jsoRxugVwXi_T2IDq9J-mEVdzS8xaOk9kuXGAef-YaQ',
        'birthday' : False,
    },
    {
        'ministry' : '802-Gather the Children',
        'gsheet_id' : '1SqqnAFEZlUdCAxp6NYKqpqGFy13L_aheVEGzHLLzOuk',
        'birthday' : False,
    },
    {
        'ministry' : '805-Monday Adult Bible Study',
        'gsheet_id' : '1rBKiweOBT-JZFbfv5Sf4_vjjOlAYnOGaZIhadtBPxSE',
        'birthday' : False,
    },
    {
        'ministry' : '807-Catechumenate/InitiationTm',
        'gsheet_id' : '1lI9eO0bryD6GFIsgJ5FipGePK4FaJ7v8eIp71DFz_6A',
        'birthday' : False,
    },
    {
        'ministry' : '808-BibleTimes Core Team',
        'gsheet_id' : '1gANQt7PBc8erErQRHrRLtkdmOF3Y7DbuLC3J6aRLu4U',
        'birthday' : False,
    },

    {
        'ministry' : '901-Youth Ministry Adult Vols',
        'gsheet_id' : '13neLXFrDTsohe_N_CVPX7ajx4pFvYb9miFtZtUqdVl4',
        'birthday' : False,
    },
    {
        'ministry' : '902-Adult Advisory Council',
        'gsheet_id' : '1yW9tUxef7slTboXYXlVoWrAH80vhP-IAZ53MBjXYUPA',
        'birthday' : False,
    },
    {
        'ministry' : '903-Confirmation Core Team',
        'gsheet_id' : '1Ay9Zp3HnHmINygZWIilVWMvlbD4IfKygyOoa9wU4FK8',
        'birthday' : False,
    },
]

workgroups = [
    {
        'workgroup' : 'Livestream Team',
        'gsheet_id' : '1Yku0IFuIKZCeUNGB5c_Ser_geYkylC2o1tiVfaNwkx8',
        'birthday'  : False,
    },
    {
        'workgroup' : 'YouthMin parent: Jr high',
        'gsheet_id' : '1VIs-AezopoWd3rpVU_kqMOTTXlBPpjI9IxAbhDpSnG4',
        'birthday'  : False,
    },
    {
        'workgroup' : 'YouthMin parent: Sr high',
        'gsheet_id' : '1B_2LeR0EbG3oDUqifvB51YhCuj30lhaYHfUb9bmCj5A',
        'birthday'  : False,
    },
]

####################################################################

def write_xlsx(members, ministry_name, name, want_birthday, log):
    # Make the microseconds be 0, just for simplicity
    now = datetime.now()
    us = timedelta(microseconds=now.microsecond)
    now = now - us

    timestamp = ('{year:04}-{mon:02}-{day:02} {hour:02}:{min:02}'
                .format(year=now.year, mon=now.month, day=now.day,
                        hour=now.hour, min=now.minute))
    filename_base = name
    if filename_base is None:
        filename_base = ministry_name
    filename_base = filename_base.replace("/", "-")
    filename = (f'{filename_base} members as of {timestamp}.xlsx')

    # Put the members in a sortable form (they're currently sorted by MID)
    sorted_members = dict()
    for m in members:
        # 'Name' will be "Last,First..."
        sorted_members[m['display_FullName'] + " " + str(m['memberDUID'])] = m

    wb = Workbook()
    ws = wb.active

    # Title rows + set column widths
    title_font = Font(color='FFFF00')
    title_fill = PatternFill(fgColor='0000FF', fill_type='solid')
    title_align = Alignment(horizontal='center')

    last_col = 'D'
    if want_birthday:
        last_col = 'E'

    row = 1
    ws.merge_cells(f'A{row}:{last_col}{row}')
    cell = f'A{row}'
    ws[cell] = f'Ministry: {ministry_name}'
    ws[cell].fill = title_fill
    ws[cell].font = title_font

    row = row + 1
    ws.merge_cells(f'A{row}:{last_col}{row}')
    cell = f'A{row}'
    ws[cell] = f'Last updated: {now}'
    ws[cell].fill = title_fill
    ws[cell].font = title_font

    row = row + 1
    ws.merge_cells(f'A{row}:{last_col}{row}')
    cell = f'A{row}'
    ws[cell] = ''
    ws[cell].fill = title_fill
    ws[cell].font = title_font

    row = row + 1
    columns = [(f'A{row}', 'Member name', 30),
               (f'B{row}', 'Address', 30),
               (f'C{row}', 'Phone / email', 50)]

    role = 'D'
    if want_birthday:
        columns.append((f'D{row}', 'Birthday', 30))
        role = 'E'
    columns.append((f'{role}{row}', 'Role', 20))

    for cell,value,width in columns:
        ws[cell] = value
        ws[cell].fill = title_fill
        ws[cell].font = title_font
        ws[cell].alignment = title_align
        ws.column_dimensions[cell[0]].width = width

    # Freeze the title row
    row = row + 1
    ws.freeze_panes = ws[f'A{row}']

    #---------------------------------------------------------------------

    def _append(row, col, value):
        if value is None or len(value.strip()) == 0:
            return row

        _ = ws.cell(row=row, column=col, value=value)
        return row + 1

    # Data rows
    for name in sorted(sorted_members):
        m = sorted_members[name]
        # The name will take 1 row
        _ = ws.cell(row=row, column=1, value=m['py friendly name LF'])

        # The address will take multiple rows
        col = 2
        last_row = row
        f = m['py family']
        last_row = _append(col=col, row=last_row, value=f['primaryAddress1'])
        last_row = _append(col=col, row=last_row, value=f['primaryAddress2'])
        val = '{cs}, {state} {zip}'.format(cs=f['primaryCity'],
                                           state=f['primaryState'],
                                           zip=f['primaryPostalCode'])
        last_row = _append(col=col, row=last_row, value=val)
        addr_last_row = last_row
        email_last_row = last_row

        # The phone / email may be more than 1 row
        col = 3
        last_row = row
        phones = ParishSoft.get_member_public_phones(m)
        for phone in phones:
            val = '{ph} {type}'.format(ph=phone['number'], type=phone['type'])
            last_row = _append(col=col, row=last_row, value=val)

        # If we have any preferred emails, list them all
        email = ParishSoft.get_member_public_email(m)
        if email is not None:
            last_row = _append(col=col, row=last_row, value=email)
            email_last_row = last_row

        # The birthday will only be 1 row
        if want_birthday:
            col = 4
            key = 'birthdate'
            if key in m and m[key] is not None:
                birthday = f'{m[key].strftime("%B")} {m[key].day}'
                _append(col=col, row=row, value=birthday)

        # Role
        col += 1
        _append(col=col, row=row, value=m['py ministry role'])

        # Between the address / phone+email, find the real last row
        last_row = max(email_last_row, addr_last_row)
        row = last_row + 1

    #---------------------------------------------------------------------

    wb.save(filename)
    log.info(f'Wrote {filename}')

    return filename

#-------------------------------------------------------------------

@retry.Retry(predicate=Google.retry_errors)
def upload_overwrite(filename, google, file_id, log):
    # Strip the trailing ".xlsx" off the Google Sheet name
    gsheet_name = filename
    if gsheet_name.endswith('.xlsx'):
        gsheet_name = gsheet_name[:-5]
    mime_type = Google.mime_types['sheet']
    #mime_type = Google.mime_types['xlsx']

    try:
        log.info(f'Uploading file update to Google file ID "{file_id}"')
        metadata = {
            'name'     : gsheet_name,
            'mimeType' : Google.mime_types['sheet'],
            'supportsAllDrives' : True,
            }
        media = MediaFileUpload(filename,
                                mimetype=Google.mime_types['xlsx'],
                                resumable=True)
        file = google.files().update(body=metadata,
                                     fileId=file_id,
                                     media_body=media,
                                     supportsAllDrives=True,
                                     fields='id').execute()
        log.debug(f'Successfully updated file: "{filename}" (ID: {file["id"]})')

    except Exception as e:
        # When errors occur, we do want to log them.  But we'll re-raise them to
        # let an upper-level error handler handle them (e.g., retry.Retry() may
        # actually re-invoke this function if it was a retry-able Google API
        # error).
        log.error('Google file update failed for some reason:')
        log.error(e)
        raise e

#-------------------------------------------------------------------

def _create_roster(ps_members, ministry_name, sheet_name,
                   birthday, gsheet_id, google, log):
    members = [ x for x in ps_members.values() ]

    # Make an xlsx
    filename = write_xlsx(members=members,
                          ministry_name=sheet_name,
                          name=ministry_name,
                          want_birthday=birthday, log=log)
    log.debug(f"Wrote temp XLSX file: {filename}")

    # Upload the xlsx to Google
    upload_overwrite(filename=filename, google=google, file_id=gsheet_id,
                     log=log)
    log.debug("Uploaded XLSX file to Google")

    # Remove the temp local XLSX file
    try:
        os.unlink(filename)
        log.debug("Unlinked temp XLSX file")
    except Exception as e:
        log.info("Failed to unlink temp XLSX file!")
        log.error(e)

#-------------------------------------------------------------------

def create_ministry_roster(ps_members, ps_ministries, ministry_sheet, google, log):
    log.info(f"Making ministry roster for: {ministry_sheet}")
    gsheet_id = ministry_sheet['gsheet_id']
    birthday  = ministry_sheet['birthday']

    key1 = 'ministry'
    key2 = 'ministries'
    if key1 in ministry_sheet:
        sheet_name = ministry_sheet[key1]
        ministries = [ sheet_name ]
    elif key2 in ministry_sheet:
        sheet_name = ', '.join(ministry_sheet[key2])
        ministries = ministry_sheet[key2]
    else:
        print(f"ERROR: Cannot find {key1} or {key2} in ministry_sheet!")
        print(ministry_sheet)
        exit(1)

    name = None
    key = 'name'
    if key in ministry_sheet:
        name = ministry_sheet[key]

    # Find the PS Ministry Membershp with the same name
    ministry = None
    for entry in ps_ministries.values():
        if entry['name'] == ministry_sheet['ministry']:
            ministry = entry
    if ministry is None:
        log.error(f"Could not find PS ministry named '{ministry_sheet['ministry']}'")
        exit(1)
    log.debug(f"Found ministry: {ministry['name']}")

    # Make a list of PS Members (objects) that are in this ministry
    members = dict()
    for entry in ministry['membership']:
        mem_duid = entry['memberId']
        if mem_duid not in ps_members:
            log.error(f"Cannot find PS Member with DUID {mem_duid} -- inactive member?")
            exit(1)

        members[mem_duid] = ps_members[mem_duid]
        # We stash this on the member so that it's easy to find later,
        # when making the roster.  This will continually be
        # overwritten on the member with the role for the current
        # ministry.
        members[mem_duid]['py ministry role'] = entry['ministryRoleName']

    if members is None or len(members) == 0:
        log.info(f"No members in ministry: {sheet_name} -- writing empty sheet")

    _create_roster(members, name, sheet_name,
                   birthday, gsheet_id, google, log)

    # Are there any sub-sheets to create?
    key = 'role sheets'
    if key not in ministry_sheet:
        log.debug("No role sub sheets -- returning")
        return

    for role_sheet in ministry_sheet[key]:
        # Filter the members list by those with the role(s) listed
        role_members = {}
        for duid, member in members.items():
            if member['py ministry role'] in role_sheet['roles']:
                role_members[duid] = member

        gsheet_id = role_sheet['gsheet_id']
        name = role_sheet['name']
        sheet_name = name
        _create_roster(role_members, name, sheet_name,
                    birthday, gsheet_id, google, log)

#-------------------------------------------------------------------

def create_workgroup_roster(ps_members, ps_mem_workgroups, workgroup_sheet, google, log):
    log.info(f"Making roster for Member Workgroup: {workgroup_sheet}")
    gsheet_id = workgroup_sheet['gsheet_id']
    birthday  = workgroup_sheet['birthday']

    key = 'workgroup'
    if key not in workgroup_sheet:
        print(f"ERROR: Cannot find {key} in workgroup_sheet!")
        print(ministry_entry)
        exit(1)

    # The "Leader" workgroups end in " Ldr"
    leader_suffix = ' Ldr'
    wg_name = workgroup_sheet[key]
    # IMPORTANT: Have the "leader suffix" group come last!  We want to
    # overwrite the "py member role" on the Member of leaders (below).
    wg_names = [ wg_name,
                 f'{wg_name}{leader_suffix}' ]

    # Find the PS Member Workgroup with the same name or ldr_name.
    # Fill the members dictionary will all of its members (the
    # dictionary will de-duplicate based on DUID).
    members = dict()
    for wg in ps_mem_workgroups.values():
        if wg['name'] not in wg_names:
            continue

        # Make a list of PS Members (objects) that are in this
        # member workgroup
        for member in wg['membership']:
            mem_duid = member['memberId']
            if mem_duid not in ps_members:
                log.error(f"Cannot find PS Member with DUID {mem_duid} -- inactive member?")
                exit(1)

            members[mem_duid] = ps_members[mem_duid]
            # We stash this on the member so that it's easy to find
            # later, when making the roster.  This will continually be
            # overwritten on the member with the role for the current
            # ministry.
            if wg['name'].endswith(leader_suffix):
                members[mem_duid]['py ministry role'] = 'Leader'
            else:
                members[mem_duid]['py ministry role'] = 'Member'

    if members is None or len(members) == 0:
        log.info(f"No members in ministry: {sheet_name} -- writing empty sheet")

    _create_roster(members, wg_name, wg_name,
                   birthday, gsheet_id, google, log)

####################################################################

def setup_cli_args():
    tools.argparser.add_argument('--logfile',
                                 help='Also save to a logfile')
    tools.argparser.add_argument('--debug',
                                 action='store_true',
                                 default=False,
                                 help='Be extra verbose')

    tools.argparser.add_argument('--ps-api-keyfile',
                                 required=True,
                                 help='File containing the ParishSoft API key')
    tools.argparser.add_argument('--ps-cache-dir',
                                 default='.',
                                 help='Directory to cache the ParishSoft data')

    global gapp_id
    tools.argparser.add_argument('--app-id',
                                 default=gapp_id,
                                 help='Filename containing Google application credentials')
    global guser_cred_file
    tools.argparser.add_argument('--user-credentials',
                                 default=guser_cred_file,
                                 help='Filename containing Google user credentials')

    args = tools.argparser.parse_args()

    # Read the PS API key
    if not os.path.exists(args.ps_api_keyfile):
        print(f"ERROR: ParishSoft API keyfile does not exist: {args.ps_api_keyfile}")
        exit(1)
    with open(args.ps_api_keyfile) as fp:
        args.api_key = fp.read().strip()

    return args

####################################################################
#
# Main
#
####################################################################

def main():
    args = setup_cli_args()
    log = ECC.setup_logging(info=True,
                            debug=args.debug,
                            logfile=args.logfile)

    log.info("Reading PS data...")
    families, members, family_workgroups, member_workgroups, ministries = \
        ParishSoft.load_families_and_members(api_key=args.api_key,
                                             cache_dir=args.ps_cache_dir,
                                             active_only=True,
                                             parishioners_only=False,
                                             log=log)

    apis = {
        'drive' : { 'scope'       : Google.scopes['drive'],
                    'api_name'    : 'drive',
                    'api_version' : 'v3', },
    }
    services = GoogleAuth.service_oauth_login(apis,
                                              app_json=args.app_id,
                                              user_json=args.user_credentials,
                                              log=log)
    google = services['drive']

    for sheet in ministry_sheets:
        create_ministry_roster(ps_members=members,
                               ps_ministries=ministries,
                               ministry_sheet=sheet,
                               google=google,
                               log=log)

    for sheet in workgroups:
        create_workgroup_roster(ps_members=members,
                                ps_mem_workgroups=member_workgroups,
                                workgroup_sheet=sheet,
                                google=google,
                                log=log)

if __name__ == '__main__':
    main()
