#!/usr/bin/env python3

# Basic script to create a list of all PDS trainings of a given type.
#
# Sidenote: we used to make XLSX/Google Sheets for sub-ministries of
# the main 313 ministry.  Hence, we have a nice subroutine that was
# capable of emitting both each of the sub-ministry sheets, and the
# union of everyone in all of those sub-ministries.  We no longer have
# those sub-ministries in PDS, but the infrastructure still remains
# here in this code.
#
# This code is still somewhat of a mish-mash of requirements that no
# longer exist, and should probably be cleaned up.

import sys
import os

import logging.handlers
import logging

# We assume that there is a "ecc-python-modules" sym link in this
# directory that points to the directory with ECC.py and friends.
moddir = os.path.join(os.getcwd(), 'ecc-python-modules')
if not os.path.exists(moddir):
    print("ERROR: Could not find the ecc-python-modules directory.")
    print("ERROR: Please make a ecc-python-modules sym link and run again.")
    exit(1)

sys.path.insert(0, moddir)

import ECC
import Google
import PDSChurch
import GoogleAuth
from google.api_core import retry

from datetime import date
from datetime import datetime
from datetime import timedelta

from pprint import pprint

from oauth2client import tools
from googleapiclient.http import MediaFileUpload

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment

from datetimerange import DateTimeRange

# Globals

default_gapp_id         = 'client_id.json'
default_guser_cred_file = 'user-credentials.json'

now   = datetime.now(tz=ECC.local_tz)
today = now.date()

# This is the PDS training name we're looking for
training_name = 'Communion Minister Certificatn'

# These are all the ministries where that training name is relevant
name_313  = '313-Communion Ministers'
gsheet_id_313 = '1ZFGpGlAnB7B_cHfxZjPX0ITGZ51tYShGwesVolRrsbw'

# This is the Google sheet ID of the "About to expire" report
about_to_expire_gsheet_id = '1hMga-0PWSAl91eL3FBBhyMwUArT13xS9YmSvVZgJ7n8'
about_to_expire = timedelta(days=6*30)

#---------------------------------------------------------------------------

def upload_workbook(filename, wb, args, google,
                    gsheet_id, log):
    @retry.Retry(predicate=Google.retry_errors)
    def _upload_overwrite(local_filename, gsheet_filename, google, file_id, log):
        try:
            log.info(f'Uploading file update to Google file ID {file_id}')
            metadata = {
                'name'     : gsheet_filename,
                'mimeType' : Google.mime_types['sheet'],
                'supportsAllDrives' : True,
            }
            media = MediaFileUpload(local_filename,
                                    mimetype=Google.mime_types['sheet'],
                                    resumable=True)
            file = google.files().update(body=metadata,
                                         fileId=file_id,
                                         media_body=media,
                                         supportsAllDrives=True,
                                         fields='id').execute()
            log.debug(f"Successfully updated file: {filename} (ID: {file['id']})")

        except Exception as e:
            # When errors occur, we do want to log them.  But we'll
            # re-raise them to let an upper-level error handler handle
            # them (e.g., retry.Retry() may actually re-invoke this
            # function if it was a retry-able Google API error).
            log.error('Google file update failed for some reason:')
            log.error(e)
            raise e

    #--------------------------------------------------------------

    # Save the xlsx
    local_filename  = f'{filename}.xlsx'
    gsheet_filename = filename
    wb.save(local_filename)
    log.info(f"Wrote temp XLSX file: {local_filename}")

    if args.dry_run:
        log.warning("Dry run: not uploading to Google")
        return

    # Upload xlsx to Google
    _upload_overwrite(local_filename=local_filename,
                      gsheet_filename=gsheet_filename,
                      google=google,
                      file_id=gsheet_id, log=log)

    # Remove temp local xlsx file
    try:
        os.unlink(local_filename)
        log.debug("Unlinked temp XLSX file")
    except:
        log.info(f"Failed to unlink temp XLSX file {local_filename}!")

#------------------------------------------------------------------

def xlsx_create_headers(ws, title_row, fixed_columns,
                        dynamic_column_name_fn, log):
    title_font = Font(color='FFFF00')
    title_fill = PatternFill(fgColor='0000FF', fill_type='solid')
    title_align = Alignment(horizontal='center', wrap_text=True)

    # Make a list of the columns first, so that we know how wide
    # it will be (i.e., what the last column used will be)
    columns = fixed_columns.copy()
    last_fixed_col = chr(ord('A') + len(columns) - 1)

    if dynamic_column_name_fn:
        dynamic_columns = dynamic_column_name_fn(log)
        columns.extend(dynamic_columns)
        last_col = chr(ord('A') + len(columns) - 1)
    else:
        last_col = last_fixed_col

    row = 1
    ws.merge_cells(f'A{row}:{last_col}{row}')
    cell = f'A{row}'
    ws[cell] = title_row
    ws[cell].fill = title_fill
    ws[cell].font = title_font

    row += 1
    ws.merge_cells(f'A{row}:{last_col}{row}')
    cell = f'A{row}'
    ws[cell] = f'Last updated: {now}'
    ws[cell].fill = title_fill
    ws[cell].font = title_font

    row += 1
    ws.merge_cells(f'A{row}:{last_col}{row}')
    cell = f'A{row}'
    ws[cell] = ''
    ws[cell].fill = title_fill
    ws[cell].font = title_font

    col = 'A'
    row += 1
    for value, width in columns:
        cell = f'{col}{row}'
        ws[cell] = value
        ws[cell].fill = title_fill
        ws[cell].font = title_font
        ws[cell].alignment = title_align
        ws.column_dimensions[cell[0]].width = width

        col = chr(ord(col) + 1)

    row += 1
    ws.freeze_panes = ws[f'A{row}']

    return last_fixed_col, row

#------------------------------------------------------------------

def in_ministry(member, ministry_name, log):
    # Is the Member active in the specified ministry?
    key = 'active_ministries'
    if key not in member:
        return false

    found = False
    for member_ministry in member[key]:
        if member_ministry['Description'] == ministry_name:
            log.debug(f"Found {member['Name']} in {ministry_name}")
            return True

    return False

###################################################################
###################################################################
###################################################################

def find_about_to_expire(pds_members, expiration, log):
    key  = 'calculated'
    key2 = 'by_result'
    key3 = 'Yes'

    check_dt     = now + expiration
    check_dt_str = check_dt.isoformat()
    now_str      = now.isoformat()

    out = list()
    for member in pds_members.values():
        if key not in member:
            continue
        if key2 not in member[key]:
            continue
        if key3 not in member[key][key2]:
            continue

        # NOTE: The code below could likely be more optimial /
        # efficient.  However, there's a bunch of different corner
        # cases, and I thought it would be best to write the code in
        # its simplest form -- even if it's not efficient.

        # Remember: the list of mandates has been consolidated/merged,
        # so each entry in "mandates" may actually represent more than
        # one mandate.  But that doesn't matter here: we just want to
        # know the length of time that this Member is covered by a
        # mandate (or set of mandates).
        mandates = member[key][key2][key3]

        # If today's date and the checked expiration date are both in
        # the same mandate range, then this Member is good -- move on
        # to check the next Member.
        covered = False
        for mandate in mandates:
            if (now_str in mandate['date_range'] and
                check_dt_str in mandate['date_range']):
                covered = True
                break

        if covered:
            continue

        # If we got here, it's because there's a discontinuity of
        # mandate coverage between now and the checked expiration
        # date.  Let's see if we can come up with a helpful message
        # for the report.

        # Calculate some values

        # The mandate that covers today
        now_mandate      = None
        # The next mandate that will be active after the mandate that
        # covers today
        next_mandate     = None

        for i, mandate in enumerate(mandates):
            if now_str in mandate['date_range']:
                now_mandate = mandate
                if i + 1 < len(mandates):
                    next_mandate = mandates[i + 1]

        log.debug(f"{member['Name']}: now {now_mandate}")
        log.debug(f"{member['Name']}: next {next_mandate}")

        # This report is only for finding people who have a current
        # mandate that is about to expire.  So skip any Member who
        # does not have a "now" mandate.
        if not now_mandate:
            continue

        # If we're here, it means thart there is a Member who has a
        # current mandate, but it will expire soon.  Make a
        # friendly/useful message
        expire_date = now_mandate["end_date"]
        message = f'Currently active mandate expires {now_mandate["end_date"]}'
        if next_mandate:
            message += f'; next mandate starts {next_mandate["start_date"]}'

        out.append({
            'member'      : member,
            'message'     : message,
            'expire date' : expire_date,
        })

    out.sort(key=lambda item: item['member']['Name'])
    return out

#------------------------------------------------------------------

def make_xlsx_about_to_expire(entries, log):
    # Make a new sheet
    wb = Workbook()
    ws = wb.active

    # Fill in the headers
    columns = [ ('Member Name',          20),
                ('Envelope number',      10),
                ('Email Address',        30),
                ('Phone Number',         20),
                ('Active in ministries', 10),
                ('Expire date',          20),
                ('Notes',                40), ]
    title = f'Communion Ministers with mandate about to expire'
    last_fixed_col, data_start_row = xlsx_create_headers(ws, title,
                                                         columns,
                                                         None, log)

    row = data_start_row
    for entry in entries:
        member = entry['member']
        email = PDSChurch.find_any_email(member)
        if email:
            email = email[0]
        else:
            email = ''

        # Check cell and home phone numbers
        phone = PDSChurch.find_member_phone(member, 'Cell')
        if not phone:
            phone = PDSChurch.find_member_phone(member, 'Home')

        active_in_ministry = 'No'
        for ministry in ministries:
            if in_ministry(member, ministry['name'], log):
                active_in_ministry = 'Yes'

        # Fill in the fixed columns
        ws[f'A{row}'] = member['email_name']
        ws[f'B{row}'] = member['family']['ParKey']
        ws[f'C{row}'] = email
        ws[f'D{row}'] = phone
        ws[f'E{row}'] = active_in_ministry
        ws[f'F{row}'] = entry['expire date']
        ws[f'G{row}'] = entry['message']

        row += 1

    return wb

#------------------------------------------------------------------

def mandate_about_to_expire_report(pds_members, expiration,
                                   google, args, log):
    filtered = find_about_to_expire(pds_members, expiration, log)
    log.info(f"Found {len(filtered)} Communion ministers with mandates that expire before {expiration}")

    wb = make_xlsx_about_to_expire(filtered, log)

    filename = 'Communion ministers with mandates about to expire'
    upload_workbook(filename, wb, args, google,
                    about_to_expire_gsheet_id, log)

###################################################################
###################################################################
###################################################################

def available_filter(member, ministry_name, log):
    # If we got here, the Member is in the desired ministry.  So we
    # can just return their training status.
    # (this key should definitely be there, but it never hurts to code
    # defensively)
    key = 'calculated'
    if key not in member:
        return False
    return member[key]['active']

def union_filter(member, ministry_name, log):
    # This filter is used for the "union" report -- i.e., the union of
    # all active Members who have an active mandate.
    #
    # If we got here, the Member is alredy in at least one of the
    # desired ministries.  So we can just return their training
    # status.
    key = 'calculated'
    if key not in member:
        return False
    return member[key]['active']

def available_column_names(log):
    cols = [ ('Current mandate ends' , 15),
             ('Notes' , 25), ]
    return cols

def available_column_data(member, first_col, log):
    key = 'calculated'
    key2 = 'by_result'
    mandates = member[key][key2]['Yes']

    # Find the mandate that we're in right now
    # (remember: this list of mandates was merged)
    now_str = now.isoformat()

    for mandate in mandates:
        if now_str in mandate['date_range']:
            values = [ mandate['end_date'] ]
            note = ''
            if mandate['end_date'] < today + about_to_expire:
                note = f'Expires in the next {int(about_to_expire.total_seconds() / 24 / 3600)} days'
            values.append(note)
            return values

    return [ '', '' ]

#------------------------------------------------------------------

def unavailable_filter(member, ministry_name, log):
    # If we got here, the Member is in the desired ministry.  So we
    # can just return their training status.
    key = 'calculated'
    if key not in member:
        return True
    return not member[key]['active']

def unavailable_column_names(log):
    cols = [ ('Notes' , 60), ]
    return cols

def unavailable_column_data(member, first_col, log):
    key = 'calculated'
    key2 = 'by_result'
    key3 = 'Yes'

    # It's possible that the Member has no mandates or no "Yes"
    # mandates at all.
    if (key not in member or
        key2 not in member[key] or
        key3 not in member[key][key2]):
        return [ 'No mandates listed in PDS' ]

    mandates = member[key][key2]['Yes']

    # We know that "now" is not in any existing mandate on this
    # Member.

    # Find the last active mandate before today.
    prev = None
    for mandate in mandates:
        if mandate['end_date'] < today:
            prev = mandate
        else:
            break

    # Find any active mandate after this one.
    next = None
    for mandate in mandates:
        if today < mandate['start_date']:
            next = mandate
        else:
            break

    if prev == None and next == None:
        message = 'No active mandates'
    elif prev == None and next != None:
        message = f'Mandate starts {next["start_date"]}'
    elif prev != None and next == None:
        message = f'Last mandate ended {prev["end_date"]}'
    else:
        message = f'Last mandate ended {prev["end_date"]}, but next mandate starts {next["start_date"]}'

    return [ message ]

#------------------------------------------------------------------

# Create a roster of everyone who has a current mandate for a given
# ministry
def make_xlsx_roster_workbook(members, ministry_name,
                              wb, sheet_name,
                              filter_fn,
                              column_name_fn,
                              column_data_fn, log):
    key = 'calculated'

    filtered = list()
    for member in members.values():
        # Filter Members who are not in this ministry
        if (ministry_name is not None and
            not in_ministry(member, ministry_name, log)):
            continue

        log.debug(f"Found {ministry_name} member: {member['Name']}")

        # Filter Members who do not match our criteria
        matched = filter_fn(member, ministry_name, log)
        if not matched:
            log.debug(f"DID NOT MATCH {sheet_name}: {ministry_name} {member['Name']}")
            continue

        filtered.append(member)

    filtered.sort(key=lambda item: item['Name'])

    log.info(f"Found {len(filtered)} members match {sheet_name} criteria in {ministry_name}")

    # Make a new sheet
    ws = wb.create_sheet(sheet_name)

    # Fill in the headers
    fixed_columns = [ ('Member Name',     20),
                      ('Envelope number', 10),
                      ('Email Address',   30),
                      ('Phone Number',    20), ]
    title = 'Communion Ministers with active mandate'
    if ministry_name is not None:
        title += f': {ministry_name}'
    last_fixed_col, data_start_row = xlsx_create_headers(ws, title,
                                                         fixed_columns,
                                                         column_name_fn, log)

    # Fill in the data
    first_col = chr(ord(last_fixed_col) + 1)
    row = data_start_row
    for member in filtered:
        email = PDSChurch.find_any_email(member)
        if email:
            email = email[0]
        else:
            email = ''

        # Check cell and home phone numbers
        phone = PDSChurch.find_member_phone(member, 'Cell')
        if not phone:
            phone = PDSChurch.find_member_phone(member, 'Home')

        # Fill in the fixed columns
        ws[f'A{row}'] = member['email_name']
        ws[f'B{row}'] = member['family']['ParKey']
        ws[f'C{row}'] = email
        ws[f'D{row}'] = phone

        # Fill in the dynamic columns
        if column_data_fn:
            data = column_data_fn(member, first_col, log)

            col = first_col
            for item in data:
                ws[f'{col}{row}'] = item
                col = chr(ord(col) + 1)

        row += 1

#------------------------------------------------------------------

# Make a sheet with all Active Members who have a current mandate,
# regardless of their participation in any ministry.
def report(pds_members, training_name,
           ministry_name, gsheet_id,
           google, args, log):
        wb = Workbook()
        make_xlsx_roster_workbook(pds_members, None,
                                  wb, "Any",
                                  union_filter,
                                  available_column_names,
                                  available_column_data,
                                  log)

        # Delete the default sheet that was automatically created
        key = 'Sheet'
        if key in wb.sheetnames:
            del wb[key]

        filename = 'All Active ECC parishioners with a current Communion Minister mandate',
        upload_workbook(filename, wb, args, google,
                        gsheet_id,
                        log)

###################################################################
###################################################################
###################################################################

def find_pds_training(pds_members, training_name, log):
    reqcount = 0
    one_day  = timedelta(days=1)

    log.info(f"Looking for certifications of type: {training_name}")
    for member in pds_members.values():
        key = 'requirements'
        if key not in member:
            continue

        for req in member[key]:
            if req['description'] != training_name:
                continue

            start_date = req['start_date']
            end_date   = req['end_date']

            if start_date == PDSChurch.date_never:
                # This is an invalid entry -- skip it
                log.warning(f"Skipping certification entry with no start date on Member: {member['Name']}")
                continue

            reqcount += 1

            key2 = 'calculated'
            if key2 not in member:
                member[key2] = {
                    'active'        : False,
                    'raw_trainings' : list(),
                    'by_result'     : dict(),
                }

            # This Member has training of the correct type.  See if it
            # is active.
            if (req['result'] == 'Yes' and
                today >= start_date and
                (end_date == PDSChurch.date_never or
                 today <= end_date)):
                member[key2]['active'] = True
                log.debug(f"Found active mandate for {member['Name']}")

            # Save a cross reference to the training record
            member[key2]['raw_trainings'].append(req)

            # Make a trivial lists by result
            key3   = 'by_result'
            result = req['result']
            if result not in member[key2][key3]:
                member[key2][key3][result] = list()
            member[key2][key3][result].append(req)

        # If we didn't find anything relevant on this member, move on
        # to the next one
        if key2 not in member:
            continue
        if key3 not in member[key2]:
            continue

        # Sort all the results by start date
        results = member[key2][key3]
        for trainings in results.values():
            trainings.sort(key=lambda item: item['start_date'])

        # Merge all the found "Yes" (i.e., fully mandated) trainings
        # that either are contiguous or overlap into a minimum number
        key4 = 'Yes'
        if key4 not in results:
            continue

        # Use a simple algorithm: look at two items on the list.  If
        # they can be merged, do so.  Repeat.
        i = 0
        results_yes = results['Yes']
        debug_print = len(results_yes) > 1
        while i + 1 < len(results_yes):
            left  = results_yes[i]
            right = results_yes[i + 1]

            should_merge = False

            # If the two items are contiguous, merge
            if left['end_date'] + one_day == right['start_date']:
                should_merge = True
            # If the two items overlap, merge
            if left['date_range'].is_intersection(right['date_range']):
                should_merge = True

            if should_merge:
                left['end_date'] = right['end_date']
                left_old_dtr     = left['date_range']
                right_old_dtr    = right['date_range']
                dtr = DateTimeRange(left_old_dtr.get_start_time_str(),
                                    right_old_dtr.get_end_time_str())
                left['date_range'] = dtr

                # Delete the right item
                del results_yes[i + 1]
            else:
                i += 1

        if debug_print:
            log.debug(member[key2])

    if reqcount == 0:
        log.info(f"No valid trainings of type: {training_name} found")
    else:
        log.info(f"Found {reqcount} training records")

###################################################################
###################################################################
###################################################################

def setup_cli_args():
    tools.argparser.add_argument('--logfile',
                                 help='Also save to a logfile')

    tools.argparser.add_argument('--debug',
                                 action='store_true',
                                 default=False,
                                 help='Be extra verbose')

    tools.argparser.add_argument('--dry-run',
                                 action='store_true',
                                 default=False,
                                 help='Do not upload to Google')

    tools.argparser.add_argument('--sqlite3-db',
                                 required=True,
                                 help='Location of PDS sqlite3 database')

    global gapp_id
    tools.argparser.add_argument('--app-id',
                                 default=default_gapp_id,
                                 help='Filename containing Google application credentials')
    global guser_cred_file
    tools.argparser.add_argument('--user-credentials',
                                 default=default_guser_cred_file,
                                 help='Filename containing Google user credentials')

    args = tools.argparser.parse_args()

    return args

#-------------------------------------------------------------------

def main():
    args = setup_cli_args()
    log = ECC.setup_logging(info=True,
                            debug=args.debug,
                            logfile=args.logfile)

    log.info("Reading PDS data...")
    (pds, pds_families,
     pds_members) = PDSChurch.load_families_and_members(filename=args.sqlite3_db,
                                                        parishioners_only=False,
                                                        log=log)

    apis = {
        'drive' : { 'scope'       : Google.scopes['drive'],
                    'api_name'    : 'drive',
                    'api_version' : 'v3', },
    }
    google = None
    if not args.dry_run:
        services = GoogleAuth.service_oauth_login(apis,
                                                app_json=args.app_id,
                                                user_json=args.user_credentials,
                                                log=log)
        google = services['drive']

    # Setup metadata on PDS Member records
    find_pds_training(pds_members, training_name, log)

    # All who have active mandates
    report(pds_members, training_name,
           name_313, gsheet_id_313,
           google, args, log)

    # Write "everyone who is about to expire" report
    mandate_about_to_expire_report(pds_members, about_to_expire,
                                   google, args, log)

    # All done
    pds.connection.close()

main()
