#!/usr/bin/env python3

#
# See https://openpyxl.readthedocs.io/en/stable/index.html
#
# pip3.6 install openpyxl
#

import sys
import os

import logging.handlers
import logging

# We assume that there is a "ecc-python-modules" sym link in this
# directory that points to the directory with ECC.py and friends.
moddir = os.path.join(os.getcwd(), 'ecc-python-modules')
if not os.path.exists(moddir):
    print("ERROR: Could not find the ecc-python-modules directory.")
    print("ERROR: Please make a ecc-python-modules sym link and run again.")
    exit(1)

sys.path.insert(0, moddir)

import ECC
import Google
import PDSChurch
import GoogleAuth
import googleapiclient
from google.api_core import retry

from datetime import datetime
from datetime import timedelta

from oauth2client import tools
from googleapiclient.http import MediaFileUpload

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment

from pprint import pprint
from pprint import pformat

# Globals

gapp_id         = 'client_id.json'
guser_cred_file = 'user-credentials.json'

ministries = [
    {
        'ministry' : '100-Parish Pastoral Council',
        'gsheet_id' : '1aIoStpSOsup8XL5eNd8nhpJwM-IqN2gTkwVf_Qvlylc',
        'birthday' : False,
    },
    {
        'ministry' : '102-Finance Advisory Council',
        'gsheet_id' : '1oGkjyLDexQyb-z53n2luFpE9vU7Gxv0rX6XirtxSjA0',
        'birthday' : False,
    },
    {
        'ministry' : '103-Worship Committee',
        'gsheet_id' : '1h_ZvhkYlnebIu0Tk7h1ldJo-VKnJsJGe1jEzY34mcd0',
        'birthday' : False,
    },
    {
        'ministry' : '104-Stewardship & E Committee',
        'gsheet_id' : '1avaHrl-sHOWOc541GclHZL3ajG30HozISR0HDoswKnM',
        'birthday' : False,
    },
    {
        'ministry' : '107-Social Resp Steering Comm',
        'gsheet_id' : '1Am3v0Pv4D9zubkGYgFbUd8e92PZnBPrbcwKrMrs8AnI',
        'birthday' : False,
    },
    {
        'ministry' : '108-Formation Team',
        'gsheet_id' : '1rWy81dH5ZSdjfXGyFN6nF9DIkVk1NS5G7RmHZDn5FTY',
        'birthday' : False,
    },
    {
        'ministry' : '110-Ten Percent Committee',
        'gsheet_id' : '18BIrnBWf_4LS9XeC9tordSD1SBgJz67a0I9Ouj6ZcEc',
        'birthday' : False,
    },
    {
        'ministry' : '111-Hispanic Ministry Team',
        'gsheet_id' : '1FuUzKKmezkfBcBJdGiX1gbw4PHmqR_nivwAjoFTLjK8',
        'birthday' : False,
    },
    {
        'ministry' : '113-Media Comms Planning Comm.',
        'gsheet_id' : '1erlqkK_F-ol8Bmo5Pg6SoAe355tOxwU3D4pU179foCU',
        'birthday' : False,
    },
    {
        'ministry' : '114-Long Range Plan 2029 Comm',
        'gsheet_id' : '1Ohroqv1OSZJuhJ0y4z2dIEt4gCkxdkyudAqZ0XpuPTQ',
        'birthday' : False,
    },
    {
        'ministry' : '115-Parish Life Committee',
        'gsheet_id' : '1bKiNztcE20jrK24ccoCzOJDuJsA-nAR-P-6BYstIDog',
        'birthday' : False,
    },
    {
        'ministry' : '116-Youth Council',
        'gsheet_id' : '1dM3gNsnTJyQDLeRSmP1iRhjWlXV0aHiOCmdLtG3KnKc',
        'birthday' : False,
    },
    {
        'ministry' : '200-Audit Committee',
        'gsheet_id' : '1OJ3z9tS9qoZMjr4EfSh5zeKy_PKmaRja9htIthjUPpI',
        'birthday' : False,
    },
    {
        'ministry' : '201-Collection Counter',
        'gsheet_id' : '1w-PQ-_jnzWWRB3id-Mh4pNHTHJzf9Al1gc_p7cS0koM',
        'birthday' : False,
    },
    {
        'ministry' : '202-Facility Mgmt & Planning',
        'gsheet_id' : '1kcna05U9Oiy0dyvinjOjLi8tXSwQ3sHBC8N8LqjDsjU',
        'birthday' : False,
    },
    {
        'ministry' : '203-Garden & Grounds',
        'gsheet_id' : '1ZrJCRCrrClQuEq9KPXiRVj2s2u78L6R4-P7hbMpYqzs',
        'birthday' : False,
    },
    {
        'ministry' : '204-Parish Office Volunteers',
        'gsheet_id' : '1yhGeBNaUTuVkEGTRkCh5SC6hQLOpdaEVNpXaq2WmMc4',
        'birthday' : False,
    },
    {
        'ministry' : '205-Participation Sheet Vol',
        'gsheet_id' : '1rPf-_YTqZoOxhj_EO7a3i4tA0-1wl-OxvLyk82m6IE0',
        'birthday' : False,
    },
    {
        'ministry' : '206-Space Arrangers',
        'gsheet_id' : '1ZIAn7MdeDVcgvHwR98dXs_ewIaU2uWbl6mvriLSk5V0',
        'birthday' : False,
    },
    {
        'ministry' : '207-Technology Committee',
        'gsheet_id' : '1Gn2m2VMabPkWJWg_NTs6XeGPf_Qi7qPELLxyOx9Q0vU',
        'birthday' : False,
    },
    {
        'ministry' : '208-Weekend Closer',
        'gsheet_id' : '1pj2r_0Xjog22g-kiSmSGkO6BHIIZqW7uXcE3hUs6Xy8',
        'birthday' : False,
    },
    {
        'ministry' : '300-Art & Environment',
        'gsheet_id' : '1uXy2zTQeeH_YtBAR46lDV5X8ADZRgPnNROtzDU_iqv0',
        'birthday' : False,
    },
    {
        'ministry' : '301-Audio/Visual/Light Minstry',
        'gsheet_id' : '1LtsNJc-9KYZkQqy2BITQ4Xgd_ns4Uo7z3YSQdnQrzN8',
        'birthday' : False,
    },
    {
        'ministry' : '303-Linens/Vestments Ministry',
        'gsheet_id' : '1ul5doCFZx4Y_L8ZNmDarVdgLy_dQo2MkE0mdo5kbrY0',
        'birthday' : False,
    },
    {
        'ministry' : '304-Liturgical Planning',
        'gsheet_id' : '11A59wLOy58-ADKm60YuVGDG0N02E8a9iY-VoCjQX9tY',
        'birthday' : False,
    },
    {
        'ministry' : '305-Movers Ministry',
        'gsheet_id' : '1q0rNZQ0Od3cCoFG2qLgKZep10GzQ3nmvVIIJuUOWBgY',
        'birthday' : False,
    },
    {
        'ministry' : '306-Music Support for Children',
        'gsheet_id' : '1ODBCOAUY5g93-ZDrgAx9-Fv5wT4VSeMYxHgpFeb2F4s',
        'birthday' : False,
    },
    {
        'ministry' : '307-Wedding Assistant',
        'gsheet_id' : '1UBzQOEpdYmb1afAnrPDlQcz0hoouvd6mkpEV6rYDMxg',
        'birthday' : False,
    },
    {
        'ministry' : '308-Worship&Music Support Team',
        'gsheet_id' : '1kHRnohAEeaC_2-yTTcgHe36yRVJ4EhWhmbBCZ0P_vI8',
        'birthday' : False,
    },
    {
        'ministry' : '309-Acolytes',
        'gsheet_id' : '1zXfAxnuQCATWQ7nU7QYn2Mvy1Ql1Huy7Ve1DW60nPKA',
        'birthday' : False,
    },
    {
        'ministry' : '310-Adult Choir',
        'gsheet_id' : '1ku8Aq9dXm_mrOq421MWVk7hAqV2Am5FFSgUACOYs2WU',
        'birthday' : False,
    },
    {
        'ministry' : '311-Bell Choir',
        'gsheet_id' : '1UTzXgO9ZLBHB0w-zAW8-u57cgWLbkWWGanJgPC9gboE',
        'birthday' : True,
    },
    {
        'ministry' : '312-Children\'s Music Ministry',
        'gsheet_id' : '11H5SbDti8Jm2HMqdo3mZ3_5iTBky4OmZGdB_WeCOZYI',
        'birthday' : False,
    },
    {
        'ministry' : '313-Communion Ministers',
        'gsheet_id' : '10Aq9XtZHL3v0m0Erm71f8BSUYr7CiEDMwzDWTSgKPJ4',
        'birthday' : False,
    },
    {
        'ministry' : '315-Funeral Mass Ministry',
        'gsheet_id' : '107rbGV_WbEz6m6m8XMliLWIMFKKMz8pQv8pdjgLFvLI',
        'birthday' : False,
    },
    {
        'ministry' : '316-Greeters',
        'gsheet_id' : '1vd7sltcU8MVwIBad__PyXXMN_cUEFvdXpL5RkYiBWJ0',
        'birthday' : False,
    },
    {
        'ministry' : '317-Instrumentalists & Cantors',
        'gsheet_id' : '1YP3sC4dcOWH9Li1rJV8D5FI9mef50xvxqOf6K1K54_U',
        'birthday' : True,
    },
    {
        'ministry' : '318-Lectors',
        'gsheet_id' : '1X796X7_wFZmYoKMzGnj2BFFCOeoncIEILv1cmq_CJB8',
        'birthday' : False,
    },
    {
        'ministry' : '319-Liturgical Dance Ministry',
        'gsheet_id' : '14s4gtJMjK0qiHvN-eQ1zaD76PoSGVd9qcfIPxTDGhrc',
        'birthday' : False,
    },
    {
        'ministry' : '321-Prayer Chain Ministry',
        'gsheet_id' : '1Z3amqu2CwsqUNw3O7ih_awgRwyXJazBzAYt93ZgoAwQ',
        'birthday' : False,
    },
    {
        'ministry' : '401-Epiphany Companions',
        'gsheet_id' : '1voFdTbY3RMs3R_X-pO6-hbjBfC9Prm80ON-lLDTl2UI',
        'birthday' : False,
    },
    {
        'ministry' : '402-New Members Coffee',
        'gsheet_id' : '10yQUVTr2EHqRwvdotQAZT33mbVwk8jtDoIwIyWW3-to',
        'birthday' : False,
    },
    {
        'ministry' : '404-Welcome Desk',
        'gsheet_id' : '1pJwD2UXiDAKng-DFwdWMp0AbZQFSJoiOsJaQgJmfvCE',
        'birthday' : False,
    },
    {
        'ministry' : '405-Parish Mission Plan Team',
        'gsheet_id' : '1YLK4dsW2-whQeKPR0g4F8q3ZuximhwLiN1gS_ps4JgQ',
        'birthday' : False,
    },
    {
        'ministry' : '406-Evangelization Team',
        'gsheet_id' : '1T2FCBImLj4akVnX1hpfymi-JKV5jPf3G0ucknzQz8Io',
        'birthday' : False,
    },
    {
        'ministry' : '407-Stewardship Team',
        'gsheet_id' : '1vGOdc2YtiNj1VnzUMy0ftrjuwVxqSJz_8XFeTxnQ8l8',
        'birthday' : False,
    },
    {
        'ministry' : '408-Engagement Team',
        'gsheet_id' : '1zaVGPdnHp5zZu3zbWfGckIaxbCeCSSTYyoe4pCIlQjs',
        'birthday' : False,
    },
    {
        'ministry' : '409-Sunday Morning Coffee',
        'gsheet_id' : '1QLcjLidJppCfhqaNBn10h1Toe81DEy48PwwJAhwDNUs',
        'birthday' : False,
    },
    {
        'ministry' : '452-Media Communications',
        'gsheet_id' : '1KkCF2V4JIK65b9QsfSG-lGMCASXPavxIxim-CfzxBYk',
        'birthday' : False,
    },
    {
        'ministry' : '501-Eucharist to Sick&Homebnd',
        'gsheet_id' : '1fKXe-NuoObjgYoFRcbuJQwIOjFIAfT7TQB87eccgfng',
        'birthday' : False,
    },
    {
        'ministry' : '504-DivorceCare',
        'gsheet_id' : '1c5S-NSgGrGipzk2Wri60pazwcR5L_ZjWQc_Gzxqc-1g',
        'birthday' : False,
    },
    {
        'ministry' : '505-Healing Blanket Ministry',
        'gsheet_id' : '1NOijo2AHwftLBGDyCiyGMEIp4Qmfw-0jVjHG9Ygz3S8',
        'birthday' : False,
    },
    {
        'ministry' : '508-Messages of Hope Ministry',
        'gsheet_id' : '1Ua7pm-4Av7vPhU8quL5sAEKfowo8O_jfEF34vD1s3Wg',
        'birthday' : False,
    },
    {
        'ministry' : '509-HOPE Support Groups',
        'gsheet_id' : '160CPnzR_Q-kph_36LLCPFMC_pLRdMFbUQ2fkQ90LEw0',
        'birthday' : False,
    },
    {
        'ministry' : '510-Flower Delivery to SHB',
        'gsheet_id' : '1Fq11QLU7OQ_yHYPeDPXT8JaA31ackl3eJ2xOvecG_go',
        'birthday' : False,
    },
    {
        'ministry' : '600-Men of Epiphany',
        'gsheet_id' : '11LCDr-Vc3jyeKh5nrd49irscdvTv3TDXhpOoFWlohgs',
        'birthday' : False,
    },
    {
        'ministry' : '601-Sages (for 50 yrs. +)',
        'gsheet_id' : '1-uvQO5RRf0K6NJlR_4Mijygn4XGk0zhvowdflKLoEUc',
        'birthday' : False,
    },
    {
        'ministry' : '602-Singles Explore Life (SEL)',
        'gsheet_id' : '1dJ-57kiTO1SsvxG8wAXd2YlQ_vY7xmKnjvI9j0k97EM',
        'birthday' : False,
    },
    {
        'ministry' : '604-Wednesdays for Women',
        'gsheet_id' : '19n3zSL9bHla98_Q6d29kcp1d2lxnSpYOU6FUtU41lf4',
        'birthday' : False,
    },
    {
        'ministry' : '609-Octoberfest Plan Team 2022',
        'gsheet_id' : '17QXoqgreLu8sUQpZfooNkto4NV06wdoE3yA6Q5vgiAw',
        'birthday' : False,
    },
    {
        'ministry' : '610-FeastOfEpiphanyPlanTeam\'23',
        'gsheet_id' : '18UNXrRqckFSO2801lmODRohk_CJVI9ko4UtRFruS9E0',
        'birthday' : False,
    },
    {
        'ministry' : '611-Bereavement Receptions',
        'gsheet_id' : '1r6FNXGn-T5anj9X7HGlB2EpZ9oXfy5dQXD6z-OQEGgE',
        'birthday' : False,
    },
    {
        'ministry' : '612-Community Life Committee',
        'gsheet_id' : '1k_hH1tEWBGuERCmFvhZxKOfAsBkqy0uZ16LAd0_jMDg',
        'birthday' : False,
    },
    {
        'ministry' : '700-Advocates for Common Good',
        'gsheet_id' : '1Iz8hz7NAhh9-dVMiC7mL8yYFi_qmM_ayB5IXhJU0uPw',
        'birthday' : False,
    },
    {
        'ministry' : '701-CLOUT',
        'gsheet_id' : '1gBQXnTgxodILkjXvBrfJ-sztVj6NZaZQI1kbddAESNE',
        'birthday' : False,
    },
    {
        'ministry' : '703-Eyeglass Ministry',
        'gsheet_id' : '1i3EBKO3Lj3lIprhnoc4VeabEIPAuce5JeV99y1zrUAc',
        'birthday' : False,
    },
    {
        'ministry' : '704-Habitat for Humanity',
        'gsheet_id' : '1HacDJsMK-oLKjuPjvrABhbYn-b0joox3GMY9uEA--yg',
        'birthday' : False,
    },
    {
        'ministry' : '705-Hunger & Poverty Ministry',
        'gsheet_id' : '1m0Cp7k0XyeJvZ8Z0IZXrudD3oALfeLr0sWvvnEEgn0o',
        'birthday' : False,
    },
    {
        'ministry' : '706-Prison Ministry',
        'gsheet_id' : '1lXkeHsyHNqHYH4zs_HiQdaq8PK8vh5W2koQE5sCq_7U',
        'birthday' : False,
    },
    {
        'ministry' : '707-St. Vincent De Paul',
        'gsheet_id' : '1tFmzvu53v-oBF3P59fl4zv46VHt4PRu066YyceuYNl4',
        'birthday' : False,
    },
    {
        'ministry' : '709-Twinning Committee:Chiapas',
        'gsheet_id' : '1GVKj83EQxNS6TFI1Coe9Kat2cyezcKwD8VihuVEWLMA',
        'birthday' : False,
    },
    {
        'ministry' : '710-Environmental Concerns',
        'gsheet_id' : '1jsoRxugVwXi_T2IDq9J-mEVdzS8xaOk9kuXGAef-YaQ',
        'birthday' : False,
    },
    {
        'ministry' : '712-Legislative Network',
        'gsheet_id' : '1SqqnAFEZlUdCAxp6NYKqpqGFy13L_aheVEGzHLLzOuk',
        'birthday' : False,
    },
    {
        'ministry' : '800-Catechists for Children',
        'gsheet_id' : '1rBKiweOBT-JZFbfv5Sf4_vjjOlAYnOGaZIhadtBPxSE',
        'birthday' : False,
    },
    {
        'ministry' : '802-Gather the Children',
        'gsheet_id' : '1lI9eO0bryD6GFIsgJ5FipGePK4FaJ7v8eIp71DFz_6A',
        'birthday' : False,
    },
    {
        'ministry' : '805-Monday Adult Bible Study',
        'gsheet_id' : '1gANQt7PBc8erErQRHrRLtkdmOF3Y7DbuLC3J6aRLu4U',
        'birthday' : False,
    },
    {
        'ministry' : '807-Catechumenate/InitiationTm',
        'gsheet_id' : '13neLXFrDTsohe_N_CVPX7ajx4pFvYb9miFtZtUqdVl4',
        'birthday' : False,
    },
]

keywords = [
    {
        'keyword'   : 'Livestream Team',
        'gsheet_id' : '1Yku0IFuIKZCeUNGB5c_Ser_geYkylC2o1tiVfaNwkx8',
        'birthday'  : False,
    },
    {
        'keyword'   : 'YouthMin parent: Jr high',
        'gsheet_id' : '1VIs-AezopoWd3rpVU_kqMOTTXlBPpjI9IxAbhDpSnG4',
        'birthday'  : False,
    },
    {
        'keyword'   : 'YouthMin parent: Sr high',
        'gsheet_id' : '1B_2LeR0EbG3oDUqifvB51YhCuj30lhaYHfUb9bmCj5A',
        'birthday'  : False,
    },
]

####################################################################

def write_xlsx(members, ministry_name, name, want_birthday, log):
    # Make the microseconds be 0, just for simplicity
    now = datetime.now()
    us = timedelta(microseconds=now.microsecond)
    now = now - us

    timestamp = ('{year:04}-{mon:02}-{day:02} {hour:02}:{min:02}'
                .format(year=now.year, mon=now.month, day=now.day,
                        hour=now.hour, min=now.minute))
    filename_base = name
    if filename_base is None:
        filename_base = ministry_name
    filename_base = filename_base.replace("/", "-")
    filename = (f'{filename_base} members as of {timestamp}.xlsx')

    # Put the members in a sortable form (they're currently sorted by MID)
    sorted_members = dict()
    for m in members:
        # 'Name' will be "Last,First..."
        sorted_members[m['Name']] = m

    wb = Workbook()
    ws = wb.active

    # Title rows + set column widths
    title_font = Font(color='FFFF00')
    title_fill = PatternFill(fgColor='0000FF', fill_type='solid')
    title_align = Alignment(horizontal='center')

    last_col = 'C'
    if want_birthday:
        last_col = 'D'

    row = 1
    ws.merge_cells(f'A{row}:{last_col}{row}')
    cell = f'A{row}'
    ws[cell] = f'Ministry: {ministry_name}'
    ws[cell].fill = title_fill
    ws[cell].font = title_font

    row = row + 1
    ws.merge_cells(f'A{row}:{last_col}{row}')
    cell = f'A{row}'
    ws[cell] = f'Last updated: {now}'
    ws[cell].fill = title_fill
    ws[cell].font = title_font

    row = row + 1
    ws.merge_cells(f'A{row}:{last_col}{row}')
    cell = f'A{row}'
    ws[cell] = ''
    ws[cell].fill = title_fill
    ws[cell].font = title_font

    row = row + 1
    columns = [(f'A{row}', 'Member name', 30),
               (f'B{row}', 'Address', 30),
               (f'C{row}', 'Phone / email', 50)]
    if want_birthday:
        columns.append((f'D{row}', 'Birthday', 30))

    for cell,value,width in columns:
        ws[cell] = value
        ws[cell].fill = title_fill
        ws[cell].font = title_font
        ws[cell].alignment = title_align
        ws.column_dimensions[cell[0]].width = width

    # Freeze the title row
    row = row + 1
    ws.freeze_panes = ws[f'A{row}']

    #---------------------------------------------------------------------

    def _append(row, col, value):
        if value is None or len(value.strip()) == 0:
            return row

        _ = ws.cell(row=row, column=col, value=value)
        return row + 1

    # Data rows
    for name in sorted(sorted_members):
        m = sorted_members[name]
        # The name will take 1 row
        _ = ws.cell(row=row, column=1, value=m['email_name'])

        # The address will take multiple rows
        col = 2
        last_row = row
        f = m['family']
        last_row = _append(col=col, row=last_row, value=f['StreetAddress1'])
        last_row = _append(col=col, row=last_row, value=f['StreetAddress2'])
        val = '{cs}, {zip}'.format(cs=f['city_state'], zip=f['StreetZip'])
        last_row = _append(col=col, row=last_row, value=val)
        addr_last_row = last_row
        email_last_row = last_row

        # The phone / email may be more than 1 row
        col = 3
        last_row = row
        key = 'phones'
        if key in m:
            for phone in m[key]:
                # Skip unlisted phone numbers
                if phone['unlisted']:
                    log.info("SKIPPED UNLISTED NUMBER FOR {n}".format(n=m['full_name']))
                    continue

                val = '{ph} {type}'.format(ph=phone['number'], type=phone['type'])
                last_row = _append(col=col, row=last_row, value=val)

        # If we have any preferred emails, list them all
        key = 'preferred_emails'
        if key in m and len(m[key]) > 0:
            for email in m[key]:
                last_row = _append(col=col, row=last_row, value=email['EMailAddress'])
                email_last_row = last_row

        # If we have no preferred emails, list the first alphabetic
        # non-preferred email
        else:
            key = 'non_preferred_emails'
            if key in m and len(m[key]) > 0:
                emails   = sorted([x['EMailAddress'] for x in m[key]])
                last_row = _append(col=col, row=last_row,
                                   value=emails[0])
                email_last_row = last_row

        # The birthday will only be 1 row
        if want_birthday:
            col = 4
            key1 = 'MonthOfBirth'
            key2 = 'DayOfBirth'
            if key1 in m and key2 in m:
                birthday = '{m} {d}'.format(m=m[key1], d=m[key2])
                # Sometimes PDS has "None" in one of these two fields
                if 'None' not in birthday:
                    _append(col=col, row=row, value=birthday)

        # Between the address / phone+email, find the real last row
        last_row = max(email_last_row, addr_last_row)
        row = last_row + 1

    #---------------------------------------------------------------------

    wb.save(filename)
    log.info(f'Wrote {filename}')

    return filename

#-------------------------------------------------------------------

@retry.Retry(predicate=Google.retry_errors)
def upload_overwrite(filename, google, file_id, log):
    # Strip the trailing ".xlsx" off the Google Sheet name
    gsheet_name = filename
    if gsheet_name.endswith('.xlsx'):
        gsheet_name = gsheet_name[:-5]

    try:
        log.info('Uploading file update to Google file ID "{id}"'
              .format(id=file_id))
        metadata = {
            'name'     : gsheet_name,
            'mimeType' : Google.mime_types['sheet'],
            'supportsAllDrives' : True,
            }
        media = MediaFileUpload(filename,
                                mimetype=Google.mime_types['sheet'],
                                resumable=True)
        file = google.files().update(body=metadata,
                                     fileId=file_id,
                                     media_body=media,
                                     supportsAllDrives=True,
                                     fields='id').execute()
        log.debug('Successfully updated file: "{filename}" (ID: {id})'
              .format(filename=filename, id=file['id']))

    except Exception as e:
        # When errors occur, we do want to log them.  But we'll re-raise them to
        # let an upper-level error handler handle them (e.g., retry.Retry() may
        # actually re-invoke this function if it was a retry-able Google API
        # error).
        log.error('Google file update failed for some reason:')
        log.error(e)
        raise e

#-------------------------------------------------------------------

def _create_roster(pds_members, ministry_name, sheet_name,
                   birthday, gsheet_id, google, log):
    # PDSChurch.filter_members() returns a dict.  Turn this into a simple
    # list of Members.
    members = [ x for x in pds_members.values() ]

    # Make an xlsx
    filename = write_xlsx(members=members,
                          ministry_name=sheet_name,
                          name=ministry_name,
                          want_birthday=birthday, log=log)
    log.debug(f"Wrote temp XLSX file: {filename}")

    # Upload the xlsx to Google
    upload_overwrite(filename=filename, google=google, file_id=gsheet_id,
                     log=log)
    log.debug("Uploaded XLSX file to Google")

    # Remove the temp local XLSX file
    try:
        os.unlink(filename)
        log.debug("Unlinked temp XLSX file")
    except Exception as e:
        log.info("Failed to unlink temp XLSX file!")
        log.error(e)

#-------------------------------------------------------------------

def create_ministry_roster(pds_members, ministry_entry, google, log):
    gsheet_id     = ministry_entry['gsheet_id']
    birthday      = ministry_entry['birthday']

    key1 = 'ministry'
    key2 = 'ministries'
    if key1 in ministry_entry:
        sheet_name = ministry_entry[key1]
        ministries = [ sheet_name ]
    elif key2 in ministry_entry:
        sheet_name = ', '.join(ministry_entry[key2])
        ministries = ministry_entry[key2]
    else:
        print(f"ERROR: Cannot find {key1} or {key2} in ministry_entry!")
        print(ministry_entry)
        exit(1)

    name = None
    key = 'name'
    if key in ministry_entry:
        name = ministry_entry[key]

    # Find the members
    members = PDSChurch.filter_members_on_ministries(pds_members, ministries)
    if members is None or len(members) == 0:
        log.info(f"No members in ministry: {sheet_name} -- writing empty sheet")

    _create_roster(members, name, sheet_name,
                   birthday, gsheet_id, google, log)

#-------------------------------------------------------------------

def create_keyword_roster(pds_members, keyword_entry, google, log):
    gsheet_id = keyword_entry['gsheet_id']
    birthday  = keyword_entry['birthday']

    key = 'keyword'
    if key not in keyword_entry:
        print(f"ERROR: Cannot find {key} in ministry_entry!")
        print(ministry_entry)
        exit(1)

    sheet_name = keyword_entry[key]
    keywords = [ sheet_name,
                 f'{sheet_name} Ldr' ]

    name = None
    key = 'name'
    if key in keyword_entry:
        name = keyword_entry[key]

    # Find the members with the keyword or keyword+' Ldr'
    members = PDSChurch.filter_members_on_keywords(pds_members, keywords)
    if members is None or len(members) == 0:
        log.info(f"No members with keyword: {sheet_name} -- writing empty sheet")

    _create_roster(members, name, sheet_name,
                   birthday, gsheet_id, google, log)

####################################################################

def setup_cli_args():
    tools.argparser.add_argument('--logfile',
                                 help='Also save to a logfile')
    tools.argparser.add_argument('--debug',
                                 action='store_true',
                                 default=False,
                                 help='Be extra verbose')

    tools.argparser.add_argument('--sqlite3-db',
                                 required=True,
                                 help='Location of PDS sqlite3 database')

    global gapp_id
    tools.argparser.add_argument('--app-id',
                                 default=gapp_id,
                                 help='Filename containing Google application credentials')
    global guser_cred_file
    tools.argparser.add_argument('--user-credentials',
                                 default=guser_cred_file,
                                 help='Filename containing Google user credentials')

    args = tools.argparser.parse_args()

    return args

####################################################################
#
# Main
#
####################################################################

def main():
    args = setup_cli_args()

    log = ECC.setup_logging(info=True,
                            debug=args.debug,
                            logfile=args.logfile)

    log.info("Reading PDS data...")
    (pds, pds_families,
     pds_members) = PDSChurch.load_families_and_members(filename=args.sqlite3_db,
                                                        parishioners_only=False,
                                                        log=log)

    apis = {
        'drive' : { 'scope'       : Google.scopes['drive'],
                    'api_name'    : 'drive',
                    'api_version' : 'v3', },
    }
    services = GoogleAuth.service_oauth_login(apis,
                                              app_json=args.app_id,
                                              user_json=args.user_credentials,
                                              log=log)
    google = services['drive']

    for keyword in keywords:
        create_keyword_roster(pds_members=pds_members,
                              keyword_entry=keyword,
                              google=google,
                              log=log)
    for ministry in ministries:
        create_ministry_roster(pds_members=pds_members,
                               ministry_entry=ministry,
                               google=google,
                               log=log)

    # All done
    pds.connection.close()

if __name__ == '__main__':
    main()
