#!/usr/bin/env python3

import sys
import os
import csv

import logging.handlers
import logging

# We assume that there is a "ecc-python-modules" sym link in this
# directory that points to the directory with ECC.py and friends.
moddir = os.path.join(os.getcwd(), 'ecc-python-modules')
if not os.path.exists(moddir):
    print("ERROR: Could not find the ecc-python-modules directory.")
    print("ERROR: Please make a ecc-python-modules sym link and run again.")
    exit(1)

sys.path.insert(0, moddir)

import ECC
import Google
import PDSChurch
import GoogleAuth
import googleapiclient
from google.api_core import retry

from datetime import datetime
from datetime import timedelta

from oauth2client import tools
from googleapiclient.http import MediaFileUpload

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment

from pprint import pprint
from pprint import pformat

# Globals

gapp_id         = 'client_id.json'
guser_cred_file = 'user-credentials.json'

keywords = [
    {
        'keyword'   : 'Homebound',
        'gsheet_id' : '1lCeMoGuVuXJpgHAKOcGCPKYlW2iD60AtOfB7QdMTw54',
    },
]

####################################################################

def write_xlsx(families, keyword_name, name, log):
    # Make the microseconds be 0, just for simplicity
    now = datetime.now()
    us = timedelta(microseconds=now.microsecond)
    now = now - us

    timestamp = ('{year:04}-{mon:02}-{day:02} {hour:02}:{min:02}'
                .format(year=now.year, mon=now.month, day=now.day,
                        hour=now.hour, min=now.minute))
    filename_base = name
    if filename_base is None:
        filename_base = keyword_name
    filename_base = filename_base.replace("/", "-")
    filename = (f'{filename_base} members as of {timestamp}.xlsx')

    # Put the members in a sortable form (they're currently sorted by MID)
    families_by_name = { f['Name'] : f for f in families }

    #---------------------------------------------------------------------

    wb = Workbook()
    ws = wb.active

    # Title rows + set column widths
    title_font = Font(color='FFFF00')
    title_fill = PatternFill(fgColor='0000FF', fill_type='solid')
    title_align = Alignment(horizontal='center')

    last_col = 'I'

    row = 1
    ws.merge_cells(f'A{row}:{last_col}{row}')
    cell = f'A{row}'
    ws[cell] = f'Keyword: {keyword_name}'
    ws[cell].fill = title_fill
    ws[cell].font = title_font

    row = row + 1
    ws.merge_cells(f'A{row}:{last_col}{row}')
    cell = f'A{row}'
    ws[cell] = f'Last updated: {now}'
    ws[cell].fill = title_fill
    ws[cell].font = title_font

    row = row + 1
    ws.merge_cells(f'A{row}:{last_col}{row}')
    cell = f'A{row}'
    ws[cell] = ''
    ws[cell].fill = title_fill
    ws[cell].font = title_font

    row = row + 1
    columns = [(f'A{row}', 'Family name', 30),
               (f'B{row}', 'Street 1', 30),
               (f'C{row}', 'Street 2', 30),
               (f'D{row}', 'City', 20),
               (f'E{row}', 'State', 10),
               (f'F{row}', 'Zip', 15),
               (f'G{row}', 'Home phone', 30),
               (f'H{row}', 'HoH email', 50),
               (f'I{row}', 'Spouse email', 50),
               ]

    for cell,value,width in columns:
        ws[cell] = value
        ws[cell].fill = title_fill
        ws[cell].font = title_font
        ws[cell].alignment = title_align
        ws.column_dimensions[cell[0]].width = width

    # Freeze the title row
    row = row + 1
    ws.freeze_panes = ws[f'A{row}']

    #---------------------------------------------------------------------

    def _append(row, col, value):
        if value is None or len(value.strip()) == 0:
            return row

        _ = ws.cell(row=row, column=col, value=value)
        return row + 1

    # Data rows
    for family_name in sorted(families_by_name):
        family = families_by_name[family_name]

        hoh, spouse, _ = \
            PDSChurch.filter_members_on_hohspouse(family['members'])

        home_phone = \
            PDSChurch.find_member_phone(hoh, 'Home', True)
        if home_phone is None:
            home_phone = \
                PDSChurch.find_member_phone(spouse, 'Home', True)
        if home_phone is None:
            home_phone = \
                PDSChurch.find_member_phone(family, 'Home', True)

        hoh_emails = PDSChurch.find_any_email(hoh)
        hoh_email = ''
        if len(hoh_emails) > 0:
            hoh_email = hoh_emails[0]

        spouse_emails = PDSChurch.find_any_email(spouse)
        spouse_email = ''
        if len(spouse_emails) > 0:
            spouse_email = spouse_emails[0]

        _ = ws.cell(row=row, column=1, value=family['MailingName'])
        _ = ws.cell(row=row, column=2, value=family['StreetAddress1'])
        _ = ws.cell(row=row, column=3, value=family['StreetAddress2'])
        _ = ws.cell(row=row, column=4, value=family['city'])
        _ = ws.cell(row=row, column=5, value=family['state'])
        _ = ws.cell(row=row, column=6, value=family['StreetZip'])
        _ = ws.cell(row=row, column=7, value=home_phone)
        _ = ws.cell(row=row, column=8, value=hoh_email)
        _ = ws.cell(row=row, column=9, value=spouse_email)

        row += 1

    wb.save(filename)
    log.info(f'Wrote {filename}')

    return filename

#-------------------------------------------------------------------

@retry.Retry(predicate=Google.retry_errors)
def upload_overwrite(filename, google, file_id, log):
    # Strip the trailing ".xlsx" off the Google Sheet name
    gsheet_name = filename
    if gsheet_name.endswith('.xlsx'):
        gsheet_name = gsheet_name[:-5]

    try:
        log.info('Uploading file update to Google file ID "{id}"'
              .format(id=file_id))
        metadata = {
            'name'     : gsheet_name,
            'mimeType' : Google.mime_types['sheet'],
            'supportsAllDrives' : True,
            }
        media = MediaFileUpload(filename,
                                mimetype=Google.mime_types['sheet'],
                                resumable=True)
        file = google.files().update(body=metadata,
                                     fileId=file_id,
                                     media_body=media,
                                     supportsAllDrives=True,
                                     fields='id').execute()
        log.debug('Successfully updated file: "{filename}" (ID: {id})'
              .format(filename=filename, id=file['id']))

    except Exception as e:
        # When errors occur, we do want to log them.  But we'll re-raise them to
        # let an upper-level error handler handle them (e.g., retry.Retry() may
        # actually re-invoke this function if it was a retry-able Google API
        # error).
        log.error('Google file update failed for some reason:')
        log.error(e)
        raise e

#-------------------------------------------------------------------

def _create_roster(families, keyword_name, sheet_name,
                   gsheet_id, google, log):
    # Make an XLSX
    filename = write_xlsx(families=families,
                          keyword_name=sheet_name,
                          name=keyword_name,
                          log=log)
    log.debug(f"Wrote temp XLSX file: {filename}")

    # Upload the xlsx to Google
    upload_overwrite(filename=filename, google=google, file_id=gsheet_id,
                     log=log)
    log.debug("Uploaded XLSX file to Google")

    # Remove the temp local XLSX file
    try:
        os.unlink(filename)
        log.debug("Unlinked temp XLSX file")
    except Exception as e:
        log.info("Failed to unlink temp XLSX file!")
        log.error(e)

#-------------------------------------------------------------------

def create_keyword_roster(pds_families, keyword_entry, google, log):
    gsheet_id = keyword_entry['gsheet_id']

    key = 'keyword'
    if key not in keyword_entry:
        print(f"ERROR: Cannot find {key} in keyword_entry!")
        print(keyword_entry)
        exit(1)

    sheet_name = keyword_entry[key]
    keywords = [ sheet_name, ]

    name = None
    key = 'name'
    if key in keyword_entry:
        name = keyword_entry[key]

    key = 'keywords'
    families = list()
    for family in pds_families.values():
        found = False
        for member in family['members']:
            if key not in member:
                continue
            for target_key in keywords:
                if target_key in member[key]:
                    found = True
                    break

        if found:
            families.append(family)

    if len(families) == 0:
        log.info(f"No families with keyword: {sheet_name} -- writing empty sheet")

    _create_roster(families, name, sheet_name,
                   gsheet_id, google, log)

####################################################################

def setup_cli_args():
    tools.argparser.add_argument('--logfile',
                                 help='Also save to a logfile')
    tools.argparser.add_argument('--debug',
                                 action='store_true',
                                 default=False,
                                 help='Be extra verbose')

    tools.argparser.add_argument('--sqlite3-db',
                                 required=True,
                                 help='Location of PDS sqlite3 database')

    global gapp_id
    tools.argparser.add_argument('--app-id',
                                 default=gapp_id,
                                 help='Filename containing Google application credentials')
    global guser_cred_file
    tools.argparser.add_argument('--user-credentials',
                                 default=guser_cred_file,
                                 help='Filename containing Google user credentials')

    args = tools.argparser.parse_args()

    return args

####################################################################
#
# Main
#
####################################################################

def main():
    args = setup_cli_args()

    log = ECC.setup_logging(info=True,
                            debug=args.debug,
                            logfile=args.logfile)

    log.info("Reading PDS data...")
    (pds, pds_families, pds_members) = \
        PDSChurch.load_families_and_members(filename=args.sqlite3_db,
                                            parishioners_only=True,
                                            log=log)

    apis = {
        'drive' : { 'scope'       : Google.scopes['drive'],
                    'api_name'    : 'drive',
                    'api_version' : 'v3', },
    }
    services = \
        GoogleAuth.service_oauth_login(apis,
                                       app_json=args.app_id,
                                       user_json=args.user_credentials,
                                       log=log)
    google = services['drive']

    for entry in keywords:
        create_keyword_roster(pds_families=pds_families,
                              keyword_entry=entry,
                              google=google,
                              log=log)

    # All done
    pds.connection.close()

if __name__ == '__main__':
    main()
