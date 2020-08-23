#!/usr/bin/env python3

# Class information for the different sheets of create-training-rosters.py

import sys
sys.path.insert(0, '../../../python')

import os

import logging.handlers
import logging

import ECC
import Google
import PDSChurch
import GoogleAuth
import copy
from abc import ABCMeta, abstractmethod

from datetime import date
from datetime import datetime
from datetime import timedelta

from oauth2client import tools

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment

from pprint import pprint
from pprint import pformat

now = datetime.now()

class MetaSheet:

    def __init__(self, ws, training_data):
        self.ws = ws
        self.values = training_data
        self.roster = list()

    def create_sheet(self, title, want_everything):
        ws = self.ws
        title_font = Font(color='FFFF00')
        title_fill = PatternFill(fgColor='0000FF', fill_type='solid')
        title_align = Alignment(horizontal='center')

        last_col = 'K'
    
        row = 1
        ws.merge_cells(f'A{row}:{last_col}{row}')
        cell = f'A{row}'
        ws[cell] = f'Training: {title}'
        ws[cell].fill = title_fill
        ws[cell].font = title_font

        row = row + 1
        ws.merge_cells(f'A{row}:{last_col}{row}')
        cell = f'A{row}'
        ws[cell] = f'Last updated: {now}'
        ws[cell].fill = title_fill
        ws[cell].font = title_font

        row = row + 1
        ws.merge_cells(f'A{row}:{last_col}{row}')
        cell = f'A{row}'
        ws[cell] = ''
        ws[cell].fill = title_fill
        ws[cell].font = title_font

        row = row + 1
        columns = [ (f'A{row}', 'Start Date'             ,   30),
                    (f'B{row}', 'End Date'               ,   30),
                    (f'C{row}', 'Member Name'            ,   30),
                    (f'D{row}', 'Email Address'          ,   30),
                    (f'E{row}', 'Phone Number'           ,   50),
                    (f'F{row}', 'Weekend?'               ,   50),
                    (f'G{row}', 'Weekday?'               ,   50),
                    (f'H{row}', 'Homebound?'             ,   50),
                    (f'I{row}', 'Notes'                  ,   50),]
        if want_everything:
            columns.append((f'J{row}', 'Stage of Certification', 50))
            columns.append((f'K{row}', 'Involved?'             , 50))

        for cell,value,width in columns:
            ws[cell] = value
            ws[cell].fill = title_fill
            ws[cell].font = title_font
            ws[cell].alignment = title_align
            ws.column_dimensions[cell[0]].width = width

        row = row + 1
        ws.freeze_panes = ws[f'A{row}']

        return row, ws

    def create_entry(self, ws, entry, row):
        ws[f'A{row}'] = entry['start_date']
        ws[f'B{row}'] = entry['end_date']
        ws[f'C{row}'] = entry['name']
        ws[f'D{row}'] = entry['email']
        ws[f'E{row}'] = entry['phone']
        ws[f'F{row}'] = entry['weekend']
        ws[f'G{row}'] = entry['weekday']
        ws[f'H{row}'] = entry['homebound']

class EverythingSheet(MetaSheet):

    def __init__(self, wb, training_data):
        ws = wb['Sheet']
        ws.title = 'Everything'
        MetaSheet.__init__(self, ws, training_data)
        
    def create_roster(self, title):
        (row, ws) = self.create_sheet(title, want_everything = True)
        for mid in sorted(self.values, reverse=True):
            for sd in sorted(self.values[mid]):
                for entry in self.values[mid][sd]:
                    self.create_entry(ws, entry, row)
                    ws[f'J{row}'] = entry['stage']
                    ws[f'K{row}'] = entry['involved']

                    row += 1
        return row

class SchedulableSheet(MetaSheet):

    def __init__(self, wb, training_data):
        ws = wb.create_sheet('Schedulable')
        MetaSheet.__init__(self, ws, training_data)
        
    def create_roster(self, title):
        (row, ws) = self.create_sheet(title, want_everything = False)
        for mid in sorted(self.values, reverse=True):
            most_recent = { 'end_date' : now.date() }
            for sd in sorted(self.values[mid]):
                for entry in self.values[mid][sd]:
                    if entry['end_date'] > most_recent['end_date']:
                        most_recent   = entry
            if most_recent['end_date'] != now.date():
                self.create_entry(ws, most_recent, row)
                row += 1
        return row

class NonSchedulableSheet(MetaSheet):

    def __init__(self, wb, training_data):
        ws = wb.create_sheet('NonSchedulable')
        MetaSheet.__init__(self, ws, training_data)
        
    def create_roster(self, title):
        (row, ws) = self.create_sheet(title, want_everything = False)
        for mid in sorted(self.values, reverse=True):
            most_recent = { 'end_date' : now.date() }
            for sd in sorted(self.values[mid]):
                for entry in self.values[mid][sd]:
                    if entry['end_date'] < now.date():
                        if (entry['end_date'] > most_recent['end_date']) or (most_recent['end_date'] == now.date()):
                            most_recent = entry
            if most_recent['end_date'] != now.date():
                self.create_entry(ws, most_recent, row)
                row += 1
        return row